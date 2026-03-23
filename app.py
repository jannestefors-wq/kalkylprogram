#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Kalkylprogram Bygg & Entreprenad – v2
"""
import streamlit as st
import pandas as pd
import json, datetime, copy, io

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    OPX_OK = True
except ImportError:
    OPX_OK = False

try:
    from reportlab.lib import colors as rl_colors
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import mm
    RL_OK = True
except ImportError:
    RL_OK = False

# ──────────────────────────────────────────────────────────────
st.set_page_config(page_title="Kalkylprogram – Bygg & Entreprenad",
                   page_icon="🏗", layout="wide",
                   initial_sidebar_state="expanded")

RADTYPER    = ["Material", "Arbete", "UE"]
ENHETER     = ["st","m","m²","m³","kg","ton","tim","pers","ls","rund"]
STATUS_LIST = ["Kalkyl","Offert","Pågående","Vunnen","Förlorad","Avslutad","Pausad"]
BYGGDEL_DEF = ["Badrum","Kök","Vägg","Etapp 1","Rivning","Snickeri",
               "Grund","Fasad","Tak","El","VVS","Övrigt"]
OMKKOSTNADER= ["Arbetsledning","Etablering","Administration",
               "Transporter","Garanti","Risk","Övrigt"]
BKI = {
    "Flerbostadshus":{"2018":175.2,"2019":180.1,"2020":183.5,"2021":192.4,
                      "2022":215.8,"2023":228.3,"2024":232.1,"2025":238.5,"2026":242.0},
    "Småhus":        {"2018":168.4,"2019":173.2,"2020":177.8,"2021":186.3,
                      "2022":208.9,"2023":221.7,"2024":225.4,"2025":231.0,"2026":234.5},
    "ROT/Ombyggnad": {"2018":171.3,"2019":176.5,"2020":180.2,"2021":189.1,
                      "2022":212.1,"2023":224.9,"2024":228.7,"2025":234.8,"2026":238.2},
}

KATEGORIER = ["Mark","Betong","Murning","Trä & stål","Isolering",
              "Skivor","Tak","Puts","Målning","Beläggning","Sakvaror",
              "El","VVS","Ventilation","Rivning","Inredning","Övrigt"]

DEFAULT_PRISBANK = [
    # Extraherat från BK 2025 (Byggmästarnas Kostnadskalkylator)
    {"kod":"BK0002","benamning":"Maskinschakt för källare och grunder Jordmån klass A","enhet":"m³","matpris":401,"arbpris":71,"leverantor":"BK 2025","kategori":"Mark"},
    {"kod":"BK0003","benamning":"Maskinschakt för yttre rörgravar etc Jordmån klass A genomsnitt ca","enhet":"m³","matpris":450,"arbpris":80,"leverantor":"BK 2025","kategori":"Mark"},
    {"kod":"BK0004","benamning":"Rörgravar, sulor, dräneringsledningar etc. Jordmån klass A, djup 0–1 m","enhet":"m³","matpris":450,"arbpris":80,"leverantor":"BK 2025","kategori":"Mark"},
    {"kod":"BK0012","benamning":"För schakt i tjälad jord och då tjälen överstiger 80 mm i tjocklek ökas priserna för handschaktning med ca 110–130 kr","enhet":"m³","matpris":204,"arbpris":36,"leverantor":"BK 2025","kategori":"Mark"},
    {"kod":"BK0013","benamning":"Fyllnad med schaktmaskin, befintliga massor jord klass A, varierande","enhet":"m³","matpris":150,"arbpris":26,"leverantor":"BK 2025","kategori":"Mark"},
    {"kod":"BK0014","benamning":"Dito, jord klass B, C, och D","enhet":"m³","matpris":194,"arbpris":34,"leverantor":"BK 2025","kategori":"Mark"},
    {"kod":"BK0015","benamning":"Dito, med anskaffade fyllnadsmassor, varierande beroende på materialets anskaffningspris","enhet":"m³","matpris":478,"arbpris":84,"leverantor":"BK 2025","kategori":"Mark"},
    {"kod":"BK0016","benamning":"Fyllnad för hand, jord klass A","enhet":"m³","matpris":500,"arbpris":88,"leverantor":"BK 2025","kategori":"Mark"},
    {"kod":"BK0017","benamning":"Grovplanering, befintliga massor, schakt- maskin inkl efterjustering ca","enhet":"m²","matpris":90,"arbpris":16,"leverantor":"BK 2025","kategori":"Mark"},
    {"kod":"BK0018","benamning":"Dito + utplanering av 100 mm befintlig matjord + gödsling + grässådd","enhet":"m²","matpris":252,"arbpris":45,"leverantor":"BK 2025","kategori":"Mark"},
    {"kod":"BK0019","benamning":"Dito, med inköpt och transporterad matjord, priset varierar med matjordskostnad","enhet":"m²","matpris":352,"arbpris":62,"leverantor":"BK 2025","kategori":"Mark"},
    {"kod":"BK0020","benamning":"Vägar och planer för gångtrafik + 35 mm bituminös beläggning","enhet":"m²","matpris":486,"arbpris":86,"leverantor":"BK 2025","kategori":"Mark"},
    {"kod":"BK0021","benamning":"Vägar och planer för körtrafik + 40 mm bituminös beläggning","enhet":"m²","matpris":594,"arbpris":105,"leverantor":"BK 2025","kategori":"Mark"},
    {"kod":"BK0024","benamning":"Asfaltbeläggningar Asfalt 60 Ab 8 t exkl underarbete ca","enhet":"m²","matpris":450,"arbpris":80,"leverantor":"BK 2025","kategori":"Beläggning"},
    {"kod":"BK0025","benamning":"Asfalt 80 Ab 12 t exkl underarbete ca","enhet":"m²","matpris":510,"arbpris":90,"leverantor":"BK 2025","kategori":"Beläggning"},
    {"kod":"BK0026","benamning":"Asfalt 100 Ab 12 t exkl underarbete ca","enhet":"m²","matpris":532,"arbpris":94,"leverantor":"BK 2025","kategori":"Beläggning"},
    {"kod":"BK0027","benamning":"Slitlager av grus och stenmjöl Grus 50 mm","enhet":"m²","matpris":90,"arbpris":16,"leverantor":"BK 2025","kategori":"Beläggning"},
    {"kod":"BK0028","benamning":"Stenmjöl 50 mm","enhet":"m²","matpris":82,"arbpris":14,"leverantor":"BK 2025","kategori":"Beläggning"},
    {"kod":"BK0029","benamning":"Gångbaneplattor av betong. Inkl sättsand Släta 350 x 350 x 50 mm","enhet":"m²","matpris":829,"arbpris":146,"leverantor":"BK 2025","kategori":"Betong"},
    {"kod":"BK0031","benamning":"Blästrade 350 x 350 x 50 mm","enhet":"m²","matpris":882,"arbpris":156,"leverantor":"BK 2025","kategori":"Beläggning"},
    {"kod":"BK0033","benamning":"Mönsterpräglade 350 x 350 x 50 mm","enhet":"m²","matpris":986,"arbpris":174,"leverantor":"BK 2025","kategori":"Beläggning"},
    {"kod":"BK0035","benamning":"Typ SF-sten 190 x 108 x 60 mm","enhet":"m²","matpris":1243,"arbpris":219,"leverantor":"BK 2025","kategori":"Beläggning"},
    {"kod":"BK0037","benamning":"Typ UNI-sten 225 x 112,5 x 60 mm","enhet":"m²","matpris":1346,"arbpris":238,"leverantor":"BK 2025","kategori":"Beläggning"},
    {"kod":"BK0039","benamning":"Rektangulär sten 210 x 105 x 60 mm","enhet":"m²","matpris":1346,"arbpris":238,"leverantor":"BK 2025","kategori":"Beläggning"},
    {"kod":"BK0041","benamning":"Priset varierar med mängd, läggningsmönster, frakter m m","enhet":"m²","matpris":2476,"arbpris":437,"leverantor":"BK 2025","kategori":"Mark"},
    {"kod":"BK0043","benamning":"Trall 90 x 27 mm","enhet":"m²","matpris":1188,"arbpris":210,"leverantor":"BK 2025","kategori":"Trä & stål"},
    {"kod":"BK0044","benamning":"Dimension 500 x 250 x 45 mm (frilagd ballast) m 732:– UE 700 x 350 x 65 mm","enhet":"m","matpris":599,"arbpris":106,"leverantor":"BK 2025","kategori":"Trä & stål"},
    {"kod":"BK0046","benamning":"Kantstöd av betong limmade eller spikade Dimension l 800 h 80","enhet":"m","matpris":518,"arbpris":92,"leverantor":"BK 2025","kategori":"Betong"},
    {"kod":"BK0050","benamning":"Kantstöd av råhuggen granit inkl sättsand Dimension RV1 b 150, h 300","enhet":"m","matpris":1554,"arbpris":274,"leverantor":"BK 2025","kategori":"Trä & stål"},
    {"kod":"BK0051","benamning":"RV2 b 120, h 300","enhet":"m","matpris":1504,"arbpris":265,"leverantor":"BK 2025","kategori":"Trä & stål"},
    {"kod":"BK0052","benamning":"RV4 b 100, h 300","enhet":"m","matpris":1450,"arbpris":256,"leverantor":"BK 2025","kategori":"Trä & stål"},
    {"kod":"BK0053","benamning":"Kantstöd av råhuggen granit satt i betong Dimension RV1 b 150, h 300","enhet":"m","matpris":1864,"arbpris":329,"leverantor":"BK 2025","kategori":"Betong"},
    {"kod":"BK0056","benamning":"Kantstöd av sågat virke impregnerad i klass A enligt SIS 05 61 10 Tryckimpr virke 50 x 100 oh","enhet":"m","matpris":176,"arbpris":31,"leverantor":"BK 2025","kategori":"Trä & stål"},
    {"kod":"BK0057","benamning":"Tryckimpr virke 50 x 125 oh","enhet":"m","matpris":176,"arbpris":31,"leverantor":"BK 2025","kategori":"Trä & stål"},
    {"kod":"BK0058","benamning":"Tryckimpr virke 38 x 125 oh","enhet":"m","matpris":150,"arbpris":26,"leverantor":"BK 2025","kategori":"Trä & stål"},
    {"kod":"BK0059","benamning":"Cykelställ med 5 platser för fastgjutning i mark","enhet":"st","matpris":5851,"arbpris":1033,"leverantor":"BK 2025","kategori":"Mark"},
    {"kod":"BK0060","benamning":"Ledningsbädd för dräneringsledning, b=400 fall B, tj= 150","enhet":"m","matpris":150,"arbpris":26,"leverantor":"BK 2025","kategori":"Mark"},
    {"kod":"BK0062","benamning":"PEH dräneringsrör ø 90","enhet":"m","matpris":150,"arbpris":26,"leverantor":"BK 2025","kategori":"Mark"},
    {"kod":"BK0066","benamning":"Ledningsbädd för dräneringsledning, fall B (0,15 m3 / lm), tj=150","enhet":"m²","matpris":154,"arbpris":27,"leverantor":"BK 2025","kategori":"Mark"},
    {"kod":"BK0068","benamning":"Dräneringslager av tvättad makadam tj<100","enhet":"m²","matpris":132,"arbpris":23,"leverantor":"BK 2025","kategori":"Mark"},
    {"kod":"BK0072","benamning":"Sprängning för källare, grunder, öppna schakt m m Avtäckning berg","enhet":"m²","matpris":185,"arbpris":33,"leverantor":"BK 2025","kategori":"Mark"},
    {"kod":"BK0073","benamning":"Bergs överyta h<1,0 m","enhet":"m²","matpris":262,"arbpris":46,"leverantor":"BK 2025","kategori":"Betong"},
    {"kod":"BK0074","benamning":"Bergs överyta h>1,0 m","enhet":"m²","matpris":194,"arbpris":34,"leverantor":"BK 2025","kategori":"Betong"},
    {"kod":"BK0076","benamning":"Tätning sprängbotten","enhet":"m²","matpris":262,"arbpris":46,"leverantor":"BK 2025","kategori":"Betong"},
    {"kod":"BK0077","benamning":"Djup intill 600 mm bottenbredd intill 600 mm","enhet":"m","matpris":1140,"arbpris":201,"leverantor":"BK 2025","kategori":"Betong"},
    {"kod":"BK0080","benamning":"Djup intill 3000 mm bottenbredd intill 600 mm","enhet":"m³","matpris":1710,"arbpris":302,"leverantor":"BK 2025","kategori":"Betong"},
    {"kod":"BK0085","benamning":"Jordsten större än 0,5 m3 men högst 1 m3/st, variabelt ca","enhet":"st","matpris":1346,"arbpris":238,"leverantor":"BK 2025","kategori":"Beläggning"},
    {"kod":"BK0086","benamning":"Transport av bergmassor Väglängd 0– 5,0 km","enhet":"m³","matpris":195,"arbpris":293,"leverantor":"BK 2025","kategori":"Mark"},
    {"kod":"BK0094","benamning":"Tillägg för komplett brandpostanordning, bygghöjd 1,7 m, inklusive brandpostledning (DN 100) och förankring BP med brandpostledning 2 m","enhet":"st","matpris":19506,"arbpris":3442,"leverantor":"BK 2025","kategori":"Mark"},
    {"kod":"BK0099","benamning":"Brunnar av betong Nedstigningsbrunn, bygghöjd 1,2 m","enhet":"st","matpris":14487,"arbpris":2557,"leverantor":"BK 2025","kategori":"Betong"},
    {"kod":"BK0101","benamning":"Tillsynsbrunn, bygghöjd ____1,2 m","enhet":"st","matpris":9528,"arbpris":1682,"leverantor":"BK 2025","kategori":"Mark"},
    {"kod":"BK0103","benamning":"Rensbrunn, bygghöjd ______1,2 m","enhet":"st","matpris":5447,"arbpris":961,"leverantor":"BK 2025","kategori":"Mark"},
    {"kod":"BK0105","benamning":"Dagvattenbrunn, inkl gallerbetäckning med vattenlås och sandfång","enhet":"st","matpris":5253,"arbpris":927,"leverantor":"BK 2025","kategori":"VVS"},
    {"kod":"BK0114","benamning":"Tillägg för brunn av plast, komplett, rakt genomlopp utan sidoanslutning Tillsynsbrunn, bygghöjd 1,2 m","enhet":"st","matpris":12638,"arbpris":2230,"leverantor":"BK 2025","kategori":"Mark"},
    {"kod":"BK0116","benamning":"Rensbrunn, bygghöjd 1,2 m","enhet":"st","matpris":13030,"arbpris":2299,"leverantor":"BK 2025","kategori":"Mark"},
    {"kod":"BK0118","benamning":"Dagvattenbrunn, djup till inkommande lednings vattengång 1,2 m med vattenlås och sandfång","enhet":"st","matpris":9141,"arbpris":1613,"leverantor":"BK 2025","kategori":"Mark"},
    {"kod":"BK0133","benamning":"Formar till grundplattor, kantformar, avsättningar etc, lågformar intill högst 300 mm höjd, en sida form","enhet":"m","matpris":160,"arbpris":139,"leverantor":"BK 2025","kategori":"Beläggning"},
    {"kod":"BK0134","benamning":"Kantbalksform till bl a källarlösa grundtyper med utbredda plattor på mark, formhöjd > 400 mm","enhet":"m²","matpris":188,"arbpris":144,"leverantor":"BK 2025","kategori":"Mark"},
    {"kod":"BK0135","benamning":"Formar till väggar, motgjuten yta 25 mm råplan och ohyvlad","enhet":"m²","matpris":339,"arbpris":396,"leverantor":"BK 2025","kategori":"Mark"},
    {"kod":"BK0138","benamning":"Plywoodform för putsfria ytor","enhet":"m²","matpris":471,"arbpris":396,"leverantor":"BK 2025","kategori":"Puts"},
    {"kod":"BK0139","benamning":"Tillägg för väggtjocklekar under 170 mm","enhet":"m²","matpris":19,"arbpris":50,"leverantor":"BK 2025","kategori":"Mark"},
    {"kod":"BK0143","benamning":"Plywoodform för putsfria takytor","enhet":"m²","matpris":276,"arbpris":272,"leverantor":"BK 2025","kategori":"Tak"},
    {"kod":"BK0144","benamning":"Formar till balkongplattor, skärmtak m m, mindre ytor, putsfria ytor ca","enhet":"m²","matpris":310,"arbpris":421,"leverantor":"BK 2025","kategori":"Tak"},
    {"kod":"BK0145","benamning":"Formar till pelare, balkar etc Balkformar 23 mm råplan","enhet":"m²","matpris":378,"arbpris":594,"leverantor":"BK 2025","kategori":"El"},
    {"kod":"BK0147","benamning":"Pelarformar 23 mm råplan","enhet":"m²","matpris":404,"arbpris":569,"leverantor":"BK 2025","kategori":"El"},
    {"kod":"BK0149","benamning":"Kvalitet B 500 BT ø 6 mm","enhet":"kg","matpris":30,"arbpris":20,"leverantor":"BK 2025","kategori":"El"},
    {"kod":"BK0152","benamning":"Nät typ 5150","enhet":"m²","matpris":48,"arbpris":20,"leverantor":"BK 2025","kategori":"El"},
    {"kod":"BK0153","benamning":"Fogband Fogband av PVC FN bredd 150 mm","enhet":"m","matpris":245,"arbpris":124,"leverantor":"BK 2025","kategori":"El"},
    {"kod":"BK0157","benamning":"Enbart avjämning med sloda och exkl armering. Tjocklek 50 mm","enhet":"m²","matpris":124,"arbpris":50,"leverantor":"BK 2025","kategori":"Betong"},
    {"kod":"BK0164","benamning":"Helgjutna betonggolv Betong C 20/25, inkl stålslipad färdig golvyta, exkl armering. Tjocklek 70 mm","enhet":"m²","matpris":183,"arbpris":99,"leverantor":"BK 2025","kategori":"Betong"},
    {"kod":"BK0169","benamning":"Tjocklek 120 mm","enhet":"m²","matpris":400,"arbpris":70,"leverantor":"BK 2025","kategori":"Betong"},
    {"kod":"BK0170","benamning":"Tjocklek 40 mm","enhet":"m²","matpris":142,"arbpris":148,"leverantor":"BK 2025","kategori":"Betong"},
    {"kod":"BK0175","benamning":"För ev fogindelning av större golvytor, exempelvis industrigolv m m tillägg per m fog inkl krympbar fogmassa, läkt, dymlingar m m Helfog","enhet":"m","matpris":240,"arbpris":248,"leverantor":"BK 2025","kategori":"Beläggning"},
    {"kod":"BK0176","benamning":"Spontad arbetsfog","enhet":"m","matpris":141,"arbpris":148,"leverantor":"BK 2025","kategori":"Betong"},
    {"kod":"BK0177","benamning":"Industrigolv Flintkotegolv, tjocklek 10–12 mm.","enhet":"m²","matpris":241,"arbpris":0,"leverantor":"BK 2025","kategori":"Beläggning"},
    {"kod":"BK0178","benamning":"Smet- och hålkälslister mot väggar","enhet":"m","matpris":162,"arbpris":0,"leverantor":"BK 2025","kategori":"Mark"},
    {"kod":"BK0179","benamning":"Platsgjuten rak trappa av betong b=1000 mm, 15 steg","enhet":"st","matpris":22972,"arbpris":23676,"leverantor":"BK 2025","kategori":"Betong"},
    {"kod":"BK0183","benamning":"Densitet 500 kg/m3 Tjocklek 50 mm","enhet":"m²","matpris":564,"arbpris":297,"leverantor":"BK 2025","kategori":"Murning"},
    {"kod":"BK0186","benamning":"Densitet 400–450 kg/m3 Tjocklek 150 mm","enhet":"m²","matpris":1064,"arbpris":322,"leverantor":"BK 2025","kategori":"Betong"},
    {"kod":"BK0191","benamning":"Till fönster- och dörröppningar Väggtjocklek 150 mm","enhet":"m","matpris":811,"arbpris":198,"leverantor":"BK 2025","kategori":"Mark"},
    {"kod":"BK0196","benamning":"Densitet 450 kg/m3 Tjocklek 150 mm","enhet":"m²","matpris":1014,"arbpris":297,"leverantor":"BK 2025","kategori":"Betong"},
    {"kod":"BK0201","benamning":"Isolerblock 290 x 190 x 590 mm","enhet":"m²","matpris":1641,"arbpris":257,"leverantor":"BK 2025","kategori":"Betong"},
    {"kod":"BK0202","benamning":"Lättklinkerblock 70 x 190 x 590 mm","enhet":"m²","matpris":573,"arbpris":297,"leverantor":"BK 2025","kategori":"Beläggning"},
    {"kod":"BK0209","benamning":"Håltegel, 62 mm sten 1/2 -stens vägg, 52 sten per m2","enhet":"m²","matpris":1212,"arbpris":371,"leverantor":"BK 2025","kategori":"Mark"},
    {"kod":"BK0211","benamning":"Massivt tegel 62 mm sten (ej normerat) 1/2 -stens vägg, 52 sten per m2","enhet":"m²","matpris":1712,"arbpris":371,"leverantor":"BK 2025","kategori":"Mark"},
    {"kod":"BK0213","benamning":"Håltegel 1/2 -stens mur, 55 sten per m2","enhet":"m²","matpris":2050,"arbpris":544,"leverantor":"BK 2025","kategori":"Murning"},
    {"kod":"BK0215","benamning":"Ett flertal olika fabrikat och typer förekommer. pris per m2","enhet":"m²","matpris":6033,"arbpris":1238,"leverantor":"BK 2025","kategori":"Murning"},
    {"kod":"BK0216","benamning":"Bjälklags- och yttertakelement Förtillverkade kassettbjälklag TT40","enhet":"m²","matpris":1316,"arbpris":232,"leverantor":"BK 2025","kategori":"Trä & stål"},
    {"kod":"BK0217","benamning":"Förtillverkade hålbjälklag HD/F 265","enhet":"m²","matpris":1200,"arbpris":212,"leverantor":"BK 2025","kategori":"Betong"},
    {"kod":"BK0218","benamning":"Förtillverkade Plattbärlag t=70-80","enhet":"m²","matpris":554,"arbpris":98,"leverantor":"BK 2025","kategori":"Betong"},
    {"kod":"BK0219","benamning":"Rak trappa, hellopp h=3000, b=900","enhet":"st","matpris":65192,"arbpris":1980,"leverantor":"BK 2025","kategori":"El"},
    {"kod":"BK0220","benamning":"Dito svängd trappa","enhet":"st","matpris":71550,"arbpris":2475,"leverantor":"BK 2025","kategori":"Betong"},
    {"kod":"BK0221","benamning":"Rak trappa, halvlopp h=1500, b=900","enhet":"st","matpris":38954,"arbpris":1485,"leverantor":"BK 2025","kategori":"Betong"},
    {"kod":"BK0223","benamning":"Spindelburna vinkelblocksteg, svängd trappa, per steg","enhet":"st","matpris":4195,"arbpris":371,"leverantor":"BK 2025","kategori":"El"},
    {"kod":"BK0224","benamning":"Densitet 400–500 kg/m3 Tjocklek 100 mm","enhet":"m²","matpris":583,"arbpris":148,"leverantor":"BK 2025","kategori":"Betong"},
    {"kod":"BK0230","benamning":"Bjälklagselement till styckebyggda småhus typ BE 2,8 Tjocklek 250 mm","enhet":"m²","matpris":1568,"arbpris":40,"leverantor":"BK 2025","kategori":"Trä & stål"},
    {"kod":"BK0231","benamning":"Typ TE 450 3,2 Tjocklek 250 mm","enhet":"m²","matpris":1324,"arbpris":74,"leverantor":"BK 2025","kategori":"Betong"},
    {"kod":"BK0233","benamning":"Bygghöjd = 6M. Kvalitetsgrupp 550 Tjocklek 250 mm","enhet":"m²","matpris":1317,"arbpris":119,"leverantor":"BK 2025","kategori":"Betong"},
    {"kod":"BK0235","benamning":"Bjälklagselement av KL-trä KL-träskiva i bjälklag L100-5S, t=100 mm","enhet":"m²","matpris":1585,"arbpris":50,"leverantor":"BK 2025","kategori":"Trä & stål"},
    {"kod":"BK0236","benamning":"KL-träskiva i bjälklag L150-5S, t=150 mm","enhet":"m²","matpris":1930,"arbpris":50,"leverantor":"BK 2025","kategori":"Trä & stål"},
    {"kod":"BK0237","benamning":"KL-träskiva i bjälklag L200-5S, t=200 mm","enhet":"m²","matpris":2219,"arbpris":50,"leverantor":"BK 2025","kategori":"Trä & stål"},
    {"kod":"BK0238","benamning":"KL-träskiva i bjälklag L240-7S, t=240 mm","enhet":"m²","matpris":2980,"arbpris":50,"leverantor":"BK 2025","kategori":"Trä & stål"},
    {"kod":"BK0239","benamning":"KL-träskiva i bjälklag L300-7S, t=300 mm","enhet":"m²","matpris":3308,"arbpris":50,"leverantor":"BK 2025","kategori":"Trä & stål"},
    {"kod":"BK0240","benamning":"Väggelement av KL-trä KL-träskiva i vägg L100-5S, t=100 mm","enhet":"m²","matpris":1585,"arbpris":50,"leverantor":"BK 2025","kategori":"Mark"},
    {"kod":"BK0241","benamning":"KL-träskiva i vägg L150-5S, t=150 mm","enhet":"m²","matpris":1930,"arbpris":50,"leverantor":"BK 2025","kategori":"Mark"},
    {"kod":"BK0242","benamning":"KL-träskiva i vägg L200-5S, t=200 mm","enhet":"m²","matpris":2219,"arbpris":50,"leverantor":"BK 2025","kategori":"Mark"},
    {"kod":"BK0243","benamning":"KL-träskiva i vägg L240-7S, t=240 mm","enhet":"m²","matpris":2980,"arbpris":50,"leverantor":"BK 2025","kategori":"Mark"},
    {"kod":"BK0244","benamning":"KL-träskiva i vägg L300-7S, t=300 mm","enhet":"m²","matpris":3308,"arbpris":50,"leverantor":"BK 2025","kategori":"Mark"},
    {"kod":"BK0245","benamning":"Lättbalk Z-profil 100 x 1,2 mm","enhet":"m","matpris":136,"arbpris":30,"leverantor":"BK 2025","kategori":"Trä & stål"},
    {"kod":"BK0251","benamning":"Regelstomme c 600 (2,0 m/m2) R 45","enhet":"m²","matpris":96,"arbpris":84,"leverantor":"BK 2025","kategori":"El"},
    {"kod":"BK0256","benamning":"Regelstomme c 600 + c 600 (4,0 m/m2) R 45","enhet":"m²","matpris":192,"arbpris":168,"leverantor":"BK 2025","kategori":"El"},
    {"kod":"BK0261","benamning":"Golv-takskenor SK 70","enhet":"m","matpris":64,"arbpris":50,"leverantor":"BK 2025","kategori":"Tak"},
    {"kod":"BK0268","benamning":"Stålbalkar och övrigt profilstål, riktpriser, UE IPE-profil","enhet":"kg","matpris":49,"arbpris":9,"leverantor":"BK 2025","kategori":"Trä & stål"},
    {"kod":"BK0271","benamning":"Smäckrare profiljärn och plattstål","enhet":"kg","matpris":94,"arbpris":16,"leverantor":"BK 2025","kategori":"Trä & stål"},
    {"kod":"BK0272","benamning":"Diverse färdiga stomkonstruktioner av resp balkar och profilstål, varierande priser i förhållande till arbetsomfattning, totalvikt m m. För ev varmförzinkning av gängse handelsstål tillkommer ett varierande pris av 12–20:– per kg","enhet":"kg","matpris":13,"arbpris":2,"leverantor":"BK 2025","kategori":"Trä & stål"},
    {"kod":"BK0273","benamning":"Kvalitet Ö-virke, råplan 34 x 95 11/2 x 4”","enhet":"m","matpris":59,"arbpris":59,"leverantor":"BK 2025","kategori":"Trä & stål"},
    {"kod":"BK0279","benamning":"Tryckimpregnerat virke, klass A, plh 34 x 95 11/2 x 4”","enhet":"m","matpris":61,"arbpris":59,"leverantor":"BK 2025","kategori":"Trä & stål"},
    {"kod":"BK0285","benamning":"Trästomme c ca 600 mm (3,0 m/m2) 45 x 70 2 x 3”","enhet":"m²","matpris":134,"arbpris":94,"leverantor":"BK 2025","kategori":"Trä & stål"},
    {"kod":"BK0291","benamning":"Fabrikat Wood Tube 70/70 (450) (2,4 m/m2) 45 x 70","enhet":"m²","matpris":102,"arbpris":99,"leverantor":"BK 2025","kategori":"Trä & stål"},
    {"kod":"BK0292","benamning":"Kvalitet V-sort 23 x 95","enhet":"m²","matpris":271,"arbpris":193,"leverantor":"BK 2025","kategori":"Trä & stål"},
    {"kod":"BK0294","benamning":"Kvalitet V-sort inkl spikreglar, kortlingar etc c ca 600 mm (3,0 m/m2) 45 x 70 mm virke, 2 x 3”","enhet":"m²","matpris":166,"arbpris":119,"leverantor":"BK 2025","kategori":"Trä & stål"},
    {"kod":"BK0300","benamning":"Bjälkar inlagda c 600 mm (2 m/m2) Exkl ev blindbottnar, läkt mm och ev krysskolvning 45 x 145 mm virke, 2 x 6”","enhet":"m²","matpris":177,"arbpris":79,"leverantor":"BK 2025","kategori":"Trä & stål"},
    {"kod":"BK0304","benamning":"Krysskolvning med 45 x 95 / kryss","enhet":"st","matpris":87,"arbpris":74,"leverantor":"BK 2025","kategori":"Trä & stål"},
    {"kod":"BK0305","benamning":"Plywood P30 tj=7 mm med 25 x 125 mm tryckimpr blindbottenläkt","enhet":"m²","matpris":252,"arbpris":134,"leverantor":"BK 2025","kategori":"Trä & stål"},
    {"kod":"BK0306","benamning":"Dito tj=9 mm","enhet":"m²","matpris":313,"arbpris":148,"leverantor":"BK 2025","kategori":"Trä & stål"},
    {"kod":"BK0307","benamning":"Kvalitet Ö-virke 45 x 45 2 x 2”","enhet":"m²","matpris":311,"arbpris":94,"leverantor":"BK 2025","kategori":"Trä & stål"},
    {"kod":"BK0311","benamning":"Virke C24 / K24 45 x 120 mm sparrar, 2 x 5”","enhet":"m²","matpris":160,"arbpris":124,"leverantor":"BK 2025","kategori":"Trä & stål"},
    {"kod":"BK0315","benamning":"Kvalitet V sort 17 mm råspontad panel","enhet":"m²","matpris":185,"arbpris":89,"leverantor":"BK 2025","kategori":"El"},
    {"kod":"BK0319","benamning":"Dubbelläktning för tegel, 25 x 25 mm c 600 ströläkt + 25 x 38 mm c 375 bärläkt","enhet":"m²","matpris":55,"arbpris":35,"leverantor":"BK 2025","kategori":"Murning"},
    {"kod":"BK0320","benamning":"Kvalitet V-sort 22 x 95 gles panel c 300","enhet":"m²","matpris":101,"arbpris":74,"leverantor":"BK 2025","kategori":"El"},
    {"kod":"BK0324","benamning":"Furumaterial III Lockläktpanel 22x170 + 16x45","enhet":"m²","matpris":455,"arbpris":332,"leverantor":"BK 2025","kategori":"El"},
    {"kod":"BK0325","benamning":"Dubbelfasspont 22x145","enhet":"m²","matpris":383,"arbpris":193,"leverantor":"BK 2025","kategori":"El"},
    {"kod":"BK0327","benamning":"III Sort hyvlat 15 mm granpanel 3/4”","enhet":"m²","matpris":407,"arbpris":218,"leverantor":"BK 2025","kategori":"El"},
    {"kod":"BK0330","benamning":"För uppsättning i tak tilläggs","enhet":"m²","matpris":19,"arbpris":50,"leverantor":"BK 2025","kategori":"Tak"},
    {"kod":"BK0331","benamning":"Krossning från 0–16 mm (varierande priser) Lättklinker levererad i storsäck 1750 l","enhet":"m³","matpris":3672,"arbpris":198,"leverantor":"BK 2025","kategori":"Beläggning"},
    {"kod":"BK0332","benamning":"Längd x bredd ca 1200x600 mm Tjocklek 50 mm","enhet":"m²","matpris":302,"arbpris":30,"leverantor":"BK 2025","kategori":"Mark"},
    {"kod":"BK0335","benamning":"Tjocklek 50 mm","enhet":"m²","matpris":107,"arbpris":30,"leverantor":"BK 2025","kategori":"Mark"},
    {"kod":"BK0339","benamning":"Längd x höjd 2000 x 460 mm Tjocklek 75 mm","enhet":"m","matpris":363,"arbpris":148,"leverantor":"BK 2025","kategori":"Mark"},
    {"kod":"BK0343","benamning":"Tjocklek 25 mm","enhet":"m²","matpris":253,"arbpris":74,"leverantor":"BK 2025","kategori":"Mark"},
    {"kod":"BK0348","benamning":"Syllisolering av mineralull Bredd 100 mm","enhet":"m","matpris":21,"arbpris":15,"leverantor":"BK 2025","kategori":"Isolering"},
    {"kod":"BK0353","benamning":"Regelskiva, KL 0,037 Tjocklek 45 mm","enhet":"m²","matpris":54,"arbpris":30,"leverantor":"BK 2025","kategori":"El"},
    {"kod":"BK0361","benamning":"Bredd 455 mm Tjocklek 45 mm","enhet":"m²","matpris":79,"arbpris":30,"leverantor":"BK 2025","kategori":"Mark"},
    {"kod":"BK0368","benamning":"Längd x bredd 1200 x 600 mm Tjocklek 80 mm","enhet":"m²","matpris":165,"arbpris":30,"leverantor":"BK 2025","kategori":"Mark"},
    {"kod":"BK0372","benamning":"Fasadskiva 33 FS Redair Tjocklek 100 mm","enhet":"m²","matpris":384,"arbpris":50,"leverantor":"BK 2025","kategori":"Mark"},
    {"kod":"BK0376","benamning":"Längd x bredd ca 2700 x 1200 mm Tjocklek 50 mm","enhet":"m²","matpris":240,"arbpris":64,"leverantor":"BK 2025","kategori":"Mark"},
    {"kod":"BK0378","benamning":"Längd x bredd 2700 x 1200 mm Tjocklek 15 mm","enhet":"m²","matpris":171,"arbpris":50,"leverantor":"BK 2025","kategori":"Mark"},
    {"kod":"BK0380","benamning":"Längd x bredd 1200 x 600 mm Tjocklek 30 mm","enhet":"m²","matpris":183,"arbpris":40,"leverantor":"BK 2025","kategori":"Mark"},
    {"kod":"BK0389","benamning":"Bjälklagsskiva 36, stenull Tjocklek 70 mm","enhet":"m²","matpris":100,"arbpris":35,"leverantor":"BK 2025","kategori":"Beläggning"},
    {"kod":"BK0395","benamning":"Bjälklagsskiva 39, stenull Tjocklek 45 mm","enhet":"m²","matpris":75,"arbpris":35,"leverantor":"BK 2025","kategori":"Beläggning"},
    {"kod":"BK0400","benamning":"Längd x bredd ca 1200 x 1160 mm Tjocklek 145 mm","enhet":"m²","matpris":243,"arbpris":50,"leverantor":"BK 2025","kategori":"Tak"},
    {"kod":"BK0404","benamning":"Längd x bredd 1200 x 600 mm Tjocklek 95 mm","enhet":"m²","matpris":78,"arbpris":35,"leverantor":"BK 2025","kategori":"Tak"},
    {"kod":"BK0407","benamning":"Byggmatta med vindskydd KL 0,037 Tjocklek 50 mm","enhet":"m²","matpris":105,"arbpris":20,"leverantor":"BK 2025","kategori":"Tak"},
    {"kod":"BK0413","benamning":"Tjocklek 80 mm","enhet":"m²","matpris":505,"arbpris":59,"leverantor":"BK 2025","kategori":"Tak"},
    {"kod":"BK0418","benamning":"PIR-isolering (fast polyuretanskum) Tjocklek 30 mm","enhet":"m²","matpris":280,"arbpris":45,"leverantor":"BK 2025","kategori":"Isolering"},
    {"kod":"BK0432","benamning":"Underlagstäckning av byggpapp på yttertak Underlagstakpapp UT typ 111","enhet":"m²","matpris":63,"arbpris":30,"leverantor":"BK 2025","kategori":"Tak"},
    {"kod":"BK0433","benamning":"Vattenavledande skikt av byggpapp på yttertak Underlagstakpapp VU typ YAP 2200","enhet":"m²","matpris":84,"arbpris":30,"leverantor":"BK 2025","kategori":"Tak"},
    {"kod":"BK0434","benamning":"Byggfolie ångbroms t=0,2 på tak","enhet":"m²","matpris":45,"arbpris":15,"leverantor":"BK 2025","kategori":"Tak"},
    {"kod":"BK0435","benamning":"Mindre ytor Tätskikt TI typ JSE.1321","enhet":"m²","matpris":241,"arbpris":42,"leverantor":"BK 2025","kategori":"Tak"},
    {"kod":"BK0438","benamning":"Tätskikt i terrassbjälklag Tätskikt TT typ JSE.1421","enhet":"m²","matpris":241,"arbpris":42,"leverantor":"BK 2025","kategori":"Tak"},
    {"kod":"BK0440","benamning":"Tätskikt på yttertak Tätskiktsklass TY typ JSE.1513","enhet":"m²","matpris":192,"arbpris":34,"leverantor":"BK 2025","kategori":"Tak"},
    {"kod":"BK0445","benamning":"Tätskiktsklass duk, typ JSE 4524","enhet":"m²","matpris":221,"arbpris":39,"leverantor":"BK 2025","kategori":"Tak"},
    {"kod":"BK0446","benamning":"Industritäckning Tvålagstäckning (typ TY JSE.1523) inkl värmeisolering Tvålag + 50 mm värmeisolering","enhet":"m²","matpris":373,"arbpris":66,"leverantor":"BK 2025","kategori":"Isolering"},
    {"kod":"BK0450","benamning":"Grundisoleringspapp Impregnerad och asfaltytbelagd, sandad på båda sidor YEP 2500 bredd 125 mm","enhet":"m","matpris":40,"arbpris":20,"leverantor":"BK 2025","kategori":"Isolering"},
    {"kod":"BK0454","benamning":"Syllisolering S-list b= 120","enhet":"m","matpris":24,"arbpris":15,"leverantor":"BK 2025","kategori":"Isolering"},
    {"kod":"BK0455","benamning":"Fuktskydd av plastfilm (plastfolie) Tjocklek 0,20 mm på vägg","enhet":"m²","matpris":29,"arbpris":20,"leverantor":"BK 2025","kategori":"Mark"},
    {"kod":"BK0457","benamning":"Luftspaltbildande plastmatta Platonmatta","enhet":"m²","matpris":133,"arbpris":50,"leverantor":"BK 2025","kategori":"Ventilation"},
    {"kod":"BK0458","benamning":"Vegetationsskikt på tätskikt, Moss-sedum t=30","enhet":"m²","matpris":295,"arbpris":52,"leverantor":"BK 2025","kategori":"Tak"},
    {"kod":"BK0459","benamning":"Dubbelfalsade taktäckningar Bandtäckning Förz plåt","enhet":"m²","matpris":989,"arbpris":175,"leverantor":"BK 2025","kategori":"Tak"},
    {"kod":"BK0460","benamning":"Förz, lackad","enhet":"m²","matpris":1039,"arbpris":183,"leverantor":"BK 2025","kategori":"Tak"},
    {"kod":"BK0462","benamning":"Skivtäckning Förz plåt","enhet":"m²","matpris":1336,"arbpris":236,"leverantor":"BK 2025","kategori":"Tak"},
    {"kod":"BK0465","benamning":"Krönlist, dubbel fals 3 delar AMA, Fig JT/69 Krönbeslag, klb=400, metall. stålplåt","enhet":"m²","matpris":1306,"arbpris":230,"leverantor":"BK 2025","kategori":"Trä & stål"},
    {"kod":"BK0466","benamning":"Bredd 250 mm Förz plåt","enhet":"m","matpris":173,"arbpris":31,"leverantor":"BK 2025","kategori":"Tak"},
    {"kod":"BK0467","benamning":"JT-.42 Vinkelrännor av plåt vid taktäckning av överläggsplattor e d Ränndalsplåt/vinkelränna AMA, Fig JT/82, Fig JT/83 Vinkelränna 1100 mm","enhet":"m","matpris":813,"arbpris":143,"leverantor":"BK 2025","kategori":"Tak"},
    {"kod":"BK0468","benamning":"Klippbredd max 450 mm Ståndskiva, metall. stålplåt","enhet":"m","matpris":241,"arbpris":43,"leverantor":"BK 2025","kategori":"Trä & stål"},
    {"kod":"BK0469","benamning":"Ståndskiva, kopparplåt","enhet":"m","matpris":779,"arbpris":138,"leverantor":"BK 2025","kategori":"Tak"},
    {"kod":"BK0470","benamning":"Nock- och vindskivsbeslag, krönbeslag, ståndskivor m m Gavelbeslag, klb=400, metall. stålplåt","enhet":"m","matpris":658,"arbpris":116,"leverantor":"BK 2025","kategori":"Trä & stål"},
    {"kod":"BK0471","benamning":"Stosar Stos, stålplåt","enhet":"st","matpris":2224,"arbpris":392,"leverantor":"BK 2025","kategori":"Trä & stål"},
    {"kod":"BK0472","benamning":"Stos, koppar","enhet":"st","matpris":3427,"arbpris":605,"leverantor":"BK 2025","kategori":"Tak"},
    {"kod":"BK0473","benamning":"Helbeslagning av ventilationsskorstenar och liknande, mått intill 2500 mm rundmått och höjd intill 800 mm, taklutning ≤ 20° vid bandtäckning, stålplåt","enhet":"st","matpris":3060,"arbpris":540,"leverantor":"BK 2025","kategori":"Trä & stål"},
    {"kod":"BK0475","benamning":"Fullt färdigt arbete exkl eventuella fotplåtar. Rännor förz plåt bredd 100 mm","enhet":"m","matpris":440,"arbpris":78,"leverantor":"BK 2025","kategori":"Tak"},
    {"kod":"BK0478","benamning":"Rännor förz, lackad bredd 100 mm","enhet":"m","matpris":323,"arbpris":57,"leverantor":"BK 2025","kategori":"Tak"},
    {"kod":"BK0481","benamning":"Rännor koppar- plåt bredd 100 mm","enhet":"m","matpris":1193,"arbpris":211,"leverantor":"BK 2025","kategori":"Tak"},
    {"kod":"BK0484","benamning":"Stuprör Komplett arbete inkl rör, svep, väggfästen, mellanstycken, vinklar, utkastare etc. Förz plåt ø 87 mm","enhet":"m","matpris":785,"arbpris":139,"leverantor":"BK 2025","kategori":"Mark"},
    {"kod":"BK0488","benamning":"Kopparplåt ø 87 mm","enhet":"m","matpris":2734,"arbpris":482,"leverantor":"BK 2025","kategori":"Tak"},
    {"kod":"BK0496","benamning":"Cirkaåtgång 2-kup 10,7 och 1-kup 13 st/m2 1-kupiga","enhet":"m²","matpris":702,"arbpris":109,"leverantor":"BK 2025","kategori":"Murning"},
    {"kod":"BK0498","benamning":"Åtgång ca 9/m2. Vikt 4,1 kg/st. 1-kupiga, standard, svart","enhet":"m²","matpris":378,"arbpris":89,"leverantor":"BK 2025","kategori":"Betong"},
    {"kod":"BK0499","benamning":"Dito, ytbehandlad, svart","enhet":"m²","matpris":349,"arbpris":89,"leverantor":"BK 2025","kategori":"Betong"},
    {"kod":"BK0502","benamning":"Nockpannor, standard, svart","enhet":"m","matpris":479,"arbpris":25,"leverantor":"BK 2025","kategori":"Betong"},
    {"kod":"BK0505","benamning":"Gavelpannor, stardard, röd","enhet":"m","matpris":297,"arbpris":25,"leverantor":"BK 2025","kategori":"El"},
    {"kod":"BK0507","benamning":"Tg takpannor sned begränsning","enhet":"m","matpris":109,"arbpris":272,"leverantor":"BK 2025","kategori":"Tak"},
    {"kod":"BK0508","benamning":"Snörasskydd 3 rör, galv, infästning för takpannor","enhet":"m","matpris":1078,"arbpris":134,"leverantor":"BK 2025","kategori":"Tak"},
    {"kod":"BK0509","benamning":"JV-.21 Taktäckningar av överläggsplattor e d av profilerad plåt Plåtprofil 20 t=0,60 på yttertak, HC 25","enhet":"m²","matpris":298,"arbpris":45,"leverantor":"BK 2025","kategori":"Tak"},
    {"kod":"BK0510","benamning":"Plåtprofil 45 t=0,60 på yttertak, HC 25","enhet":"m²","matpris":218,"arbpris":45,"leverantor":"BK 2025","kategori":"Tak"},
    {"kod":"BK0511","benamning":"Plåtprofil 70 t=0,65 bärande på yttertak, HC 25","enhet":"m²","matpris":270,"arbpris":54,"leverantor":"BK 2025","kategori":"Tak"},
    {"kod":"BK0512","benamning":"LHP 115 t=0,9 bärande i yttertak, vfz","enhet":"m²","matpris":262,"arbpris":54,"leverantor":"BK 2025","kategori":"Tak"},
    {"kod":"BK0513","benamning":"Fabrikslackerad profilerad plåt Tegelprofilerad plåt, t=0,5 HC 25","enhet":"m²","matpris":279,"arbpris":45,"leverantor":"BK 2025","kategori":"Murning"},
    {"kod":"BK0514","benamning":"Lackad stålplåt på träreglar Plåtprofil 20 t=0,60 på vägg, HC 25","enhet":"m²","matpris":316,"arbpris":89,"leverantor":"BK 2025","kategori":"Mark"},
    {"kod":"BK0515","benamning":"Plåtprofil 40 t=0,60 på vägg, HC 25","enhet":"m²","matpris":221,"arbpris":89,"leverantor":"BK 2025","kategori":"Mark"},
    {"kod":"BK0516","benamning":"Plåtprofil 70 t=0,65 på vägg HC 25","enhet":"m²","matpris":284,"arbpris":89,"leverantor":"BK 2025","kategori":"Mark"},
    {"kod":"BK0518","benamning":"Gipsbaserad kompositskiva vindskydd t=9,5 b=900","enhet":"m²","matpris":253,"arbpris":64,"leverantor":"BK 2025","kategori":"Skivor"},
    {"kod":"BK0519","benamning":"Normal, typ GN, tjocklek Bredd 900 mm 13 mm","enhet":"m²","matpris":106,"arbpris":109,"leverantor":"BK 2025","kategori":"Betong"},
    {"kod":"BK0523","benamning":"Uppsättning i tak Gipsplank t=13, b=600","enhet":"m","matpris":118,"arbpris":99,"leverantor":"BK 2025","kategori":"Tak"},
    {"kod":"BK0524","benamning":"Gipskortplank t=13, b=600","enhet":"m","matpris":181,"arbpris":99,"leverantor":"BK 2025","kategori":"Skivor"},
    {"kod":"BK0525","benamning":"Hörnskydd till gipsskivor Hörnskydd HS","enhet":"m","matpris":47,"arbpris":35,"leverantor":"BK 2025","kategori":"Skivor"},
    {"kod":"BK0530","benamning":"Plywood på vägg inomhus Plywood t=12 på vägg","enhet":"m²","matpris":242,"arbpris":99,"leverantor":"BK 2025","kategori":"Mark"},
    {"kod":"BK0531","benamning":"Plywood t=15 på vägg","enhet":"m²","matpris":257,"arbpris":99,"leverantor":"BK 2025","kategori":"Mark"},
    {"kod":"BK0532","benamning":"Kvalitet BB/x Tjocklek 7 mm","enhet":"m²","matpris":340,"arbpris":84,"leverantor":"BK 2025","kategori":"Tak"},
    {"kod":"BK0536","benamning":"Golvskivor Plywood t=15 på golv P30","enhet":"m²","matpris":333,"arbpris":89,"leverantor":"BK 2025","kategori":"Skivor"},
    {"kod":"BK0537","benamning":"Plywood t=18 på golv P30","enhet":"m²","matpris":361,"arbpris":89,"leverantor":"BK 2025","kategori":"Beläggning"},
    {"kod":"BK0538","benamning":"Träfiberskiva t=3,2 hård board på yttertak","enhet":"m²","matpris":59,"arbpris":35,"leverantor":"BK 2025","kategori":"Trä & stål"},
    {"kod":"BK0539","benamning":"Träfiberskiva t=6,4 oljehärdad i tak","enhet":"m²","matpris":268,"arbpris":89,"leverantor":"BK 2025","kategori":"Trä & stål"},
    {"kod":"BK0540","benamning":"TRÄFIBERSKIVOR Asfaboard t=12 på vägg","enhet":"m²","matpris":127,"arbpris":64,"leverantor":"BK 2025","kategori":"Mark"},
    {"kod":"BK0541","benamning":"Uppsättning på vägg Spånskiva t=10 byggskiva","enhet":"m²","matpris":143,"arbpris":89,"leverantor":"BK 2025","kategori":"Mark"},
    {"kod":"BK0542","benamning":"Spånskiva t=12 byggskiva","enhet":"m²","matpris":150,"arbpris":89,"leverantor":"BK 2025","kategori":"Skivor"},
    {"kod":"BK0543","benamning":"Spånskiva t=16 standardskiva","enhet":"m²","matpris":192,"arbpris":89,"leverantor":"BK 2025","kategori":"Skivor"},
    {"kod":"BK0544","benamning":"Spånskiva t=19 standardskiva","enhet":"m²","matpris":194,"arbpris":89,"leverantor":"BK 2025","kategori":"Skivor"},
    {"kod":"BK0545","benamning":"Spånskiva t=22 standardskiva","enhet":"m²","matpris":224,"arbpris":89,"leverantor":"BK 2025","kategori":"Skivor"},
    {"kod":"BK0546","benamning":"Fiberriktad spånskiva OSB3, Nexfor, t=11, b=900","enhet":"m²","matpris":155,"arbpris":89,"leverantor":"BK 2025","kategori":"Skivor"},
    {"kod":"BK0547","benamning":"Spånskiva t=22 regelgolv","enhet":"m²","matpris":253,"arbpris":84,"leverantor":"BK 2025","kategori":"Beläggning"},
    {"kod":"BK0548","benamning":"Tjocklek 16 mm","enhet":"m²","matpris":256,"arbpris":84,"leverantor":"BK 2025","kategori":"Skivor"},
    {"kod":"BK0550","benamning":"Väggskivor med fals på två långsidor Spånskiva t=10 byggskiva på vägg","enhet":"m²","matpris":143,"arbpris":89,"leverantor":"BK 2025","kategori":"Mark"},
    {"kod":"BK0551","benamning":"Spånskiva t=12 byggskiva på vägg","enhet":"m²","matpris":150,"arbpris":89,"leverantor":"BK 2025","kategori":"Mark"},
    {"kod":"BK0552","benamning":"Invändiga arbeten (Väggar) Slamning, vanlig - manuell","enhet":"m²","matpris":57,"arbpris":84,"leverantor":"BK 2025","kategori":"Mark"},
    {"kod":"BK0553","benamning":"Dito, mindre utrymmen","enhet":"m²","matpris":71,"arbpris":129,"leverantor":"BK 2025","kategori":"Puts"},
    {"kod":"BK0554","benamning":"Slamning, stockad - manuell","enhet":"m²","matpris":103,"arbpris":148,"leverantor":"BK 2025","kategori":"El"},
    {"kod":"BK0555","benamning":"Slätputs invändigt, tj=6 mm","enhet":"m²","matpris":218,"arbpris":153,"leverantor":"BK 2025","kategori":"Puts"},
    {"kod":"BK0557","benamning":"Tjocklek max ca 25 mm Horisontala ytor, golv","enhet":"m²","matpris":262,"arbpris":183,"leverantor":"BK 2025","kategori":"Beläggning"},
    {"kod":"BK0558","benamning":"Vertikala ytor, väggar","enhet":"m²","matpris":310,"arbpris":307,"leverantor":"BK 2025","kategori":"Mark"},
    {"kod":"BK0559","benamning":"Horisontala underytor, tak","enhet":"m²","matpris":345,"arbpris":401,"leverantor":"BK 2025","kategori":"Tak"},
    {"kod":"BK0560","benamning":"Sika Tricosal, Silix etc","enhet":"m²","matpris":18,"arbpris":15,"leverantor":"BK 2025","kategori":"Betong"},
    {"kod":"BK0561","benamning":"Stålglättade betongsocklar Höjd upp till 100 mm","enhet":"m","matpris":132,"arbpris":148,"leverantor":"BK 2025","kategori":"Betong"},
    {"kod":"BK0565","benamning":"Hålkälar, 20–100 mm radie","enhet":"m","matpris":49,"arbpris":99,"leverantor":"BK 2025","kategori":"Betong"},
    {"kod":"BK0567","benamning":"Underlag: betong, tegel Grundning + stänkputs","enhet":"m²","matpris":152,"arbpris":119,"leverantor":"BK 2025","kategori":"Betong"},
    {"kod":"BK0568","benamning":"Grundning + grovputs + stänkputs fin","enhet":"m²","matpris":303,"arbpris":322,"leverantor":"BK 2025","kategori":"Puts"},
    {"kod":"BK0569","benamning":"Dito med spritputs","enhet":"m²","matpris":539,"arbpris":391,"leverantor":"BK 2025","kategori":"Puts"},
    {"kod":"BK0570","benamning":"Grundning + DSM-bruk + silikatputs riven inkl färg och stålnät","enhet":"m²","matpris":1156,"arbpris":124,"leverantor":"BK 2025","kategori":"Trä & stål"},
    {"kod":"BK0571","benamning":"Tunnputs, typ stänk","enhet":"m²","matpris":150,"arbpris":119,"leverantor":"BK 2025","kategori":"Puts"},
    {"kod":"BK0572","benamning":"Sockelputs Stålslipad sockelputs","enhet":"m²","matpris":310,"arbpris":307,"leverantor":"BK 2025","kategori":"Trä & stål"},
    {"kod":"BK0573","benamning":"Spritputs inkl stockning","enhet":"m²","matpris":590,"arbpris":520,"leverantor":"BK 2025","kategori":"Puts"},
    {"kod":"BK0575","benamning":"Invändig putsställning för enbart vägg","enhet":"m²","matpris":84,"arbpris":35,"leverantor":"BK 2025","kategori":"Mark"},
    {"kod":"BK0587","benamning":"Sprutad latexfärg 16–00008 SP","enhet":"m²","matpris":35,"arbpris":6,"leverantor":"BK 2025","kategori":"Mark"},
    {"kod":"BK0591","benamning":"På slät putsyta 12–00010, struken silikatfärg","enhet":"m²","matpris":76,"arbpris":14,"leverantor":"BK 2025","kategori":"Puts"},
    {"kod":"BK0592","benamning":"På ojämn yta 12–00010, ojämn putsyta","enhet":"m²","matpris":103,"arbpris":18,"leverantor":"BK 2025","kategori":"Puts"},
    {"kod":"BK0599","benamning":"På putsytor 13–00007, putsad yta","enhet":"m²","matpris":40,"arbpris":7,"leverantor":"BK 2025","kategori":"Puts"},
    {"kod":"BK0601","benamning":"På slät betong och lättbetong 23–00707, slät betong/lättbetong","enhet":"m²","matpris":65,"arbpris":12,"leverantor":"BK 2025","kategori":"Betong"},
    {"kod":"BK0603","benamning":"På obehandlade ytor 75–29262","enhet":"m²","matpris":109,"arbpris":19,"leverantor":"BK 2025","kategori":"Målning"},
    {"kod":"BK0607","benamning":"På fabriksgrundade ytor 75–20008","enhet":"m²","matpris":65,"arbpris":12,"leverantor":"BK 2025","kategori":"Målning"},
    {"kod":"BK0611","benamning":"Lackfärg 15–00010","enhet":"m²","matpris":128,"arbpris":22,"leverantor":"BK 2025","kategori":"Målning"},
    {"kod":"BK0612","benamning":"Täckfärg på trägolv. Intill 50 m2 65–00010, klarlack","enhet":"m²","matpris":90,"arbpris":16,"leverantor":"BK 2025","kategori":"Trä & stål"},
    {"kod":"BK0614","benamning":"På putsade ytor 16–01023 (flerbostadshus)","enhet":"m²","matpris":203,"arbpris":36,"leverantor":"BK 2025","kategori":"Puts"},
    {"kod":"BK0618","benamning":"På slät betong och lättbetong 26–00723","enhet":"m²","matpris":203,"arbpris":36,"leverantor":"BK 2025","kategori":"Betong"},
    {"kod":"BK0621","benamning":"På skivor 56–02819","enhet":"m²","matpris":204,"arbpris":36,"leverantor":"BK 2025","kategori":"Skivor"},
    {"kod":"BK0625","benamning":"På putsytor, inkl Borosan/Decorama 11–01004, sprutspacklad","enhet":"m²","matpris":138,"arbpris":24,"leverantor":"BK 2025","kategori":"Puts"},
    {"kod":"BK0629","benamning":"På slät betong och lättbetong inkl Borosan/Decorama 21–00704, slät betong/lättbetong","enhet":"m²","matpris":123,"arbpris":22,"leverantor":"BK 2025","kategori":"Betong"},
    {"kod":"BK0633","benamning":"På träbaserade skivor m m inkl Borosan/Decorama 51–03402","enhet":"m²","matpris":129,"arbpris":23,"leverantor":"BK 2025","kategori":"Trä & stål"},
    {"kod":"BK0636","benamning":"Sockel- och fasadbeklädnad Huggen yta inkl infästn, tjocklek=30","enhet":"m²","matpris":4029,"arbpris":711,"leverantor":"BK 2025","kategori":"El"},
    {"kod":"BK0637","benamning":"Polerad yta inkl infästn, tjocklek=70","enhet":"m²","matpris":5610,"arbpris":990,"leverantor":"BK 2025","kategori":"Beläggning"},
    {"kod":"BK0638","benamning":"Inkl sättsteg. Tjocklek 40 mm Polerad plan yta","enhet":"m","matpris":3373,"arbpris":595,"leverantor":"BK 2025","kategori":"Beläggning"},
    {"kod":"BK0639","benamning":"Dito sättsteg rak trappa","enhet":"m","matpris":2638,"arbpris":466,"leverantor":"BK 2025","kategori":"Beläggning"},
    {"kod":"BK0640","benamning":"Dito sättsteg svängd trappa","enhet":"m","matpris":3811,"arbpris":673,"leverantor":"BK 2025","kategori":"Beläggning"},
    {"kod":"BK0641","benamning":"Planhuggen yta","enhet":"m","matpris":2479,"arbpris":438,"leverantor":"BK 2025","kategori":"Beläggning"},
    {"kod":"BK0642","benamning":"Krysshamrad yta","enhet":"m","matpris":2479,"arbpris":438,"leverantor":"BK 2025","kategori":"Beläggning"},
    {"kod":"BK0643","benamning":"Fasad- och väggbeklädnad Kalksten Jämtland inkl infästning, tj=40","enhet":"m²","matpris":4243,"arbpris":749,"leverantor":"BK 2025","kategori":"Mark"},
    {"kod":"BK0644","benamning":"Dito, <20 m2, tj=20","enhet":"m²","matpris":2122,"arbpris":374,"leverantor":"BK 2025","kategori":"Beläggning"},
    {"kod":"BK0645","benamning":"Trappor av kalksten Kalksten Jämtland plan + sättsteg tj=30","enhet":"m","matpris":2836,"arbpris":500,"leverantor":"BK 2025","kategori":"Beläggning"},
    {"kod":"BK0647","benamning":"Kalkstenssockel h=100","enhet":"m","matpris":327,"arbpris":58,"leverantor":"BK 2025","kategori":"Beläggning"},
    {"kod":"BK0648","benamning":"Fasad- och väggbeklädnad Polerad sten, tjocklek 25–30 mm","enhet":"m²","matpris":4957,"arbpris":875,"leverantor":"BK 2025","kategori":"Mark"},
    {"kod":"BK0649","benamning":"Golv av 30 mm marmor Polerad yta inkl infästning","enhet":"m²","matpris":2122,"arbpris":374,"leverantor":"BK 2025","kategori":"Beläggning"},
    {"kod":"BK0650","benamning":"Polerad i mönster inkl infästning","enhet":"m²","matpris":4957,"arbpris":875,"leverantor":"BK 2025","kategori":"Beläggning"},
    {"kod":"BK0651","benamning":"För trappor utan sättsteg avgår _____ m steg 972:– UE Golvsocklar höjd intill 80 mm, rak","enhet":"m","matpris":345,"arbpris":61,"leverantor":"BK 2025","kategori":"Beläggning"},
    {"kod":"BK0654","benamning":"Trappsocklar, rak i trapplutning","enhet":"m","matpris":1843,"arbpris":325,"leverantor":"BK 2025","kategori":"Beläggning"},
    {"kod":"BK0655","benamning":"Slät klovyta, tjocklek ca 10–15 mm med sågade kanter","enhet":"m²","matpris":1536,"arbpris":271,"leverantor":"BK 2025","kategori":"Beläggning"},
    {"kod":"BK0656","benamning":"Plansten, bredd ca 350 mm, tjocklek 30–40 mm Klovyta, sågade kanter","enhet":"m","matpris":1198,"arbpris":212,"leverantor":"BK 2025","kategori":"Beläggning"},
    {"kod":"BK0657","benamning":"För sättsteg tilläggs","enhet":"m","matpris":306,"arbpris":54,"leverantor":"BK 2025","kategori":"Beläggning"},
    {"kod":"BK0658","benamning":"Klinker på golv i sättbruk","enhet":"m²","matpris":1159,"arbpris":204,"leverantor":"BK 2025","kategori":"Beläggning"},
    {"kod":"BK0659","benamning":"Dito halksäkra på golv i sättbruk","enhet":"m²","matpris":1214,"arbpris":214,"leverantor":"BK 2025","kategori":"Beläggning"},
    {"kod":"BK0660","benamning":"Klinker på golv i torra utrymmen","enhet":"m²","matpris":896,"arbpris":158,"leverantor":"BK 2025","kategori":"Beläggning"},
    {"kod":"BK0661","benamning":"Dito på golv i torra små utrymmen","enhet":"m²","matpris":895,"arbpris":158,"leverantor":"BK 2025","kategori":"Beläggning"},
    {"kod":"BK0662","benamning":"Klinkerplattor på golv i våtrum","enhet":"m²","matpris":1241,"arbpris":219,"leverantor":"BK 2025","kategori":"Beläggning"},
    {"kod":"BK0663","benamning":"Dito halksäkra på golv i våtrum","enhet":"m²","matpris":1372,"arbpris":242,"leverantor":"BK 2025","kategori":"Beläggning"},
    {"kod":"BK0664","benamning":"Klinkerplattor plan- och sättsteg i sättbruk __m2 1291:– UE Klinkersockel, h=100","enhet":"m","matpris":245,"arbpris":43,"leverantor":"BK 2025","kategori":"Beläggning"},
    {"kod":"BK0665","benamning":"Dito hålkäl, h=100","enhet":"m","matpris":452,"arbpris":80,"leverantor":"BK 2025","kategori":"Beläggning"},
    {"kod":"BK0666","benamning":"Klinker på vägg i torra utrymmen","enhet":"m²","matpris":940,"arbpris":166,"leverantor":"BK 2025","kategori":"Mark"},
    {"kod":"BK0667","benamning":"Dito på vägg i våtrum V12","enhet":"m²","matpris":1148,"arbpris":202,"leverantor":"BK 2025","kategori":"Mark"},
    {"kod":"BK0668","benamning":"Ek 7mm Kortstav","enhet":"m²","matpris":498,"arbpris":163,"leverantor":"BK 2025","kategori":"Beläggning"},
    {"kod":"BK0672","benamning":"Bok 7mm Kortstav","enhet":"m²","matpris":596,"arbpris":163,"leverantor":"BK 2025","kategori":"Beläggning"},
    {"kod":"BK0675","benamning":"Ek 14 mm Kortstav","enhet":"m²","matpris":985,"arbpris":163,"leverantor":"BK 2025","kategori":"Beläggning"},
    {"kod":"BK0679","benamning":"Bok 14 mm Kortstav","enhet":"m²","matpris":965,"arbpris":163,"leverantor":"BK 2025","kategori":"Beläggning"},
    {"kod":"BK0681","benamning":"Ek 22 mm Kortstav","enhet":"m²","matpris":1306,"arbpris":163,"leverantor":"BK 2025","kategori":"Beläggning"},
    {"kod":"BK0684","benamning":"Bok 22 mm Kortstav","enhet":"m²","matpris":823,"arbpris":173,"leverantor":"BK 2025","kategori":"Beläggning"},
    {"kod":"BK0687","benamning":"MOSAIKMASSA Cementmosaikmassa, standardfärg","enhet":"m²","matpris":1336,"arbpris":236,"leverantor":"BK 2025","kategori":"Beläggning"},
    {"kod":"BK0688","benamning":"Cementmosaiksockel höjd 70 mm","enhet":"m","matpris":269,"arbpris":47,"leverantor":"BK 2025","kategori":"El"},
    {"kod":"BK0689","benamning":"Lagerstandard, varmförzinkat, inkl ram 1000 x 500 mm","enhet":"st","matpris":4176,"arbpris":346,"leverantor":"BK 2025","kategori":"Beläggning"},
    {"kod":"BK0692","benamning":"Torkmatta Original, komplett inkl ram 600 x 400 mm","enhet":"st","matpris":2366,"arbpris":272,"leverantor":"BK 2025","kategori":"Beläggning"},
    {"kod":"BK0699","benamning":"Torkmatta av räfflat gummi, inkl ram Svart 590 x 390 mm","enhet":"st","matpris":2565,"arbpris":272,"leverantor":"BK 2025","kategori":"Beläggning"},
    {"kod":"BK0703","benamning":"Grå 590 x 390 mm","enhet":"st","matpris":2741,"arbpris":272,"leverantor":"BK 2025","kategori":"Beläggning"},
    {"kod":"BK0707","benamning":"Drevning av karmar Drevning mellan karm och vägg 12 x 100 __ m 31:– 0,04 Dito tvåsidig","enhet":"m","matpris":18,"arbpris":30,"leverantor":"BK 2025","kategori":"Mark"},
    {"kod":"BK0708","benamning":"Öppningsbara, färdigmålade 6 x 6 M","enhet":"st","matpris":7748,"arbpris":505,"leverantor":"BK 2025","kategori":"Beläggning"},
    {"kod":"BK0723","benamning":"Färdigmålade fönster av trä 6 x 6 M","enhet":"st","matpris":6923,"arbpris":564,"leverantor":"BK 2025","kategori":"Trä & stål"},
    {"kod":"BK0745","benamning":"Fönster av trä med utsida av aluminium 5 x 10 M","enhet":"st","matpris":9690,"arbpris":564,"leverantor":"BK 2025","kategori":"Trä & stål"},
    {"kod":"BK0762","benamning":"Fönster av plast 10 x 10 M","enhet":"st","matpris":3853,"arbpris":505,"leverantor":"BK 2025","kategori":"Sakvaror"},
    {"kod":"BK0768","benamning":"Enkeldörrar av trä, färdigmålade 8 x 20 M","enhet":"st","matpris":17386,"arbpris":505,"leverantor":"BK 2025","kategori":"Trä & stål"},
    {"kod":"BK0774","benamning":"Parfönsterdörrar av trä, färdigmålade 12 x 21 M","enhet":"st","matpris":28721,"arbpris":614,"leverantor":"BK 2025","kategori":"Trä & stål"},
    {"kod":"BK0777","benamning":"Fönsterdörrar av trä/aluminium, enkeldörrar 8 x 20 M","enhet":"st","matpris":20987,"arbpris":505,"leverantor":"BK 2025","kategori":"Trä & stål"},
    {"kod":"BK0783","benamning":"Parfönsterdörrar av trä/aluminium 12 x 21 M","enhet":"st","matpris":35605,"arbpris":614,"leverantor":"BK 2025","kategori":"Trä & stål"},
    {"kod":"BK0785","benamning":"Fönsterdörrar av aluminium, enkeldörrar 9 x 21 M","enhet":"st","matpris":23533,"arbpris":614,"leverantor":"BK 2025","kategori":"Sakvaror"},
    {"kod":"BK0788","benamning":"Parfönsterdörrar av aluminium 12 x 21 M","enhet":"st","matpris":30410,"arbpris":614,"leverantor":"BK 2025","kategori":"Sakvaror"},
    {"kod":"BK0791","benamning":"Träfönster, färdigmålade 6 x 6 M","enhet":"st","matpris":3798,"arbpris":505,"leverantor":"BK 2025","kategori":"Trä & stål"},
    {"kod":"BK0812","benamning":"Träfönster med utsida av aluminium 6 x 6 M","enhet":"st","matpris":4426,"arbpris":505,"leverantor":"BK 2025","kategori":"Trä & stål"},
    {"kod":"BK0832","benamning":"Fönster helt av aluminium 6 x 6 M","enhet":"st","matpris":3641,"arbpris":505,"leverantor":"BK 2025","kategori":"Sakvaror"},
    {"kod":"BK0852","benamning":"Inkl karm och emballage Furu 9 x 21, målad","enhet":"st","matpris":15725,"arbpris":614,"leverantor":"BK 2025","kategori":"Beläggning"},
    {"kod":"BK0864","benamning":"Inkl isolerglas, mönstrat Furu 3 x 21 målad","enhet":"st","matpris":11490,"arbpris":505,"leverantor":"BK 2025","kategori":"Beläggning"},
    {"kod":"BK0870","benamning":"A60, inkl instickskarm och emballage Stål 9 x 21","enhet":"st","matpris":6998,"arbpris":1238,"leverantor":"BK 2025","kategori":"Trä & stål"},
    {"kod":"BK0874","benamning":"Parytterdörr 15 x 21","enhet":"st","matpris":14968,"arbpris":1485,"leverantor":"BK 2025","kategori":"Sakvaror"},
    {"kod":"BK0877","benamning":"Lätta innerdörrar 7–9 x 21, målad","enhet":"st","matpris":1409,"arbpris":337,"leverantor":"BK 2025","kategori":"Sakvaror"},
    {"kod":"BK0878","benamning":"Allmogedörrar 7–9 x 21, allmoge, målad","enhet":"st","matpris":1985,"arbpris":337,"leverantor":"BK 2025","kategori":"Sakvaror"},
    {"kod":"BK0879","benamning":"EI30/Rw35dB, 9x21 Inbrottsklass 1, fanerad","enhet":"st","matpris":13645,"arbpris":614,"leverantor":"BK 2025","kategori":"Beläggning"},
    {"kod":"BK0881","benamning":"Inbrottsklass 2, fanerad","enhet":"st","matpris":13645,"arbpris":614,"leverantor":"BK 2025","kategori":"Beläggning"},
    {"kod":"BK0883","benamning":"Lätt innerdörr 7, 8, 9 x 21, slät, målad","enhet":"st","matpris":1409,"arbpris":337,"leverantor":"BK 2025","kategori":"Sakvaror"},
    {"kod":"BK0884","benamning":"Dito, slät fanerad","enhet":"st","matpris":2411,"arbpris":337,"leverantor":"BK 2025","kategori":"Beläggning"},
    {"kod":"BK0885","benamning":"Slät, glasad, massiv innerdörr, 8, 9 x 21 Målad","enhet":"st","matpris":5253,"arbpris":614,"leverantor":"BK 2025","kategori":"Sakvaror"},
    {"kod":"BK0888","benamning":"Släta pardörrar exkl beslagning och foder Lätt, målad parinnerdörr 13, 15 x 21","enhet":"st","matpris":4974,"arbpris":475,"leverantor":"BK 2025","kategori":"Sakvaror"},
    {"kod":"BK0889","benamning":"Skjutdörrar, enkeldörrar, inkl karm och beslagning Infälld skjutdörr 8, 9 x 21, målad","enhet":"st","matpris":3966,"arbpris":842,"leverantor":"BK 2025","kategori":"Sakvaror"},
    {"kod":"BK0890","benamning":"Dito fanerad","enhet":"st","matpris":5068,"arbpris":842,"leverantor":"BK 2025","kategori":"Beläggning"},
    {"kod":"BK0891","benamning":"Karm av trä, målad 7, 8, 9 x 21","enhet":"st","matpris":1998,"arbpris":337,"leverantor":"BK 2025","kategori":"Trä & stål"},
    {"kod":"BK0894","benamning":"Soprumsdörrar inkl drevning Dörr av trä 9 x 21 FM","enhet":"st","matpris":8505,"arbpris":1777,"leverantor":"BK 2025","kategori":"Trä & stål"},
    {"kod":"BK0895","benamning":"Dörr av stål 10 x 21 FM","enhet":"st","matpris":7615,"arbpris":1510,"leverantor":"BK 2025","kategori":"Trä & stål"},
    {"kod":"BK0896","benamning":"Port 25 x 21 Furu, obehandlad","enhet":"st","matpris":44991,"arbpris":1817,"leverantor":"BK 2025","kategori":"Beläggning"},
    {"kod":"BK0898","benamning":"Kompletta dörrar exkl handtag Bastudörr 7 x 19 furu oglasad","enhet":"st","matpris":5963,"arbpris":337,"leverantor":"BK 2025","kategori":"Sakvaror"},
    {"kod":"BK0900","benamning":"Bastudörr 7, 8 x 20 glas","enhet":"st","matpris":6579,"arbpris":337,"leverantor":"BK 2025","kategori":"Sakvaror"},
    {"kod":"BK0901","benamning":"Färdigmålad Kallförrådsdörr 9 x 20 furu","enhet":"st","matpris":6643,"arbpris":614,"leverantor":"BK 2025","kategori":"Sakvaror"},
    {"kod":"BK0905","benamning":"Varmförrådsdörr inkl drevning Furu karmyttermått 9 x 20 – 10 x 21","enhet":"st","matpris":7143,"arbpris":614,"leverantor":"BK 2025","kategori":"Sakvaror"},
    {"kod":"BK0906","benamning":"Entrépartier av metall Enkeldörrspartier inkl 3-glas","enhet":"m²","matpris":7375,"arbpris":1301,"leverantor":"BK 2025","kategori":"Sakvaror"},
    {"kod":"BK0907","benamning":"Invändiga dörrar oisolerade 2-glas","enhet":"m²","matpris":5437,"arbpris":959,"leverantor":"BK 2025","kategori":"Sakvaror"},
    {"kod":"BK0908","benamning":"Pardörrspartier ca 5–6 m2 area Ytterpartier med isolering och 3-glas","enhet":"m²","matpris":5824,"arbpris":1028,"leverantor":"BK 2025","kategori":"Isolering"},
    {"kod":"BK0909","benamning":"Innerpartier oisolerade och 2-glas","enhet":"m²","matpris":5437,"arbpris":959,"leverantor":"BK 2025","kategori":"Beläggning"},
    {"kod":"BK0910","benamning":"Butiksfönsterpartier i storlek mellan ca 5 och 8 m2 med isolerad konstruktion och 3-glasfönster","enhet":"m²","matpris":3162,"arbpris":558,"leverantor":"BK 2025","kategori":"Sakvaror"},
    {"kod":"BK0911","benamning":"Handdrivna skjutportar i storlek mellan ca 8 och 12 m2, industri- och fabriksportar","enhet":"m²","matpris":2856,"arbpris":504,"leverantor":"BK 2025","kategori":"El"},
    {"kod":"BK0912","benamning":"Enkeldörr med karm EI 60 Branddörr 10 x 21","enhet":"st","matpris":7228,"arbpris":1510,"leverantor":"BK 2025","kategori":"Sakvaror"},
    {"kod":"BK0915","benamning":"Parbladig dörr med karm EI 60 Brandpardörr 15 x 21","enhet":"st","matpris":13327,"arbpris":1510,"leverantor":"BK 2025","kategori":"Sakvaror"},
    {"kod":"BK0919","benamning":"Arkivdörrar Arkivdörr 10 x 21 EI 120","enhet":"st","matpris":26557,"arbpris":1287,"leverantor":"BK 2025","kategori":"Sakvaror"},
    {"kod":"BK0920","benamning":"Lucka klass EI 60 07 x 08 vikt ca 25 kg","enhet":"st","matpris":4276,"arbpris":1510,"leverantor":"BK 2025","kategori":"Beläggning"},
    {"kod":"BK0923","benamning":"Dörrblad inkl karm Kylrumsdörr 9 x 21","enhet":"st","matpris":14900,"arbpris":1287,"leverantor":"BK 2025","kategori":"Sakvaror"},
    {"kod":"BK0924","benamning":"Frysrumsdörrar Frysrumsdörr 9 x 21","enhet":"st","matpris":31756,"arbpris":1287,"leverantor":"BK 2025","kategori":"Sakvaror"},
    {"kod":"BK0925","benamning":"Dörr typ YD, karmtyp 21, EI 60 Stålytterdörr YDE 9 x 20","enhet":"st","matpris":8346,"arbpris":1510,"leverantor":"BK 2025","kategori":"Trä & stål"},
    {"kod":"BK0929","benamning":"Diverse beslag m m DORMA 83","enhet":"st","matpris":3645,"arbpris":282,"leverantor":"BK 2025","kategori":"Beläggning"},
    {"kod":"BK0930","benamning":"Brevinkastbeslag med namnskylt","enhet":"st","matpris":1808,"arbpris":0,"leverantor":"BK 2025","kategori":"Beläggning"},
    {"kod":"BK0931","benamning":"Draghandtag, varierande","enhet":"st","matpris":2086,"arbpris":124,"leverantor":"BK 2025","kategori":"Beläggning"},
    {"kod":"BK0932","benamning":"Dörrtrycken, varierande priser","enhet":"st","matpris":918,"arbpris":50,"leverantor":"BK 2025","kategori":"Sakvaror"},
    {"kod":"BK0933","benamning":"Dörrstoppare av gummi","enhet":"st","matpris":68,"arbpris":40,"leverantor":"BK 2025","kategori":"Sakvaror"},
    {"kod":"BK0934","benamning":"Foderlister Färdigmålade 12 x 43","enhet":"st","matpris":0,"arbpris":530,"leverantor":"BK 2025","kategori":"Beläggning"},
    {"kod":"BK0937","benamning":"Bok fabriksbehandlad 12 x 43","enhet":"st","matpris":0,"arbpris":530,"leverantor":"BK 2025","kategori":"Beläggning"},
    {"kod":"BK0939","benamning":"Allmoge 21 x 70","enhet":"st","matpris":0,"arbpris":530,"leverantor":"BK 2025","kategori":"Beläggning"},
    {"kod":"BK0942","benamning":"Fönsterbänkar av kalksten, standard Fönsterbänk 25 x 150 inkl konsoler","enhet":"m","matpris":798,"arbpris":198,"leverantor":"BK 2025","kategori":"Beläggning"},
    {"kod":"BK0943","benamning":"Fönsterbänkar av marmor 20 x 150 polerad inkl konsoler","enhet":"m","matpris":917,"arbpris":198,"leverantor":"BK 2025","kategori":"Sakvaror"},
    {"kod":"BK0944","benamning":"Fönsterbänk 20 x 160 mdf vit, inkl konsoler Längd 1000 mm","enhet":"st","matpris":515,"arbpris":198,"leverantor":"BK 2025","kategori":"Skivor"},
    {"kod":"BK0948","benamning":"Plastkupoler invändigt mått 1000 x 1000 mm Fast kupol 2-lags","enhet":"st","matpris":9043,"arbpris":742,"leverantor":"BK 2025","kategori":"Tak"},
    {"kod":"BK0950","benamning":"Öppningsbar kupol 2-lags","enhet":"st","matpris":11302,"arbpris":990,"leverantor":"BK 2025","kategori":"Tak"},
    {"kod":"BK0952","benamning":"Undertak i bärverk, kant A Mjuk mineralfiber","enhet":"m²","matpris":363,"arbpris":64,"leverantor":"BK 2025","kategori":"Tak"},
    {"kod":"BK0953","benamning":"Våtpressad mineralfiber","enhet":"m²","matpris":316,"arbpris":56,"leverantor":"BK 2025","kategori":"Tak"},
    {"kod":"BK0954","benamning":"Mjuk mineralfiber, diagonalmonterad","enhet":"m²","matpris":456,"arbpris":80,"leverantor":"BK 2025","kategori":"Tak"},
    {"kod":"BK0955","benamning":"Undertak i bärverk Mjuk mineralfiber, kant D","enhet":"m²","matpris":461,"arbpris":81,"leverantor":"BK 2025","kategori":"Tak"},
    {"kod":"BK0956","benamning":"Hygienskiva, kant A","enhet":"m²","matpris":475,"arbpris":84,"leverantor":"BK 2025","kategori":"Tak"},
    {"kod":"BK0957","benamning":"Mineralfiber punktklistrad, kant B","enhet":"m²","matpris":403,"arbpris":71,"leverantor":"BK 2025","kategori":"Tak"},
    {"kod":"BK0959","benamning":"Undertak i bärverk Träullsplatta grå","enhet":"m²","matpris":455,"arbpris":80,"leverantor":"BK 2025","kategori":"Trä & stål"},
    {"kod":"BK0961","benamning":"Diagonalmonterat bärverk Träullsplatta trävit","enhet":"m²","matpris":613,"arbpris":108,"leverantor":"BK 2025","kategori":"Trä & stål"},
    {"kod":"BK0962","benamning":"Direktmonterade Träullsplatta grå","enhet":"m²","matpris":455,"arbpris":80,"leverantor":"BK 2025","kategori":"Trä & stål"},
    {"kod":"BK0964","benamning":"Undertak i bärverk Mjuk mineralfiber, kant A","enhet":"m²","matpris":403,"arbpris":71,"leverantor":"BK 2025","kategori":"Tak"},
    {"kod":"BK0965","benamning":"Absorbent i MT-bärverk Bullerskiva tj= 50, natur","enhet":"m²","matpris":286,"arbpris":50,"leverantor":"BK 2025","kategori":"Tak"},
    {"kod":"BK0968","benamning":"Skruvad Bullerskiva tj= 50, natur","enhet":"m²","matpris":275,"arbpris":49,"leverantor":"BK 2025","kategori":"Tak"},
    {"kod":"BK0969","benamning":"Skärmhöjd = 2100 mm. Plastlaminat 600 mm","enhet":"m","matpris":3251,"arbpris":1005,"leverantor":"BK 2025","kategori":"Tak"},
    {"kod":"BK0971","benamning":"PVC-plast 600 mm","enhet":"m","matpris":3293,"arbpris":1005,"leverantor":"BK 2025","kategori":"Tak"},
    {"kod":"BK0973","benamning":"Infästning för takpannor Med 1 skyddsrör galv","enhet":"m","matpris":305,"arbpris":163,"leverantor":"BK 2025","kategori":"Tak"},
    {"kod":"BK0974","benamning":"Med 2 skyddsrör galv","enhet":"m","matpris":401,"arbpris":163,"leverantor":"BK 2025","kategori":"VVS"},
    {"kod":"BK0975","benamning":"Fasadstege, fast","enhet":"m","matpris":1126,"arbpris":199,"leverantor":"BK 2025","kategori":"Tak"},
    {"kod":"BK0976","benamning":"Husstegar av aluminium Längd 4,2 m","enhet":"st","matpris":4009,"arbpris":50,"leverantor":"BK 2025","kategori":"Tak"},
    {"kod":"BK0977","benamning":"Furu b=900 mm, h=3000 mm Rak","enhet":"st","matpris":70069,"arbpris":2237,"leverantor":"BK 2025","kategori":"Tak"},
    {"kod":"BK0979","benamning":"Med durkplåtsteg h=3000 mm Rak trappa b=700 mm","enhet":"st","matpris":40238,"arbpris":3960,"leverantor":"BK 2025","kategori":"Tak"},
    {"kod":"BK0981","benamning":"Spiraltrappa r=900 mm","enhet":"st","matpris":65134,"arbpris":4950,"leverantor":"BK 2025","kategori":"Tak"},
    {"kod":"BK0982","benamning":"Med gallerduksteg h=3000 mm Rak trappa b=700 mm","enhet":"st","matpris":35058,"arbpris":3960,"leverantor":"BK 2025","kategori":"Tak"},
    {"kod":"BK0985","benamning":"Helloppstrappor h=3000 mm Rak trappa b=900 mm","enhet":"st","matpris":65009,"arbpris":1525,"leverantor":"BK 2025","kategori":"El"},
    {"kod":"BK0986","benamning":"Svängd trappa r=900 mm","enhet":"st","matpris":71273,"arbpris":1782,"leverantor":"BK 2025","kategori":"Tak"},
    {"kod":"BK0987","benamning":"Halvloppstrappor h=1500 mm Rak trappa b=700 mm","enhet":"st","matpris":38871,"arbpris":1277,"leverantor":"BK 2025","kategori":"Tak"},
    {"kod":"BK0989","benamning":"Räcken Trappräcke av stål","enhet":"m","matpris":1571,"arbpris":277,"leverantor":"BK 2025","kategori":"Trä & stål"},
    {"kod":"BK0990","benamning":"Dito med furuhandledare","enhet":"m","matpris":2040,"arbpris":360,"leverantor":"BK 2025","kategori":"Tak"},
    {"kod":"BK0991","benamning":"Ledstänger Ledstång av plastklätt profilstål","enhet":"m","matpris":786,"arbpris":139,"leverantor":"BK 2025","kategori":"Trä & stål"},
    {"kod":"BK0992","benamning":"Ledstång av rostfritt stål","enhet":"m","matpris":1285,"arbpris":227,"leverantor":"BK 2025","kategori":"Trä & stål"},
    {"kod":"BK0993","benamning":"Smidesledstång med furuhandledare","enhet":"m","matpris":1049,"arbpris":185,"leverantor":"BK 2025","kategori":"Tak"},
    {"kod":"BK0994","benamning":"Ledstång av bok ø40","enhet":"m","matpris":617,"arbpris":248,"leverantor":"BK 2025","kategori":"Tak"},
    {"kod":"BK0995","benamning":"Ledstång av furu","enhet":"m","matpris":245,"arbpris":248,"leverantor":"BK 2025","kategori":"Tak"},
    {"kod":"BK0996","benamning":"Sockellist av trä Furu 12 x 43","enhet":"m","matpris":49,"arbpris":40,"leverantor":"BK 2025","kategori":"Trä & stål"},
    {"kod":"BK1000","benamning":"Bok 12 x 56 fabriksbeh","enhet":"m","matpris":160,"arbpris":40,"leverantor":"BK 2025","kategori":"Tak"},
    {"kod":"BK1001","benamning":"Ek 12 x 56 fabriksbeh","enhet":"m","matpris":144,"arbpris":40,"leverantor":"BK 2025","kategori":"Tak"},
    {"kod":"BK1002","benamning":"MDF 12 x 69 fabriksbeh","enhet":"m","matpris":77,"arbpris":40,"leverantor":"BK 2025","kategori":"Skivor"},
    {"kod":"BK1003","benamning":"Foderlist Furu 12 x 43","enhet":"m","matpris":63,"arbpris":35,"leverantor":"BK 2025","kategori":"Tak"},
    {"kod":"BK1006","benamning":"Fabriksbehandlad Ek 8 x 15","enhet":"m","matpris":55,"arbpris":35,"leverantor":"BK 2025","kategori":"Tak"},
    {"kod":"BK1008","benamning":"Nivålister Bok 58 x 20","enhet":"m","matpris":180,"arbpris":50,"leverantor":"BK 2025","kategori":"Tak"},
    {"kod":"BK1011","benamning":"Kvartslist Ek 12 x 12 fabriksbeh","enhet":"m","matpris":58,"arbpris":35,"leverantor":"BK 2025","kategori":"Tak"},
    {"kod":"BK1012","benamning":"Björk 15 x 15 fabriksbeh","enhet":"m","matpris":131,"arbpris":35,"leverantor":"BK 2025","kategori":"Tak"},
    {"kod":"BK1013","benamning":"Trekantslist Trekantlist 21 mm","enhet":"m","matpris":38,"arbpris":30,"leverantor":"BK 2025","kategori":"Tak"},
    {"kod":"BK1015","benamning":"Skugglist/Taklist Furu 21 x 33","enhet":"m","matpris":61,"arbpris":45,"leverantor":"BK 2025","kategori":"Tak"},
    {"kod":"BK1017","benamning":"Hålkälslist Furu 15 x 33","enhet":"m","matpris":48,"arbpris":45,"leverantor":"BK 2025","kategori":"Tak"},
    {"kod":"BK1020","benamning":"Anslagstavlor med våningsregister 50 x 40 cm","enhet":"st","matpris":694,"arbpris":178,"leverantor":"BK 2025","kategori":"Inredning"},
    {"kod":"BK1022","benamning":"XBC.5 Speglar Toalettspeglar 600 x 450 mm","enhet":"st","matpris":250,"arbpris":178,"leverantor":"BK 2025","kategori":"Inredning"},
    {"kod":"BK1023","benamning":"Garderob exkl trådbackar bredd 600 mm","enhet":"st","matpris":3903,"arbpris":446,"leverantor":"BK 2025","kategori":"Inredning"},
    {"kod":"BK1024","benamning":"Garderob med parluckor bredd 800 mm","enhet":"st","matpris":5235,"arbpris":446,"leverantor":"BK 2025","kategori":"Inredning"},
    {"kod":"BK1025","benamning":"Linneskåp bredd 600 mm","enhet":"st","matpris":5275,"arbpris":446,"leverantor":"BK 2025","kategori":"Inredning"},
    {"kod":"BK1026","benamning":"Hyllskåp bredd 600 mm","enhet":"st","matpris":4498,"arbpris":446,"leverantor":"BK 2025","kategori":"Inredning"},
    {"kod":"BK1027","benamning":"Städskåp bredd 600 mm","enhet":"st","matpris":5798,"arbpris":446,"leverantor":"BK 2025","kategori":"Inredning"},
    {"kod":"BK1028","benamning":"Skåp för inbyggnad bredd 600 mm","enhet":"st","matpris":4903,"arbpris":446,"leverantor":"BK 2025","kategori":"Inredning"},
    {"kod":"BK1029","benamning":"Bänkskåp höjd 720 mm (totalhöjd 900 mm) bredd 600 mm","enhet":"st","matpris":2495,"arbpris":391,"leverantor":"BK 2025","kategori":"Inredning"},
    {"kod":"BK1030","benamning":"Grytskåp höjd 720 mm bredd 600 mm","enhet":"st","matpris":2539,"arbpris":391,"leverantor":"BK 2025","kategori":"Inredning"},
    {"kod":"BK1031","benamning":"Lådfack höjd 720 mm bredd 600 mm","enhet":"st","matpris":8801,"arbpris":391,"leverantor":"BK 2025","kategori":"Inredning"},
    {"kod":"BK1032","benamning":"Diskbänksskåp höjd 720 mm bredd 600 mm","enhet":"st","matpris":4032,"arbpris":391,"leverantor":"BK 2025","kategori":"Inredning"},
    {"kod":"BK1033","benamning":"Väggskåp K-höjd 700 mm bredd 600 mm","enhet":"st","matpris":1919,"arbpris":282,"leverantor":"BK 2025","kategori":"Mark"},
    {"kod":"BK1034","benamning":"Fläktskåp bredd 600 mm","enhet":"st","matpris":1785,"arbpris":282,"leverantor":"BK 2025","kategori":"Inredning"},
    {"kod":"BK1035","benamning":"Fläkthylla bredd 600 mm","enhet":"st","matpris":2889,"arbpris":282,"leverantor":"BK 2025","kategori":"Inredning"},
    {"kod":"BK1036","benamning":"Skåp över kyl bredd 600 mm höjd 390 mm","enhet":"st","matpris":2691,"arbpris":282,"leverantor":"BK 2025","kategori":"Inredning"},
    {"kod":"BK1037","benamning":"Badrumsskåp typ Svedbergs Tvilling 1650 metall","enhet":"st","matpris":1398,"arbpris":203,"leverantor":"BK 2025","kategori":"Inredning"},
    {"kod":"BK1038","benamning":"Tvilling 1650 trä","enhet":"st","matpris":2176,"arbpris":203,"leverantor":"BK 2025","kategori":"Trä & stål"},
    {"kod":"BK1039","benamning":"Enkelhyllor Längd 1000 mm","enhet":"st","matpris":1506,"arbpris":228,"leverantor":"BK 2025","kategori":"El"},
    {"kod":"BK1041","benamning":"Dubbelhyllor Längd 1000 mm","enhet":"st","matpris":1550,"arbpris":228,"leverantor":"BK 2025","kategori":"El"},
    {"kod":"BK1043","benamning":"Toalettpappershållare, fast metallpinne, krom","enhet":"st","matpris":452,"arbpris":79,"leverantor":"BK 2025","kategori":"Tak"},
    {"kod":"BK1044","benamning":"Handdukshängare BB1803 vitlack 3-krok _ st 393:– 0,16 Balkonghängare, plast","enhet":"st","matpris":195,"arbpris":74,"leverantor":"BK 2025","kategori":"Inredning"},
    {"kod":"BK1045","benamning":"XBD.6 Behållare, lådor och kärl Soppåshållare med lock","enhet":"st","matpris":2200,"arbpris":74,"leverantor":"BK 2025","kategori":"Inredning"},
    {"kod":"BK1046","benamning":"XBE.2 Bänkskivor Bänkskiva av laminat","enhet":"m","matpris":937,"arbpris":99,"leverantor":"BK 2025","kategori":"Skivor"},
    {"kod":"BK1049","benamning":"Släta Typ F L=1000 mm","enhet":"st","matpris":4228,"arbpris":248,"leverantor":"BK 2025","kategori":"Inredning"},
    {"kod":"BK1051","benamning":"GM L=1600 mm","enhet":"st","matpris":7474,"arbpris":248,"leverantor":"BK 2025","kategori":"Inredning"},
    {"kod":"BK1052","benamning":"HM L=1800 mm","enhet":"st","matpris":7040,"arbpris":248,"leverantor":"BK 2025","kategori":"Inredning"},
    {"kod":"BK1053","benamning":"Mexi fasadsten, slät 250 x 120 x 65","enhet":"m²","matpris":1417,"arbpris":0,"leverantor":"BK 2025","kategori":"Beläggning"},
    {"kod":"BK1056","benamning":"Mexi fasadsten, rundhuggen 250 x 100 x 65","enhet":"m²","matpris":1386,"arbpris":0,"leverantor":"BK 2025","kategori":"Beläggning"},
    {"kod":"BK1059","benamning":"Mexi fasadsten, rakhuggen 250 x 85 x 65","enhet":"m²","matpris":1184,"arbpris":0,"leverantor":"BK 2025","kategori":"Beläggning"},
    {"kod":"BK1062","benamning":"Märklast 17,0 kN/m. Höjd 200 mm, 300 mm Lättbetongbalk 150x200 mm","enhet":"lm","matpris":878,"arbpris":0,"leverantor":"BK 2025","kategori":"Betong"},
    {"kod":"BK1071","benamning":"Manuellt <100 mm","enhet":"m²","matpris":595,"arbpris":893,"leverantor":"BK 2025","kategori":"El"},
    {"kod":"BK1074","benamning":"Maskinellt <100 mm","enhet":"m²","matpris":483,"arbpris":725,"leverantor":"BK 2025","kategori":"El"},
    {"kod":"BK1077","benamning":"Manuellt mindre utrymmen","enhet":"m²","matpris":179,"arbpris":269,"leverantor":"BK 2025","kategori":"El"},
    {"kod":"BK1078","benamning":"Maskinellt större utrymmen","enhet":"m²","matpris":150,"arbpris":226,"leverantor":"BK 2025","kategori":"El"},
    {"kod":"BK1079","benamning":"Maskinellt tj < 200 mm","enhet":"m²","matpris":624,"arbpris":937,"leverantor":"BK 2025","kategori":"El"},
    {"kod":"BK1083","benamning":"Fuktisolerande golvmassa i våtutrymme","enhet":"m²","matpris":209,"arbpris":313,"leverantor":"BK 2025","kategori":"Beläggning"},
    {"kod":"BK1085","benamning":"Kilade reglar","enhet":"m²","matpris":61,"arbpris":91,"leverantor":"BK 2025","kategori":"Betong"},
    {"kod":"BK1086","benamning":"Uppstolpade reglar","enhet":"m²","matpris":61,"arbpris":91,"leverantor":"BK 2025","kategori":"Betong"},
    {"kod":"BK1087","benamning":"Spontad panel på golv","enhet":"m²","matpris":65,"arbpris":98,"leverantor":"BK 2025","kategori":"Beläggning"},
    {"kod":"BK1088","benamning":"Board på golv","enhet":"m²","matpris":39,"arbpris":59,"leverantor":"BK 2025","kategori":"Beläggning"},
    {"kod":"BK1089","benamning":"Lätt fyllning i golv (spån, mu)","enhet":"m²","matpris":16,"arbpris":24,"leverantor":"BK 2025","kategori":"Beläggning"},
    {"kod":"BK1090","benamning":"Tung fyllning i golv","enhet":"m²","matpris":74,"arbpris":112,"leverantor":"BK 2025","kategori":"Beläggning"},
    {"kod":"BK1091","benamning":"Cellplast i golv","enhet":"m²","matpris":23,"arbpris":34,"leverantor":"BK 2025","kategori":"Isolering"},
    {"kod":"BK1092","benamning":"Cellplast under golv","enhet":"m²","matpris":60,"arbpris":91,"leverantor":"BK 2025","kategori":"Isolering"},
    {"kod":"BK1093","benamning":"Väggar av armerad betong Tjocklek <100 mm","enhet":"m²","matpris":726,"arbpris":1090,"leverantor":"BK 2025","kategori":"Mark"},
    {"kod":"BK1097","benamning":"Väggar av murtegel Tjocklek <150 mm","enhet":"m²","matpris":256,"arbpris":384,"leverantor":"BK 2025","kategori":"Mark"},
    {"kod":"BK1102","benamning":"Väggar av betonghålsten Tjocklek <150 mm","enhet":"m²","matpris":314,"arbpris":472,"leverantor":"BK 2025","kategori":"Mark"},
    {"kod":"BK1105","benamning":"Väggar av lättklinkerblock (Leca) Tjocklek <100 mm","enhet":"m²","matpris":100,"arbpris":150,"leverantor":"BK 2025","kategori":"Mark"},
    {"kod":"BK1109","benamning":"Vägg av lättbetongblock Tjocklek <100 mm","enhet":"m²","matpris":96,"arbpris":144,"leverantor":"BK 2025","kategori":"Mark"},
    {"kod":"BK1113","benamning":"Vägg av lättbetongelement Tjocklek <100 mm","enhet":"m²","matpris":96,"arbpris":144,"leverantor":"BK 2025","kategori":"Mark"},
    {"kod":"BK1117","benamning":"Systeminnervägg, typ Eurowand","enhet":"m²","matpris":80,"arbpris":119,"leverantor":"BK 2025","kategori":"Mark"},
    {"kod":"BK1133","benamning":"Övrig rivning yttervägg Utvändig puts","enhet":"m²","matpris":85,"arbpris":127,"leverantor":"BK 2025","kategori":"Mark"},
    {"kod":"BK1134","benamning":"Utvändig panel på vägg","enhet":"m²","matpris":68,"arbpris":102,"leverantor":"BK 2025","kategori":"Mark"},
    {"kod":"BK1135","benamning":"Lockläkt på vägg","enhet":"m²","matpris":25,"arbpris":37,"leverantor":"BK 2025","kategori":"Mark"},
    {"kod":"BK1136","benamning":"Profilerad plåt på vägg","enhet":"m²","matpris":19,"arbpris":29,"leverantor":"BK 2025","kategori":"Mark"},
    {"kod":"BK1137","benamning":"Slät plåt på vägg","enhet":"m²","matpris":19,"arbpris":28,"leverantor":"BK 2025","kategori":"Mark"},
    {"kod":"BK1138","benamning":"Fasadbeklädnad av skivor","enhet":"m²","matpris":32,"arbpris":47,"leverantor":"BK 2025","kategori":"Skivor"},
    {"kod":"BK1139","benamning":"Fasadtegel tj <100","enhet":"m²","matpris":198,"arbpris":296,"leverantor":"BK 2025","kategori":"Murning"},
    {"kod":"BK1145","benamning":"Läkt på vägg","enhet":"m²","matpris":37,"arbpris":55,"leverantor":"BK 2025","kategori":"Mark"},
    {"kod":"BK1146","benamning":"Spontad panel i yttervägg","enhet":"m²","matpris":68,"arbpris":102,"leverantor":"BK 2025","kategori":"Mark"},
    {"kod":"BK1147","benamning":"Träregelstomme enkel i vägg","enhet":"m²","matpris":52,"arbpris":78,"leverantor":"BK 2025","kategori":"Mark"},
    {"kod":"BK1149","benamning":"Stålregelstomme enkel i vägg","enhet":"m²","matpris":50,"arbpris":74,"leverantor":"BK 2025","kategori":"Mark"},
    {"kod":"BK1157","benamning":"Limträpelare mindre dimension","enhet":"m","matpris":14,"arbpris":21,"leverantor":"BK 2025","kategori":"Trä & stål"},
    {"kod":"BK1159","benamning":"Bär- och ströläkt på yttertak","enhet":"m²","matpris":10,"arbpris":16,"leverantor":"BK 2025","kategori":"Tak"},
    {"kod":"BK1160","benamning":"Underlagspapp på yttertak","enhet":"m²","matpris":13,"arbpris":20,"leverantor":"BK 2025","kategori":"Tak"},
    {"kod":"BK1162","benamning":"Spontade bräder på yttertak","enhet":"m²","matpris":32,"arbpris":48,"leverantor":"BK 2025","kategori":"Tak"},
    {"kod":"BK1163","benamning":"Profilerad plåt på yttertak","enhet":"m²","matpris":24,"arbpris":37,"leverantor":"BK 2025","kategori":"Tak"},
    {"kod":"BK1164","benamning":"Slät plåt på yttertak","enhet":"m²","matpris":24,"arbpris":37,"leverantor":"BK 2025","kategori":"Tak"},
    {"kod":"BK1165","benamning":"Tätskikt av papp/duk på yttertak_","enhet":"m²","matpris":17,"arbpris":25,"leverantor":"BK 2025","kategori":"Tak"},
    {"kod":"BK1168","benamning":"Bärande plåt på yttertak","enhet":"m²","matpris":52,"arbpris":79,"leverantor":"BK 2025","kategori":"Tak"},
    {"kod":"BK1171","benamning":"TT-kassetter av betong","enhet":"m²","matpris":276,"arbpris":414,"leverantor":"BK 2025","kategori":"Betong"},
    {"kod":"BK1172","benamning":"Lättbalk av stål, Z-/C-profil i tak","enhet":"m²","matpris":25,"arbpris":37,"leverantor":"BK 2025","kategori":"Trä & stål"},
    {"kod":"BK1174","benamning":"Lättbalk av trä i tak","enhet":"m²","matpris":25,"arbpris":38,"leverantor":"BK 2025","kategori":"Trä & stål"},
    {"kod":"BK1175","benamning":"Limträbalk/limträås h < 300","enhet":"m","matpris":31,"arbpris":46,"leverantor":"BK 2025","kategori":"Trä & stål"},
    {"kod":"BK1184","benamning":"Räcke för takhuv","enhet":"st","matpris":56,"arbpris":83,"leverantor":"BK 2025","kategori":"Tak"},
    {"kod":"BK1196","benamning":"Uppstolpade takstolar","enhet":"m²","matpris":20,"arbpris":31,"leverantor":"BK 2025","kategori":"Tak"},
    {"kod":"BK1230","benamning":"Tillägg för rivning och montering av fönster över 3,25 m höjd","enhet":"st","matpris":60,"arbpris":89,"leverantor":"BK 2025","kategori":"Sakvaror"},
    {"kod":"BK1231","benamning":"Tillägg för demontering av hela fönster/fönsterdörr","enhet":"st","matpris":83,"arbpris":125,"leverantor":"BK 2025","kategori":"Sakvaror"},
    {"kod":"BK1232","benamning":"Tillägg för demontering av hela fönster/fönsterdörr vid putsad fasad","enhet":"st","matpris":107,"arbpris":160,"leverantor":"BK 2025","kategori":"Puts"},
    {"kod":"BK1248","benamning":"Borttagning av lös puts på vägg (hela väggytan mäts) 10%","enhet":"m²","matpris":7,"arbpris":11,"leverantor":"BK 2025","kategori":"Mark"},
    {"kod":"BK1256","benamning":"Borttagning av lös puts i tak (hela takytan mäts) 10%","enhet":"m²","matpris":8,"arbpris":13,"leverantor":"BK 2025","kategori":"Tak"},
    {"kod":"BK1262","benamning":"Tillägg för rivning av ytterligare lag golvmatta eller plastmatta","enhet":"m²","matpris":7,"arbpris":11,"leverantor":"BK 2025","kategori":"Beläggning"},
    {"kod":"BK1263","benamning":"Tillägg för återanvändning av lister","enhet":"m²","matpris":5,"arbpris":7,"leverantor":"BK 2025","kategori":"Tak"},
    {"kod":"BK1324","benamning":"Borrning betong i golv/bjälklag tj <160 ø30","enhet":"st","matpris":90,"arbpris":136,"leverantor":"BK 2025","kategori":"Betong"},
    {"kod":"BK1326","benamning":"Borrning betong i vägg tj <160 ø30","enhet":"st","matpris":91,"arbpris":136,"leverantor":"BK 2025","kategori":"Mark"},
    {"kod":"BK1329","benamning":"Borrning tegel i vägg tj <150 ø30","enhet":"st","matpris":90,"arbpris":136,"leverantor":"BK 2025","kategori":"Mark"},
    {"kod":"BK1332","benamning":"Bilning i betonggolv/betongbjälklag Bilning för avloppsgrodor","enhet":"m","matpris":475,"arbpris":713,"leverantor":"BK 2025","kategori":"Betong"},
    {"kod":"BK1333","benamning":"Bilning för avloppsgrodor","enhet":"st","matpris":1426,"arbpris":2138,"leverantor":"BK 2025","kategori":"VVS"},
    {"kod":"BK1334","benamning":"Sågning betong i golv/bjälklag tj <100","enhet":"st","matpris":120,"arbpris":180,"leverantor":"BK 2025","kategori":"Betong"},
    {"kod":"BK1338","benamning":"Sågning betong i vägg tj <100","enhet":"m","matpris":206,"arbpris":310,"leverantor":"BK 2025","kategori":"Mark"},
    {"kod":"BK1342","benamning":"Sågning för avloppsgrodor","enhet":"m","matpris":343,"arbpris":514,"leverantor":"BK 2025","kategori":"VVS"},
    {"kod":"BK1343","benamning":"Sågning tegel i vägg tj <150","enhet":"m","matpris":281,"arbpris":421,"leverantor":"BK 2025","kategori":"Mark"},
    {"kod":"BK1347","benamning":"Sågning fasadtegel tj = 60","enhet":"m","matpris":112,"arbpris":169,"leverantor":"BK 2025","kategori":"Murning"},
    {"kod":"BK1350","benamning":"Stålglättade betonggolv, tjocklek 3 cm","enhet":"m²","matpris":0,"arbpris":14850,"leverantor":"BK 2025","kategori":"Betong"},
    {"kod":"BK1351","benamning":"Dito brädrivna","enhet":"m²","matpris":0,"arbpris":15840,"leverantor":"BK 2025","kategori":"Betong"},
    {"kod":"BK1352","benamning":"Formsättning (gängse äldre traditionell metod): Formar till grundmurar och väggar","enhet":"m²","matpris":0,"arbpris":7425,"leverantor":"BK 2025","kategori":"Mark"},
    {"kod":"BK1353","benamning":"Formar till bjälklagsplattor","enhet":"m²","matpris":0,"arbpris":12375,"leverantor":"BK 2025","kategori":"Beläggning"},
    {"kod":"BK1354","benamning":"Formar till pelare och balkar","enhet":"m²","matpris":0,"arbpris":6435,"leverantor":"BK 2025","kategori":"El"},
    {"kod":"BK1355","benamning":"Tegelmurning, normaltegel till puts ____________ sten 600–800 Fasadtegelmurning, exkl fogning _____________ sten 500–650 Fogstrykning","enhet":"m²","matpris":0,"arbpris":24750,"leverantor":"BK 2025","kategori":"Murning"},
    {"kod":"BK1356","benamning":"Lättbetongsmurning, 20–25 cm vägg","enhet":"m²","matpris":0,"arbpris":11385,"leverantor":"BK 2025","kategori":"Mark"},
    {"kod":"BK1357","benamning":"Mellanväggar av 7–10 cm plattor","enhet":"m²","matpris":0,"arbpris":19800,"leverantor":"BK 2025","kategori":"Mark"},
    {"kod":"BK1358","benamning":"Invändig slätputs","enhet":"m²","matpris":0,"arbpris":14850,"leverantor":"BK 2025","kategori":"Puts"},
    {"kod":"BK1359","benamning":"Invändig slamning","enhet":"m²","matpris":0,"arbpris":24750,"leverantor":"BK 2025","kategori":"Mark"},
    {"kod":"BK1360","benamning":"Rörning med specialmatta","enhet":"m²","matpris":0,"arbpris":19800,"leverantor":"BK 2025","kategori":"VVS"},
]

EXEMPELPROJEKT = [{'projektnamn': 'Badrumsrenovering Ringvägen 12', 'projektnummer': '2025-001', 'kund': 'Anders & Maria Lindström', 'bestallar': 'Anders Lindström', 'adress': 'Ringvägen 12, 123 45 Huddinge', 'datum': '2025-03-01', 'kalkylansvarig': 'Jan Stefors', 'status': 'Offert', 'kommentar': 'Badrumsrenovering ca 7m². Befintlig badkar byts mot dusch. Ny klinker golv och vägg. ROT-avdrag tillämpas.', 'byggdelar': ['Rivning', 'Tätskikt', 'Klinker', 'VVS', 'El', 'Övrigt'], 'omkostnader': {'Arbetsledning': 4500.0, 'Etablering': 2000.0, 'Administration': 1500.0, 'Transporter': 2500.0, 'Garanti': 0.0, 'Risk': 0.0, 'Övrigt': 0.0}, 'paslag': {'Omkostnad %': 10.0, 'Risk %': 5.0, 'Vinst %': 8.0, 'Rabatt %': 0.0}, 'rader': [{'Radtyp': 'UE', 'Benämning': 'Rivning befintligt badrum inkl borttransport', 'Byggdel': 'Rivning', 'Mängd': 1.0, 'Enhet': 'ls', 'Timmar': 0.0, 'Á-pris': 18500, 'Påslag %': 0.0, 'Kostnad': 18500.0, 'Försäljning': 18500.0, 'Leverantör': 'UE Rivning AB', 'Kod': '', 'Kommentar': 'Fast pris inkl container'}, {'Radtyp': 'Material', 'Benämning': 'Undergolv betong C25/30 50mm', 'Byggdel': 'Tätskikt', 'Mängd': 7.0, 'Enhet': 'm²', 'Timmar': 0.0, 'Á-pris': 174, 'Påslag %': 0.0, 'Kostnad': 1218.0, 'Försäljning': 1218.0, 'Leverantör': 'BK 2025', 'Kod': 'ESE', 'Kommentar': ''}, {'Radtyp': 'Arbete', 'Benämning': 'Gjutning undergolv', 'Byggdel': 'Tätskikt', 'Mängd': 0.0, 'Enhet': 'tim', 'Timmar': 6.0, 'Á-pris': 595, 'Påslag %': 0.0, 'Kostnad': 3570.0, 'Försäljning': 3570.0, 'Leverantör': 'Eget', 'Kod': '', 'Kommentar': ''}, {'Radtyp': 'Material', 'Benämning': 'Tätskikt badrum vattentät duk', 'Byggdel': 'Tätskikt', 'Mängd': 14.0, 'Enhet': 'm²', 'Timmar': 0.0, 'Á-pris': 485, 'Påslag %': 0.0, 'Kostnad': 6790.0, 'Försäljning': 6790.0, 'Leverantör': 'BK 2025', 'Kod': 'PSF', 'Kommentar': 'Golv + väggar 1m upp'}, {'Radtyp': 'Arbete', 'Benämning': 'Montering tätskikt', 'Byggdel': 'Tätskikt', 'Mängd': 0.0, 'Enhet': 'tim', 'Timmar': 10.0, 'Á-pris': 595, 'Påslag %': 0.0, 'Kostnad': 5950.0, 'Försäljning': 5950.0, 'Leverantör': 'Eget', 'Kod': '', 'Kommentar': ''}, {'Radtyp': 'Material', 'Benämning': 'Klinker 300x300 stengodsklinker', 'Byggdel': 'Klinker', 'Mängd': 7.0, 'Enhet': 'm²', 'Timmar': 0.0, 'Á-pris': 780, 'Påslag %': 0.0, 'Kostnad': 5460.0, 'Försäljning': 5460.0, 'Leverantör': 'BK 2025', 'Kod': 'NSS.1', 'Kommentar': 'Golvklinker halkfri R11'}, {'Radtyp': 'Arbete', 'Benämning': 'Läggning klinker golv inkl fogning', 'Byggdel': 'Klinker', 'Mängd': 0.0, 'Enhet': 'tim', 'Timmar': 12.0, 'Á-pris': 595, 'Påslag %': 0.0, 'Kostnad': 7140.0, 'Försäljning': 7140.0, 'Leverantör': 'Eget', 'Kod': '', 'Kommentar': ''}, {'Radtyp': 'Material', 'Benämning': 'Klinker 200x200 stengodsklinker', 'Byggdel': 'Klinker', 'Mängd': 22.0, 'Enhet': 'm²', 'Timmar': 0.0, 'Á-pris': 825, 'Påslag %': 0.0, 'Kostnad': 18150.0, 'Försäljning': 18150.0, 'Leverantör': 'BK 2025', 'Kod': 'NSS.1', 'Kommentar': 'Väggklinker vit 200x200'}, {'Radtyp': 'Arbete', 'Benämning': 'Läggning klinker vägg inkl fogning', 'Byggdel': 'Klinker', 'Mängd': 0.0, 'Enhet': 'tim', 'Timmar': 28.0, 'Á-pris': 595, 'Påslag %': 0.0, 'Kostnad': 16660.0, 'Försäljning': 16660.0, 'Leverantör': 'Eget', 'Kod': '', 'Kommentar': ''}, {'Radtyp': 'UE', 'Benämning': 'VVS komplett: dusch, wc, tvättställ, rör', 'Byggdel': 'VVS', 'Mängd': 1.0, 'Enhet': 'ls', 'Timmar': 0.0, 'Á-pris': 38500, 'Påslag %': 0.0, 'Kostnad': 38500.0, 'Försäljning': 38500.0, 'Leverantör': 'VVS Proffs AB', 'Kod': '', 'Kommentar': 'Fast pris inkl material'}, {'Radtyp': 'UE', 'Benämning': 'El: belysning, golvvärme, ventilationsfläkt', 'Byggdel': 'El', 'Mängd': 1.0, 'Enhet': 'ls', 'Timmar': 0.0, 'Á-pris': 12500, 'Påslag %': 0.0, 'Kostnad': 12500.0, 'Försäljning': 12500.0, 'Leverantör': 'EL-service', 'Kod': '', 'Kommentar': 'Fast pris inkl material'}, {'Radtyp': 'Material', 'Benämning': 'Spegel + badrumsskåp', 'Byggdel': 'Övrigt', 'Mängd': 1.0, 'Enhet': 'st', 'Timmar': 0.0, 'Á-pris': 3800, 'Påslag %': 0.0, 'Kostnad': 3800.0, 'Försäljning': 3800.0, 'Leverantör': 'IKEA', 'Kod': '', 'Kommentar': ''}, {'Radtyp': 'Arbete', 'Benämning': 'Montering, slutsnickeri och städning', 'Byggdel': 'Övrigt', 'Mängd': 0.0, 'Enhet': 'tim', 'Timmar': 8.0, 'Á-pris': 595, 'Påslag %': 0.0, 'Kostnad': 4760.0, 'Försäljning': 4760.0, 'Leverantör': 'Eget', 'Kod': '', 'Kommentar': ''}]}, {'projektnamn': 'Tillbyggnad Villa Björkgatan 5', 'projektnummer': '2025-002', 'kund': 'Familjen Svensson', 'bestallar': 'Petra Svensson', 'adress': 'Björkgatan 5, 702 28 Örebro', 'datum': '2025-04-15', 'kalkylansvarig': 'Jan Stefors', 'status': 'Kalkyl', 'kommentar': 'Tillbyggnad av befintlig villa, plan 1. Ca 30m² nya boarea. Inkl grund, stomme, tak, fasad och invändigt. Bygglov beviljat 2025-02-10.', 'byggdelar': ['Grund', 'Stomme', 'Tak', 'Fasad', 'Isolering', 'Invändigt', 'El & VVS'], 'omkostnader': {'Arbetsledning': 15000.0, 'Etablering': 8000.0, 'Administration': 5000.0, 'Transporter': 6000.0, 'Garanti': 3000.0, 'Risk': 0.0, 'Övrigt': 2000.0}, 'paslag': {'Omkostnad %': 12.0, 'Risk %': 5.0, 'Vinst %': 10.0, 'Rabatt %': 0.0}, 'rader': [{'Radtyp': 'Material', 'Benämning': 'Maskinschakt källare klass A', 'Byggdel': 'Grund', 'Mängd': 35.0, 'Enhet': 'm³', 'Timmar': 0.0, 'Á-pris': 472, 'Påslag %': 0.0, 'Kostnad': 16520.0, 'Försäljning': 16520.0, 'Leverantör': 'BK 2025', 'Kod': 'BC', 'Kommentar': ''}, {'Radtyp': 'Material', 'Benämning': 'Markskiva cellplast 100mm', 'Byggdel': 'Grund', 'Mängd': 30.0, 'Enhet': 'm²', 'Timmar': 0.0, 'Á-pris': 222, 'Påslag %': 0.0, 'Kostnad': 6660.0, 'Försäljning': 6660.0, 'Leverantör': 'BK 2025', 'Kod': 'IBC', 'Kommentar': ''}, {'Radtyp': 'Material', 'Benämning': 'Betong C25/30 gjutning i hus', 'Byggdel': 'Grund', 'Mängd': 6.0, 'Enhet': 'm³', 'Timmar': 0.0, 'Á-pris': 2835, 'Påslag %': 0.0, 'Kostnad': 17010.0, 'Försäljning': 17010.0, 'Leverantör': 'BK 2025', 'Kod': 'ES', 'Kommentar': ''}, {'Radtyp': 'Material', 'Benämning': 'Armering B500 ø8-12mm', 'Byggdel': 'Grund', 'Mängd': 280.0, 'Enhet': 'kg', 'Timmar': 0.0, 'Á-pris': 37, 'Påslag %': 0.0, 'Kostnad': 10360.0, 'Försäljning': 10360.0, 'Leverantör': 'BK 2025', 'Kod': 'ESC', 'Kommentar': ''}, {'Radtyp': 'Arbete', 'Benämning': 'Gjutning grund inkl form', 'Byggdel': 'Grund', 'Mängd': 0.0, 'Enhet': 'tim', 'Timmar': 24.0, 'Á-pris': 595, 'Påslag %': 0.0, 'Kostnad': 14280.0, 'Försäljning': 14280.0, 'Leverantör': 'Eget', 'Kod': '', 'Kommentar': ''}, {'Radtyp': 'Material', 'Benämning': 'Ytterväggsregel 45x170 c600', 'Byggdel': 'Stomme', 'Mängd': 55.0, 'Enhet': 'm²', 'Timmar': 0.0, 'Á-pris': 344, 'Påslag %': 0.0, 'Kostnad': 18920.0, 'Försäljning': 18920.0, 'Leverantör': 'BK 2025', 'Kod': 'HSD.11', 'Kommentar': ''}, {'Radtyp': 'Material', 'Benämning': 'Bjälkar 45x195 c600 bjälklag', 'Byggdel': 'Stomme', 'Mängd': 30.0, 'Enhet': 'm²', 'Timmar': 0.0, 'Á-pris': 322, 'Påslag %': 0.0, 'Kostnad': 9660.0, 'Försäljning': 9660.0, 'Leverantör': 'BK 2025', 'Kod': 'HSD.12', 'Kommentar': ''}, {'Radtyp': 'Material', 'Benämning': 'Innerväggsregel 45x95 c600', 'Byggdel': 'Stomme', 'Mängd': 18.0, 'Enhet': 'm²', 'Timmar': 0.0, 'Á-pris': 243, 'Påslag %': 0.0, 'Kostnad': 4374.0, 'Försäljning': 4374.0, 'Leverantör': 'BK 2025', 'Kod': 'HSD.11', 'Kommentar': ''}, {'Radtyp': 'Arbete', 'Benämning': 'Resning och montering stomme', 'Byggdel': 'Stomme', 'Mängd': 0.0, 'Enhet': 'tim', 'Timmar': 40.0, 'Á-pris': 595, 'Påslag %': 0.0, 'Kostnad': 23800.0, 'Försäljning': 23800.0, 'Leverantör': 'Eget', 'Kod': '', 'Kommentar': ''}, {'Radtyp': 'Material', 'Benämning': 'Takstol fackverkstol c900 spw 8-12m', 'Byggdel': 'Tak', 'Mängd': 30.0, 'Enhet': 'm²', 'Timmar': 0.0, 'Á-pris': 338, 'Påslag %': 0.0, 'Kostnad': 10140.0, 'Försäljning': 10140.0, 'Leverantör': 'BK 2025', 'Kod': 'HSD.131', 'Kommentar': ''}, {'Radtyp': 'Material', 'Benämning': 'Inbrädning yttertak 19mm råspontad', 'Byggdel': 'Tak', 'Mängd': 33.0, 'Enhet': 'm²', 'Timmar': 0.0, 'Á-pris': 309, 'Påslag %': 0.0, 'Kostnad': 10197.0, 'Försäljning': 10197.0, 'Leverantör': 'BK 2025', 'Kod': 'HSD.133', 'Kommentar': ''}, {'Radtyp': 'Material', 'Benämning': 'Tegelpannor betong normalformat', 'Byggdel': 'Tak', 'Mängd': 33.0, 'Enhet': 'm²', 'Timmar': 0.0, 'Á-pris': 520, 'Påslag %': 0.0, 'Kostnad': 17160.0, 'Försäljning': 17160.0, 'Leverantör': 'BK 2025', 'Kod': 'JSD.1', 'Kommentar': ''}, {'Radtyp': 'Arbete', 'Benämning': 'Takläggning', 'Byggdel': 'Tak', 'Mängd': 0.0, 'Enhet': 'tim', 'Timmar': 20.0, 'Á-pris': 595, 'Påslag %': 0.0, 'Kostnad': 11900.0, 'Försäljning': 11900.0, 'Leverantör': 'Eget', 'Kod': '', 'Kommentar': ''}, {'Radtyp': 'Material', 'Benämning': 'Takränna plåt ø125mm', 'Byggdel': 'Tak', 'Mängd': 8.0, 'Enhet': 'm', 'Timmar': 0.0, 'Á-pris': 245, 'Påslag %': 0.0, 'Kostnad': 1960.0, 'Försäljning': 1960.0, 'Leverantör': 'BK 2025', 'Kod': 'JSD.5', 'Kommentar': ''}, {'Radtyp': 'Material', 'Benämning': 'Stuprör plåt ø87mm', 'Byggdel': 'Tak', 'Mängd': 5.0, 'Enhet': 'm', 'Timmar': 0.0, 'Á-pris': 198, 'Påslag %': 0.0, 'Kostnad': 990.0, 'Försäljning': 990.0, 'Leverantör': 'BK 2025', 'Kod': 'JSD.5', 'Kommentar': ''}, {'Radtyp': 'Material', 'Benämning': 'Mineralull yttervägg 145mm KL0.037', 'Byggdel': 'Fasad', 'Mängd': 55.0, 'Enhet': 'm²', 'Timmar': 0.0, 'Á-pris': 165, 'Påslag %': 0.0, 'Kostnad': 9075.0, 'Försäljning': 9075.0, 'Leverantör': 'BK 2025', 'Kod': 'IBE', 'Kommentar': ''}, {'Radtyp': 'Material', 'Benämning': 'Fasadpanel lockläkt 22x170', 'Byggdel': 'Fasad', 'Mängd': 55.0, 'Enhet': 'm²', 'Timmar': 0.0, 'Á-pris': 787, 'Påslag %': 0.0, 'Kostnad': 43285.0, 'Försäljning': 43285.0, 'Leverantör': 'BK 2025', 'Kod': 'HSD.16', 'Kommentar': ''}, {'Radtyp': 'Arbete', 'Benämning': 'Montering fasad inkl isolering', 'Byggdel': 'Fasad', 'Mängd': 0.0, 'Enhet': 'tim', 'Timmar': 36.0, 'Á-pris': 595, 'Påslag %': 0.0, 'Kostnad': 21420.0, 'Försäljning': 21420.0, 'Leverantör': 'Eget', 'Kod': '', 'Kommentar': ''}, {'Radtyp': 'Material', 'Benämning': 'Fönster trä/alu 10x12 3-glas', 'Byggdel': 'Fasad', 'Mängd': 3.0, 'Enhet': 'st', 'Timmar': 0.0, 'Á-pris': 10800, 'Påslag %': 0.0, 'Kostnad': 32400.0, 'Försäljning': 32400.0, 'Leverantör': 'BK 2025', 'Kod': 'PSC', 'Kommentar': ''}, {'Radtyp': 'Arbete', 'Benämning': 'Montering fönster', 'Byggdel': 'Fasad', 'Mängd': 0.0, 'Enhet': 'tim', 'Timmar': 9.0, 'Á-pris': 595, 'Påslag %': 0.0, 'Kostnad': 5355.0, 'Försäljning': 5355.0, 'Leverantör': 'Eget', 'Kod': '', 'Kommentar': ''}, {'Radtyp': 'Material', 'Benämning': 'Ytterdörr stål isolerad 9x21', 'Byggdel': 'Fasad', 'Mängd': 1.0, 'Enhet': 'st', 'Timmar': 0.0, 'Á-pris': 12500, 'Påslag %': 0.0, 'Kostnad': 12500.0, 'Försäljning': 12500.0, 'Leverantör': 'BK 2025', 'Kod': 'PSB', 'Kommentar': ''}, {'Radtyp': 'Material', 'Benämning': 'Mineralull bjälklag 145mm KL0.036', 'Byggdel': 'Isolering', 'Mängd': 30.0, 'Enhet': 'm²', 'Timmar': 0.0, 'Á-pris': 209, 'Påslag %': 0.0, 'Kostnad': 6270.0, 'Försäljning': 6270.0, 'Leverantör': 'BK 2025', 'Kod': 'IBF', 'Kommentar': ''}, {'Radtyp': 'Arbete', 'Benämning': 'Isolering invändig inkl ångspärr', 'Byggdel': 'Isolering', 'Mängd': 0.0, 'Enhet': 'tim', 'Timmar': 16.0, 'Á-pris': 595, 'Påslag %': 0.0, 'Kostnad': 9520.0, 'Försäljning': 9520.0, 'Leverantör': 'Eget', 'Kod': '', 'Kommentar': ''}, {'Radtyp': 'Material', 'Benämning': 'Spånskiva golvskiva 22mm', 'Byggdel': 'Invändigt', 'Mängd': 30.0, 'Enhet': 'm²', 'Timmar': 0.0, 'Á-pris': 198, 'Påslag %': 0.0, 'Kostnad': 5940.0, 'Försäljning': 5940.0, 'Leverantör': 'BK 2025', 'Kod': 'NSS.5', 'Kommentar': ''}, {'Radtyp': 'Material', 'Benämning': 'Laminat 12mm AC5', 'Byggdel': 'Invändigt', 'Mängd': 28.0, 'Enhet': 'm²', 'Timmar': 0.0, 'Á-pris': 595, 'Påslag %': 0.0, 'Kostnad': 16660.0, 'Försäljning': 16660.0, 'Leverantör': 'BK 2025', 'Kod': 'NSS.2', 'Kommentar': ''}, {'Radtyp': 'Material', 'Benämning': 'Innerdörr mdf 9x21 komplett', 'Byggdel': 'Invändigt', 'Mängd': 2.0, 'Enhet': 'st', 'Timmar': 0.0, 'Á-pris': 3200, 'Påslag %': 0.0, 'Kostnad': 6400.0, 'Försäljning': 6400.0, 'Leverantör': 'BK 2025', 'Kod': 'PSB', 'Kommentar': ''}, {'Radtyp': 'Arbete', 'Benämning': 'Invändig finish, lister, målning', 'Byggdel': 'Invändigt', 'Mängd': 0.0, 'Enhet': 'tim', 'Timmar': 32.0, 'Á-pris': 595, 'Påslag %': 0.0, 'Kostnad': 19040.0, 'Försäljning': 19040.0, 'Leverantör': 'Eget', 'Kod': '', 'Kommentar': ''}, {'Radtyp': 'UE', 'Benämning': 'El: komplett ny elinstallation', 'Byggdel': 'El & VVS', 'Mängd': 1.0, 'Enhet': 'ls', 'Timmar': 0.0, 'Á-pris': 28000, 'Påslag %': 0.0, 'Kostnad': 28000.0, 'Försäljning': 28000.0, 'Leverantör': 'EL AB', 'Kod': '', 'Kommentar': 'Fast pris inkl material'}, {'Radtyp': 'UE', 'Benämning': 'VVS: radiator, rör, koppling', 'Byggdel': 'El & VVS', 'Mängd': 1.0, 'Enhet': 'ls', 'Timmar': 0.0, 'Á-pris': 22000, 'Påslag %': 0.0, 'Kostnad': 22000.0, 'Försäljning': 22000.0, 'Leverantör': 'VVS AB', 'Kod': '', 'Kommentar': 'Fast pris inkl material'}]}, {'projektnamn': 'Fasadrenovering Solhemsgatan 14-18', 'projektnummer': '2025-003', 'kund': 'Brf Solhem', 'bestallar': 'Styrelsen Brf Solhem', 'adress': 'Solhemsgatan 14-18, 163 41 Spånga', 'datum': '2025-05-01', 'kalkylansvarig': 'Jan Stefors', 'status': 'Kalkyl', 'kommentar': 'Fasadrenovering 3 huskroppar. Byte av puts, tilläggsisolering 70mm, ny ädelputs och fönsterbyte. Ca 1 200m² fasad. Upphandlas som totalentreprenad.', 'byggdelar': ['Ställning', 'Rivning', 'Isolering', 'Puts', 'Fönster', 'Målning', 'Mark'], 'omkostnader': {'Arbetsledning': 45000.0, 'Etablering': 25000.0, 'Administration': 18000.0, 'Transporter': 12000.0, 'Garanti': 15000.0, 'Risk': 0.0, 'Övrigt': 5000.0}, 'paslag': {'Omkostnad %': 12.0, 'Risk %': 6.0, 'Vinst %': 9.0, 'Rabatt %': 2.0}, 'rader': [{'Radtyp': 'UE', 'Benämning': 'Fasadställning hyra 3 mån inkl montage', 'Byggdel': 'Ställning', 'Mängd': 1200.0, 'Enhet': 'm²', 'Timmar': 0.0, 'Á-pris': 185, 'Påslag %': 0.0, 'Kostnad': 222000.0, 'Försäljning': 222000.0, 'Leverantör': 'Ställnings AB', 'Kod': '', 'Kommentar': 'Inkl montering och demontering'}, {'Radtyp': 'UE', 'Benämning': 'Borthuggning befintlig puts klass B', 'Byggdel': 'Rivning', 'Mängd': 1200.0, 'Enhet': 'm²', 'Timmar': 0.0, 'Á-pris': 148, 'Påslag %': 0.0, 'Kostnad': 177600.0, 'Försäljning': 177600.0, 'Leverantör': 'Rivning AB', 'Kod': '', 'Kommentar': ''}, {'Radtyp': 'UE', 'Benämning': 'Borttransport rivmassor container', 'Byggdel': 'Rivning', 'Mängd': 1.0, 'Enhet': 'ls', 'Timmar': 0.0, 'Á-pris': 42000, 'Påslag %': 0.0, 'Kostnad': 42000.0, 'Försäljning': 42000.0, 'Leverantör': 'Rivning AB', 'Kod': '', 'Kommentar': ''}, {'Radtyp': 'Material', 'Benämning': 'Mineralull yttervägg 70mm KL0.037', 'Byggdel': 'Isolering', 'Mängd': 1200.0, 'Enhet': 'm²', 'Timmar': 0.0, 'Á-pris': 101, 'Påslag %': 0.0, 'Kostnad': 121200.0, 'Försäljning': 121200.0, 'Leverantör': 'BK 2025', 'Kod': 'IBE', 'Kommentar': ''}, {'Radtyp': 'Arbete', 'Benämning': 'Montering mineralull inkl mekaniska fästen', 'Byggdel': 'Isolering', 'Mängd': 0.0, 'Enhet': 'tim', 'Timmar': 180.0, 'Á-pris': 595, 'Påslag %': 0.0, 'Kostnad': 107100.0, 'Försäljning': 107100.0, 'Leverantör': 'Eget', 'Kod': '', 'Kommentar': ''}, {'Radtyp': 'Material', 'Benämning': 'Grundning + stänkputs utvändigt', 'Byggdel': 'Puts', 'Mängd': 1200.0, 'Enhet': 'm²', 'Timmar': 0.0, 'Á-pris': 271, 'Påslag %': 0.0, 'Kostnad': 325200.0, 'Försäljning': 325200.0, 'Leverantör': 'BK 2025', 'Kod': 'LBS', 'Kommentar': ''}, {'Radtyp': 'Material', 'Benämning': 'Grundning + grovputs + stänkputs fin', 'Byggdel': 'Puts', 'Mängd': 1200.0, 'Enhet': 'm²', 'Timmar': 0.0, 'Á-pris': 625, 'Påslag %': 0.0, 'Kostnad': 750000.0, 'Försäljning': 750000.0, 'Leverantör': 'BK 2025', 'Kod': 'LBS', 'Kommentar': ''}, {'Radtyp': 'Material', 'Benämning': 'Ädelputsfasad inkl stålnät', 'Byggdel': 'Puts', 'Mängd': 1200.0, 'Enhet': 'm²', 'Timmar': 0.0, 'Á-pris': 1280, 'Påslag %': 0.0, 'Kostnad': 1536000.0, 'Försäljning': 1536000.0, 'Leverantör': 'BK 2025', 'Kod': 'LBS', 'Kommentar': ''}, {'Radtyp': 'Arbete', 'Benämning': 'Puts och ytbehandling fasad', 'Byggdel': 'Puts', 'Mängd': 0.0, 'Enhet': 'tim', 'Timmar': 200.0, 'Á-pris': 595, 'Påslag %': 0.0, 'Kostnad': 119000.0, 'Försäljning': 119000.0, 'Leverantör': 'Eget', 'Kod': '', 'Kommentar': ''}, {'Radtyp': 'Material', 'Benämning': 'Fönster trä/alu 10x12 2-glas', 'Byggdel': 'Fönster', 'Mängd': 48.0, 'Enhet': 'st', 'Timmar': 0.0, 'Á-pris': 8900, 'Påslag %': 0.0, 'Kostnad': 427200.0, 'Försäljning': 427200.0, 'Leverantör': 'BK 2025', 'Kod': 'PSC', 'Kommentar': 'Lägenheter'}, {'Radtyp': 'Material', 'Benämning': 'Fönster trä/alu 12x14 2-glas', 'Byggdel': 'Fönster', 'Mängd': 12.0, 'Enhet': 'st', 'Timmar': 0.0, 'Á-pris': 11500, 'Påslag %': 0.0, 'Kostnad': 138000.0, 'Försäljning': 138000.0, 'Leverantör': 'BK 2025', 'Kod': 'PSC', 'Kommentar': 'Trapphus och gavlar'}, {'Radtyp': 'Arbete', 'Benämning': 'Montering fönster inkl tätning', 'Byggdel': 'Fönster', 'Mängd': 0.0, 'Enhet': 'tim', 'Timmar': 120.0, 'Á-pris': 595, 'Påslag %': 0.0, 'Kostnad': 71400.0, 'Försäljning': 71400.0, 'Leverantör': 'Eget', 'Kod': '', 'Kommentar': ''}, {'Radtyp': 'Material', 'Benämning': 'Silikatfärg slät putsyta', 'Byggdel': 'Målning', 'Mängd': 1200.0, 'Enhet': 'm²', 'Timmar': 0.0, 'Á-pris': 90, 'Påslag %': 0.0, 'Kostnad': 108000.0, 'Försäljning': 108000.0, 'Leverantör': 'BK 2025', 'Kod': 'LCS', 'Kommentar': '2 strykningar'}, {'Radtyp': 'Arbete', 'Benämning': 'Målning fasad', 'Byggdel': 'Målning', 'Mängd': 0.0, 'Enhet': 'tim', 'Timmar': 160.0, 'Á-pris': 595, 'Påslag %': 0.0, 'Kostnad': 95200.0, 'Försäljning': 95200.0, 'Leverantör': 'Eget', 'Kod': '', 'Kommentar': ''}, {'Radtyp': 'Material', 'Benämning': 'Sockelputs stålslipat', 'Byggdel': 'Målning', 'Mängd': 120.0, 'Enhet': 'm²', 'Timmar': 0.0, 'Á-pris': 617, 'Påslag %': 0.0, 'Kostnad': 74040.0, 'Försäljning': 74040.0, 'Leverantör': 'BK 2025', 'Kod': 'LBS', 'Kommentar': ''}, {'Radtyp': 'Material', 'Benämning': 'Gångbaneplattor betong 350x350x50mm', 'Byggdel': 'Mark', 'Mängd': 85.0, 'Enhet': 'm²', 'Timmar': 0.0, 'Á-pris': 975, 'Påslag %': 0.0, 'Kostnad': 82875.0, 'Försäljning': 82875.0, 'Leverantör': 'BK 2025', 'Kod': 'BC', 'Kommentar': 'Reparation efter ställning'}, {'Radtyp': 'UE', 'Benämning': 'Slutbesiktning och dokumentation', 'Byggdel': 'Mark', 'Mängd': 1.0, 'Enhet': 'ls', 'Timmar': 0.0, 'Á-pris': 18500, 'Påslag %': 0.0, 'Kostnad': 18500.0, 'Försäljning': 18500.0, 'Leverantör': 'Besiktnings AB', 'Kod': '', 'Kommentar': ''}]}]

# ──────────────────────────────────────────────────────────────
#  UTILITIES
# ──────────────────────────────────────────────────────────────
def sf(v, d=0.0):
    try: return float(str(v).replace(" ","").replace(",","."))
    except: return d

def kr(v):
    try: return f"{float(v):,.0f} kr".replace(","," ")
    except: return "0 kr"

def pct(v):
    try: return f"{float(v):.1f} %"
    except: return "0,0 %"

def empty_projekt():
    return {"projektnamn":"","projektnummer":"","kund":"","bestallar":"",
            "adress":"","datum":datetime.date.today().isoformat(),
            "kalkylansvarig":"","status":"Kalkyl","kommentar":"",
            "rader":[],"byggdelar":list(BYGGDEL_DEF),
            "omkostnader":{k:0.0 for k in OMKKOSTNADER},
            "paslag":{"Omkostnad %":10.0,"Risk %":5.0,"Vinst %":8.0,"Rabatt %":0.0}}

def empty_rad():
    return {"Radtyp":"Material","Benämning":"","Byggdel":"",
            "Mängd":0.0,"Enhet":"st","Timmar":0.0,
            "Á-pris":0.0,"Kostnad":0.0,"Påslag %":0.0,"Försäljning":0.0,
            "Leverantör":"","Kod":"","Kommentar":""}

def berakna(r):
    m=sf(r.get("Mängd",0)); t=sf(r.get("Timmar",0))
    a=sf(r.get("Á-pris",0)); p=sf(r.get("Påslag %",0))
    k=(t*a) if r.get("Radtyp")=="Arbete" else (m*a)
    r["Kostnad"]=round(k,2); r["Försäljning"]=round(k*(1+p/100),2); return r

def summera(proj):
    rader=proj.get("rader",[]); dir_k=sum(sf(r.get("Kostnad",0)) for r in rader)
    omk=proj.get("omkostnader",{}); omk_s=sum(sf(v) for v in omk.values())
    p=proj.get("paslag",{}); o=sf(p.get("Omkostnad %",0)); ri=sf(p.get("Risk %",0))
    v=sf(p.get("Vinst %",0)); d=sf(p.get("Rabatt %",0))
    omk_b=dir_k*o/100; ris_b=dir_k*ri/100; sjk=dir_k+omk_s+omk_b+ris_b
    vin_b=sjk*v/100; fp=sjk+vin_b; fp_n=fp-fp*d/100
    tb=fp_n-dir_k; mg=(tb/fp_n*100) if fp_n else 0
    return {"dir_k":dir_k,"omk_s":omk_s,"omk_b":omk_b,"ris_b":ris_b,
            "sjk":sjk,"vin_b":vin_b,"fp":fp_n,"tb":tb,"mg":mg}

def sync_active():
    """Spara aktivt projekt tillbaka till projektlistan."""
    idx  = st.session_state.get("aktivt_idx", 0)
    lista= st.session_state.get("projekt_lista", [])
    if lista and idx < len(lista):
        lista[idx] = st.session_state.projekt

def init():
    if "projekt_lista" not in st.session_state:
        st.session_state.projekt_lista = copy.deepcopy(EXEMPELPROJEKT)
        st.session_state.aktivt_idx    = 0
        st.session_state.projekt       = copy.deepcopy(EXEMPELPROJEKT[0])
    if "prisbank" not in st.session_state: st.session_state.prisbank=list(DEFAULT_PRISBANK)
    if "mallar"   not in st.session_state: st.session_state.mallar=[]

# ──────────────────────────────────────────────────────────────
#  CSS
# ──────────────────────────────────────────────────────────────
st.markdown("""
<style>
.stApp{background:#f4f6f9}
.block-container{padding-top:1rem;padding-bottom:1rem}
h1,h2,h3{color:#1a3a5c}
.section-hdr{background:#1a3a5c;color:white;padding:8px 14px;border-radius:6px;
             font-weight:600;margin-bottom:10px;font-size:1rem}
.info-box{background:white;border-left:4px solid #1a3a5c;padding:10px 14px;
          border-radius:4px;margin:6px 0}
div[data-testid="stMetricValue"]{font-size:1.3rem;font-weight:700;color:#1a3a5c}
div[data-testid="stMetricLabel"]{color:#6b7280;font-size:.85rem}
/* ── Flikar – tydligare ── */
.stTabs [data-baseweb="tab-list"]{gap:4px;border-bottom:2px solid #1a3a5c;padding-bottom:0}
.stTabs [data-baseweb="tab"]{background:#e8edf4;border-radius:6px 6px 0 0;
    padding:8px 16px;font-weight:600;font-size:.92rem;color:#1a3a5c;border:none}
.stTabs [aria-selected="true"]{background:#1a3a5c !important;color:white !important}
/* ── Utskrift ── */
@media print{
  section[data-testid="stSidebar"]{display:none}
  .stTabs [data-baseweb="tab-list"]{display:none}
  .block-container{padding:0}
  button,footer{display:none}
}
</style>""", unsafe_allow_html=True)

# ──────────────────────────────────────────────────────────────
#  SIDEBAR
# ──────────────────────────────────────────────────────────────
def sidebar():
    sync_active()
    proj =st.session_state.projekt
    lista=st.session_state.projekt_lista
    akt  =st.session_state.get("aktivt_idx",0)

    with st.sidebar:
        st.markdown("## 🏗 Kalkylprogram")
        st.caption("Bygg & Entreprenad  v2")
        st.divider()

        # ── NAVIGERING ────────────────────────────────────────
        st.markdown("**Välj sida**")
        st.radio("nav", [
            "🏠  Start",
            "📋  Projektinfo",
            "🔢  Kalkyl",
            "💰  Prisbank",
            "📑  Mallar",
            "🏗  Byggdelar",
            "📊  Slutsida",
        ], key="sida", label_visibility="collapsed")
        st.divider()

        # ── PROJEKT ───────────────────────────────────────────
        st.markdown("**Projekt**")
        namn_lista=[p.get("projektnamn","") or f"Projekt {i+1}"
                    for i,p in enumerate(lista)]
        valt=st.selectbox("Välj projekt",namn_lista,index=akt,
                          label_visibility="collapsed")
        ny_idx=namn_lista.index(valt)
        if ny_idx!=akt:
            lista[akt]=copy.deepcopy(proj)
            st.session_state.projekt_lista=lista
            st.session_state.aktivt_idx=ny_idx
            st.session_state.projekt=copy.deepcopy(lista[ny_idx])
            st.rerun()

        c1,c2=st.columns(2)
        if c1.button("➕ Nytt",use_container_width=True):
            lista[akt]=copy.deepcopy(proj)
            ny=empty_projekt()
            lista.append(ny)
            st.session_state.projekt_lista=lista
            st.session_state.aktivt_idx=len(lista)-1
            st.session_state.projekt=ny
            st.rerun()
        if c2.button("🗑 Ta bort",use_container_width=True):
            if len(lista)>1:
                lista.pop(akt)
                ny_akt=min(akt,len(lista)-1)
                st.session_state.projekt_lista=lista
                st.session_state.aktivt_idx=ny_akt
                st.session_state.projekt=copy.deepcopy(lista[ny_akt])
                st.rerun()
            else:
                st.session_state.projekt=empty_projekt()
                lista[0]=st.session_state.projekt
                st.rerun()

        up=st.file_uploader("📂 Öppna (.json)",type=["json"],
                             label_visibility="collapsed")
        if up:
            try:
                imp=json.loads(up.read().decode())
                lista[akt]=copy.deepcopy(proj)
                lista.append(imp)
                st.session_state.projekt_lista=lista
                st.session_state.aktivt_idx=len(lista)-1
                st.session_state.projekt=imp
                st.rerun()
            except: st.error("Kunde inte öppna filen.")

        name=(proj.get("projektnamn","projekt") or "projekt").replace(" ","_")
        st.download_button("💾 Spara projekt",
            data=json.dumps(proj,ensure_ascii=False,indent=2).encode(),
            file_name=f"{name}.json",mime="application/json",
            use_container_width=True)
        st.divider()

        # ── SUMMERING ─────────────────────────────────────────
        s=summera(proj)
        st.markdown(f"**{proj.get('projektnamn','') or '—'}**")
        st.caption(f"{proj.get('status','')}  ·  {len(lista)} projekt")
        c1,c2=st.columns(2)
        c1.metric("Rader", len(proj.get("rader",[])))
        c2.metric("Marginal", pct(s["mg"]))
        st.metric("Försäljning", kr(s["fp"]))
        st.metric("TB", kr(s["tb"]))
        st.divider()
        st.markdown("""<a href="javascript:window.print()" style="
          display:block;text-align:center;background:#1a3a5c;color:white;
          padding:8px 0;border-radius:6px;font-weight:600;font-size:.9rem;
          text-decoration:none">🖨 Skriv ut / Spara PDF</a>""",
          unsafe_allow_html=True)

# ──────────────────────────────────────────────────────────────
#  TAB START
# ──────────────────────────────────────────────────────────────
def tab_start():
    proj=st.session_state.projekt
    s=summera(proj)
    st.markdown('<div class="section-hdr">Projektöversikt</div>',unsafe_allow_html=True)
    c1,c2,c3,c4=st.columns(4)
    c1.metric("Direkta kostnader", kr(s["dir_k"]))
    c2.metric("Försäljningspris",  kr(s["fp"]))
    c3.metric("Täckningsbidrag",   kr(s["tb"]))
    c4.metric("Marginal",          pct(s["mg"]))
    st.divider()
    l,r=st.columns(2)
    with l:
        st.markdown("**Projektinfo**")
        for k,v in [("Projektnamn",proj.get("projektnamn","–")),
                    ("Projektnr",  proj.get("projektnummer","–")),
                    ("Kund",       proj.get("kund","–")),
                    ("Status",     proj.get("status","–")),
                    ("Ansvarig",   proj.get("kalkylansvarig","–")),
                    ("Datum",      proj.get("datum","–"))]:
            st.markdown(f'<div class="info-box"><b>{k}:</b> {v}</div>',unsafe_allow_html=True)
    with r:
        st.markdown("**Ekonomi**")
        df=pd.DataFrame([
            ("Direkta kostnader",   f"{s['dir_k']:,.0f}"),
            ("Omkostnader",         f"{s['omk_s']:,.0f}"),
            ("Självkostnad",        f"{s['sjk']:,.0f}"),
            ("Försäljningspris",    f"{s['fp']:,.0f}"),
            ("TB",                  f"{s['tb']:,.0f}"),
            ("Marginal",            f"{s['mg']:.1f} %"),
        ], columns=["Post","kr"])
        st.dataframe(df,hide_index=True,use_container_width=True)

    # ── Navigationsknappar ────────────────────────────────────
    st.divider()
    st.markdown("**Gå till**")
    st.markdown("""
<style>
.nav-grid{display:grid;grid-template-columns:repeat(3,1fr);gap:12px;margin-top:8px}
.nav-btn{background:#1a3a5c;color:white;border:none;border-radius:10px;
         padding:20px 10px;text-align:center;cursor:pointer;font-size:1rem;
         font-weight:600;text-decoration:none;display:block;transition:background .2s}
.nav-btn:hover{background:#2a5a8c;color:white}
.nav-icon{font-size:1.8rem;display:block;margin-bottom:6px}
</style>
<div class="nav-grid">
  <a class="nav-btn" href="?sida=Projektinfo" target="_self">
    <span class="nav-icon">📋</span>Projektinfo
  </a>
  <a class="nav-btn" href="?sida=Kalkyl" target="_self">
    <span class="nav-icon">🔢</span>Kalkyl
  </a>
  <a class="nav-btn" href="?sida=Prisbank" target="_self">
    <span class="nav-icon">💰</span>Prisbank
  </a>
  <a class="nav-btn" href="?sida=Mallar" target="_self">
    <span class="nav-icon">📑</span>Mallar
  </a>
  <a class="nav-btn" href="?sida=Byggdelar" target="_self">
    <span class="nav-icon">🏗</span>Byggdelar
  </a>
  <a class="nav-btn" href="?sida=Slutsida" target="_self">
    <span class="nav-icon">📊</span>Slutsida
  </a>
</div>""", unsafe_allow_html=True)
    # Streamlit-knappar som faktiskt fungerar för navigering
    st.write("")
    b1,b2,b3=st.columns(3)
    b4,b5,b6=st.columns(3)
    if b1.button("📋  Projektinfo",  use_container_width=True):
        st.session_state.sida="📋  Projektinfo";  st.rerun()
    if b2.button("🔢  Kalkyl",       use_container_width=True):
        st.session_state.sida="🔢  Kalkyl";       st.rerun()
    if b3.button("💰  Prisbank",     use_container_width=True):
        st.session_state.sida="💰  Prisbank";     st.rerun()
    if b4.button("📑  Mallar",       use_container_width=True):
        st.session_state.sida="📑  Mallar";       st.rerun()
    if b5.button("🏗  Byggdelar",    use_container_width=True):
        st.session_state.sida="🏗  Byggdelar";    st.rerun()
    if b6.button("📊  Slutsida",     use_container_width=True):
        st.session_state.sida="📊  Slutsida";     st.rerun()

# ──────────────────────────────────────────────────────────────
#  TAB PROJEKT
# ──────────────────────────────────────────────────────────────
def tab_projekt():
    proj=st.session_state.projekt
    st.markdown('<div class="section-hdr">Projektinformation</div>',unsafe_allow_html=True)
    with st.form("proj"):
        c1,c2=st.columns(2)
        with c1:
            proj["projektnamn"]   =st.text_input("Projektnamn *",   proj.get("projektnamn",""))
            proj["projektnummer"] =st.text_input("Projektnummer",   proj.get("projektnummer",""))
            proj["kund"]          =st.text_input("Kund",            proj.get("kund",""))
            proj["bestallar"]     =st.text_input("Beställare",      proj.get("bestallar",""))
        with c2:
            proj["adress"]        =st.text_input("Adress",          proj.get("adress",""))
            proj["datum"]         =st.text_input("Datum",           proj.get("datum",""))
            proj["kalkylansvarig"]=st.text_input("Kalkylansvarig",  proj.get("kalkylansvarig",""))
            idx=STATUS_LIST.index(proj.get("status","Kalkyl")) if proj.get("status") in STATUS_LIST else 0
            proj["status"]        =st.selectbox("Status", STATUS_LIST, index=idx)
        proj["kommentar"]=st.text_area("Kommentar", proj.get("kommentar",""))
        if st.form_submit_button("💾 Spara", type="primary"):
            st.session_state.projekt=proj; st.success("Sparat.")
    st.divider()
    st.markdown("**Byggdelar** – en per rad")
    txt=st.text_area("","\n".join(proj.get("byggdelar",BYGGDEL_DEF)),height=180,
                     label_visibility="collapsed")
    if st.button("Uppdatera byggdelar"):
        proj["byggdelar"]=[b.strip() for b in txt.splitlines() if b.strip()]
        st.session_state.projekt=proj; st.success("Klart.")

# ──────────────────────────────────────────────────────────────
#  TAB KALKYL  (omdesignad)
# ──────────────────────────────────────────────────────────────
def tab_kalkyl():
    proj =st.session_state.projekt
    rader=proj.get("rader",[])
    bdlar=proj.get("byggdelar",BYGGDEL_DEF)
    pb   =st.session_state.prisbank

    st.markdown('<div class="section-hdr">Kalkyl</div>',unsafe_allow_html=True)

    # ── Lägg till rad – tre sätt ──────────────────────────────
    with st.expander("➕  Lägg till rad", expanded=len(rader)==0):
        mode=st.radio("Lägg till via:",
                      ["Prisbank","Tom rad","Kopiera från mall"],
                      horizontal=True, label_visibility="collapsed")

        if mode=="Prisbank":
            if not pb:
                st.info("Prisbanken är tom. Gå till fliken **Prisbank** för att lägga till artiklar.")
            else:
                st.caption("💡 Priser från **Prisbank** (BK 2025 + egna). Mat-pris = komplett á-pris inkl arbete. Välj Typ=UE för underentreprenörsarbeten.")
                c1,c2,c3,c4,c5=st.columns([3,1,1,1,1])
                sok=c1.text_input("Sök artikel", placeholder="Namn eller kod...",
                                  label_visibility="collapsed")
                träffar=[p for p in pb if not sok or sok.lower() in p.get("benamning","").lower()
                         or sok.lower() in p.get("kod","").lower()]
                if träffar:
                    vald=c1.selectbox("Artikel",[f"{p.get('kod','')+' – ' if p.get('kod') else ''}{p['benamning']} ({p.get('enhet','st')} | mat:{p.get('matpris',0):.0f} | arb:{p.get('arbpris',0):.0f})"
                                                  for p in träffar],
                                       label_visibility="collapsed")
                    idx_v=0
                    for i,p in enumerate(träffar):
                        lbl=f"{p.get('kod','')+' – ' if p.get('kod') else ''}{p['benamning']} ({p.get('enhet','st')} | mat:{p.get('matpris',0):.0f} | arb:{p.get('arbpris',0):.0f})"
                        if lbl==vald: idx_v=i; break
                    item=träffar[idx_v]
                    radtyp=c2.selectbox("Typ", RADTYPER, label_visibility="collapsed")
                    mangd=c3.number_input("Mängd",value=1.0,min_value=0.0,
                                          label_visibility="collapsed")
                    bdl=c4.selectbox("Byggdel",["–"]+bdlar,
                                      label_visibility="collapsed")
                    paslag=c5.number_input("Påslag %",value=0.0,min_value=0.0,
                                            label_visibility="collapsed")
                    if st.button("Lägg till i kalkyl ➜", type="primary"):
                        apris=sf(item.get("arbpris",0)) if radtyp=="Arbete" else sf(item.get("matpris",0))
                        ny=empty_rad()
                        ny.update({"Radtyp":radtyp,"Benämning":item["benamning"],
                                   "Kod":item.get("kod",""),"Enhet":item.get("enhet","st"),
                                   "Á-pris":apris,"Mängd":mangd,"Påslag %":paslag,
                                   "Byggdel":bdl if bdl!="–" else "",
                                   "Leverantör":item.get("leverantor","")})
                        if radtyp=="Arbete": ny["Timmar"]=mangd; ny["Mängd"]=0
                        berakna(ny)
                        proj["rader"].append(ny)
                        st.session_state.projekt=proj; st.rerun()
                else:
                    st.warning("Inga träffar i prisbanken.")

        elif mode=="Tom rad":
            with st.form("ny_rad"):
                c1,c2,c3=st.columns(3)
                typ =c1.selectbox("Typ",RADTYPER)
                ben =c1.text_input("Benämning *")
                bdl =c1.selectbox("Byggdel",["–"]+bdlar)
                enh =c2.selectbox("Enhet",ENHETER)
                mng =c2.number_input("Mängd",value=0.0,min_value=0.0)
                tim =c2.number_input("Timmar",value=0.0,min_value=0.0)
                apr =c3.number_input("Á-pris (kr)",value=0.0,min_value=0.0)
                pas =c3.number_input("Påslag %",value=0.0,min_value=0.0)
                lev =c3.text_input("Leverantör")
                if st.form_submit_button("Lägg till ➜", type="primary"):
                    if not ben: st.warning("Ange benämning.")
                    else:
                        ny={"Radtyp":typ,"Benämning":ben,"Byggdel":bdl if bdl!="–" else "",
                            "Mängd":mng,"Enhet":enh,"Timmar":tim,"Á-pris":apr,
                            "Påslag %":pas,"Leverantör":lev,"Kod":"","Kommentar":"",
                            "Kostnad":0.0,"Försäljning":0.0}
                        berakna(ny); proj["rader"].append(ny)
                        st.session_state.projekt=proj; st.rerun()

        else:  # Mall
            mallar=st.session_state.mallar
            if not mallar:
                st.info("Inga mallar sparade. Gå till fliken **Mallar**.")
            else:
                vald=st.selectbox("Välj mall",[m["namn"] for m in mallar])
                m=next((x for x in mallar if x["namn"]==vald),None)
                if m:
                    st.caption(f"{len(m.get('rader',[]))} rader i mallen")
                    if st.button(f"Lägg till '{vald}' ➜", type="primary"):
                        for r in m.get("rader",[]):
                            ny=copy.deepcopy(r); berakna(ny)
                            proj["rader"].append(ny)
                        st.session_state.projekt=proj; st.rerun()

    # ── Filter ───────────────────────────────────────────────
    fc1,fc2,fc3,fc4=st.columns([2,2,2,1])
    f_bdl=fc1.selectbox("Byggdel",["Alla"]+bdlar,label_visibility="visible")
    f_typ=fc2.selectbox("Typ",["Alla"]+RADTYPER,label_visibility="visible")
    f_sok=fc3.text_input("Sök",placeholder="Sök benämning...",label_visibility="visible")
    if fc4.button("✕ Rensa",use_container_width=True): st.rerun()

    visa=[r for r in rader
          if (f_bdl=="Alla" or r.get("Byggdel","")==f_bdl)
          and (f_typ=="Alla" or r.get("Radtyp","")==f_typ)
          and (not f_sok or f_sok.lower() in r.get("Benämning","").lower())]

    # ── Kalkylbord – smala kolumner ───────────────────────────
    if visa:
        df=pd.DataFrame(visa)
        # Ensure all columns exist
        for col in empty_rad():
            if col not in df.columns: df[col]=None

        edited=st.data_editor(
            df[["Radtyp","Benämning","Byggdel","Mängd","Enhet","Timmar",
                "Á-pris","Kostnad","Påslag %","Försäljning"]],
            column_config={
                "Radtyp":     st.column_config.SelectboxColumn("Typ",    options=RADTYPER,width="small"),
                "Benämning":  st.column_config.TextColumn("Benämning",   width="medium"),
                "Byggdel":    st.column_config.SelectboxColumn("Byggdel",options=["–"]+bdlar,width="small"),
                "Mängd":      st.column_config.NumberColumn("Mängd",     format="%.2f",width="small"),
                "Enhet":      st.column_config.SelectboxColumn("Enhet",  options=ENHETER,width="small"),
                "Timmar":     st.column_config.NumberColumn("Tim",       format="%.1f",width="small"),
                "Á-pris":     st.column_config.NumberColumn("Á-pris",    format="%.0f",width="small"),
                "Kostnad":    st.column_config.NumberColumn("Kostnad",   format="%.0f",width="small",disabled=True),
                "Påslag %":   st.column_config.NumberColumn("Pål%",      format="%.1f",width="small"),
                "Försäljning":st.column_config.NumberColumn("Försälj.",  format="%.0f",width="small",disabled=True),
            },
            use_container_width=True, hide_index=True,
            num_rows="dynamic", height=min(60+len(visa)*35, 500),
            key="kalkyl_ed"
        )

        c1,c2,c3=st.columns([2,1,1])
        if c1.button("✓  Beräkna & spara", type="primary"):
            ny_rader=[]
            for _,row in edited.iterrows():
                r=row.to_dict()
                if r.get("Byggdel")=="–": r["Byggdel"]=""
                berakna(r)
                ny_rader.append(r)
            if f_bdl=="Alla" and f_typ=="Alla" and not f_sok:
                proj["rader"]=ny_rader
            else:
                orig_ids={id(r):i for i,r in enumerate(rader)}
                other=[r for r in rader if r not in visa]
                proj["rader"]=other+ny_rader
            st.session_state.projekt=proj; st.success("Kalkyl sparad."); st.rerun()

        if c2.button("💾  Spara som mall"):
            namn=st.session_state.get("_mall_namn","")
            if visa:
                st.session_state["_visa_mall_input"]=True
        if st.session_state.get("_visa_mall_input"):
            namn=st.text_input("Mallnamn",key="mall_namn_inp")
            if st.button("Spara mall ➜"):
                if namn:
                    st.session_state.mallar.append(
                        {"namn":namn,"rader":copy.deepcopy(visa),
                         "skapad":datetime.date.today().isoformat()})
                    st.session_state["_visa_mall_input"]=False
                    st.success(f"Mall '{namn}' sparad."); st.rerun()
    else:
        st.info("Inga kalkylrader ännu. Lägg till rader via panelen ovan.")

    # ── Summering ─────────────────────────────────────────────
    if rader:
        st.divider()
        tot_k=sum(sf(r.get("Kostnad",0)) for r in rader)
        tot_f=sum(sf(r.get("Försäljning",0)) for r in rader)
        m1,m2,m3,m4=st.columns(4)
        m1.metric("Rader",len(rader))
        m2.metric("Direkta kostnader",kr(tot_k))
        m3.metric("Försäljning",kr(tot_f))
        m4.metric("TB",kr(tot_f-tot_k))

    # ── Export ────────────────────────────────────────────────
    if rader:
        with st.expander("📤 Exportera"):
            xc1,xc2=st.columns(2)
            if OPX_OK:
                buf=io.BytesIO(); _xl_kalkyl(proj,buf); buf.seek(0)
                xc1.download_button("📊 Excel-kalkyl",data=buf,
                    file_name=f"{name_safe(proj)}_kalkyl.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            if RL_OK:
                buf=io.BytesIO(); _pdf_kalkyl(proj,buf); buf.seek(0)
                xc2.download_button("📄 PDF-kalkyl",data=buf,
                    file_name=f"{name_safe(proj)}_kalkyl.pdf",
                    mime="application/pdf")

# ──────────────────────────────────────────────────────────────
#  TAB PRISBANK  (omdesignad)
# ──────────────────────────────────────────────────────────────
def tab_prisbank():
    st.markdown('<div class="section-hdr">Prisbank – artiklar & à-priser</div>',
                unsafe_allow_html=True)

    pb=st.session_state.prisbank

    # ── Tre sektioner ─────────────────────────────────────────
    t1,t2,t3=st.tabs(["📋  Alla artiklar","➕  Lägg till manuellt","📥  Importera Excel"])

    # ── 1. Visa och redigera alla priser ──────────────────────
    with t1:
        if not pb:
            st.info("Prisbanken är tom. Importera en prislista eller lägg till artiklar manuellt.")
            if st.button("🔄 Återställ BK 2025 standardpriser", type="primary"):
                st.session_state.prisbank=list(DEFAULT_PRISBANK); st.rerun()
        else:
            fc1,fc2=st.columns([2,1])
            sok=fc1.text_input("🔍 Sök",placeholder="Namn eller kod...",label_visibility="collapsed")
            kat=fc2.selectbox("Kategori",["Alla"]+KATEGORIER,label_visibility="collapsed")
            träffar=[p for p in pb if
                (not sok or sok.lower() in p.get("benamning","").lower() or sok.lower() in p.get("kod","").lower())
                and (kat=="Alla" or p.get("kategori","")==kat)]

            st.caption(f"📋 {len(träffar)} av {len(pb)} artiklar  |  Källa: BK 2025 (Byggmästarnas Kostnadskalkylator)")

            # Ensure kategori field exists
            for p in pb:
                if "kategori" not in p: p["kategori"]="Övrigt"

            df=pd.DataFrame(träffar if träffar else [{"kod":"","benamning":"","enhet":"st","matpris":0,"arbpris":0,"leverantor":"BK 2025","kategori":"Övrigt"}])
            cols_show=["kategori","kod","benamning","enhet","matpris","arbpris","leverantor"]
            for c in cols_show:
                if c not in df.columns: df[c]=""
            edited=st.data_editor(
                df[cols_show],
                column_config={
                    "kategori":   st.column_config.SelectboxColumn("Kategori",  options=KATEGORIER,width="small"),
                    "kod":        st.column_config.TextColumn("Kod",            width="small"),
                    "benamning":  st.column_config.TextColumn("Benämning",      width="medium"),
                    "enhet":      st.column_config.SelectboxColumn("Enhet",     options=ENHETER,width="small"),
                    "matpris":    st.column_config.NumberColumn("Mat-pris",     format="%.0f",width="small",
                                     help="Komplett á-pris enl BK 2025 (material+arbete). Används som UE- eller materialpris."),
                    "arbpris":    st.column_config.NumberColumn("Arb-pris",     format="%.0f",width="small",
                                     help="Uppskattad arbetskostnad kr/enhet (tim × 495 kr)"),
                    "leverantor": st.column_config.TextColumn("Källa",          width="small"),
                },
                use_container_width=True, hide_index=True,
                num_rows="dynamic", height=400, key="pb_ed"
            )
            c1,c2=st.columns([2,1])
            if c1.button("💾  Spara ändringar", type="primary"):
                ny=[r for r in edited.to_dict("records")
                    if r.get("benamning","").strip()]
                if not sok:
                    st.session_state.prisbank=ny
                else:
                    others=[p for p in pb if not any(
                        p.get("benamning")==t.get("benamning") for t in träffar)]
                    st.session_state.prisbank=others+ny
                st.success(f"Prisbank sparad – {len(st.session_state.prisbank)} artiklar.")
                st.rerun()

            buf=io.BytesIO()
            pd.DataFrame(st.session_state.prisbank).to_excel(buf,index=False)
            buf.seek(0)
            c2.download_button("⬇ Ladda ner Excel",data=buf,
                file_name="prisbank.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        # BKI
        st.divider()
        st.markdown("**BKI-indexomräkning**")
        bki_c1,bki_c2,bki_c3,bki_c4=st.columns(4)
        bki_typ=bki_c1.selectbox("Typ",list(BKI.keys()),key="bki_t")
        bki_bas=bki_c2.selectbox("Basår",list(BKI[bki_typ].keys()),index=2,key="bki_b")
        bki_ny =bki_c3.selectbox("Till år",list(BKI[bki_typ].keys()),index=7,key="bki_n")
        try:
            fak=BKI[bki_typ][bki_ny]/BKI[bki_typ][bki_bas]
            bki_c4.metric("Index-faktor",f"{fak:.4f}")
        except: pass
        if pb and st.button("Räkna om alla priser med faktorn"):
            try:
                for p in st.session_state.prisbank:
                    p["matpris"]=round(sf(p.get("matpris",0))*fak,2)
                    p["arbpris"]=round(sf(p.get("arbpris",0))*fak,2)
                st.success(f"Alla priser räknade om med faktor {fak:.4f}"); st.rerun()
            except: st.error("Fyll i BKI-inställningarna ovan.")

    # ── 2. Lägg till manuellt ─────────────────────────────────
    with t2:
        st.markdown("**Lägg till egen artikel:**")
        st.caption("Mat-pris = komplett á-pris (material+arbete) för UE/material. Arb-pris = om du vill särskilja enbart arbetskostnaden.")
        with st.form("ny_artikel"):
            c1,c2,c3=st.columns(3)
            kat =c1.selectbox("Kategori",KATEGORIER)
            kod =c1.text_input("Kod (valfri)")
            ben =c1.text_input("Benämning *")
            enh =c1.selectbox("Enhet",ENHETER)
            mat =c2.number_input("Mat-pris (kr/enhet)",value=0.0,
                                  help="Komplett á-pris inkl material och arbete")
            arb =c2.number_input("Arb-pris (kr/enhet)",value=0.0,
                                  help="Enbart arbetskostnad per enhet")
            lev =c3.text_input("Källa/leverantör")
            komm=c3.text_area("Notering",height=80)
            if st.form_submit_button("➕ Lägg till i prisbank", type="primary"):
                if not ben: st.warning("Ange benämning.")
                else:
                    st.session_state.prisbank.append(
                        {"kod":kod,"benamning":ben,"enhet":enh,
                         "matpris":mat,"arbpris":arb,
                         "leverantor":lev,"kategori":kat,"notering":komm})
                    st.success(f"'{ben}' tillagd."); st.rerun()

    # ── 3. Importera Excel ────────────────────────────────────
    with t3:
        st.markdown("""
        **Importera prislista från Excel**

        Filen ska ha kolumner med:
        - **Kod** – artikelnummer (valfri)
        - **Benämning** – artikelns namn
        - **Enhet** – t.ex. st, m², tim
        - **Materialpris** – pris per enhet
        - **Arbetspris** – arbetstid × timpris per enhet

        Kolumnnamnen behöver inte vara exakta – programmet försöker matcha automatiskt.
        """)

        up=st.file_uploader("Välj Excel-fil (.xlsx/.xls)",type=["xlsx","xls"])
        if up:
            try:
                xl=pd.ExcelFile(up)
                sheet=xl.sheet_names[0]
                if len(xl.sheet_names)>1:
                    sheet=st.selectbox("Välj flik",xl.sheet_names)
                df=xl.parse(sheet)
                df.columns=[str(c).strip().lower() for c in df.columns]

                # Auto-map columns
                cm={}
                for col in df.columns:
                    if any(x in col for x in ["kod","code","nr","id"]): cm["kod"]=col
                    elif any(x in col for x in ["benäm","namn","desc","beskriv","text"]): cm["benamning"]=col
                    elif any(x in col for x in ["enh","unit"]): cm["enhet"]=col
                    elif any(x in col for x in ["mat","material","pris","kostnad"]) and "arb" not in col: cm["matpris"]=col
                    elif any(x in col for x in ["arb","arbete","lön","tim","work"]): cm["arbpris"]=col

                st.markdown(f"**Hittade kolumner:** {', '.join(df.columns.tolist())}")
                st.markdown("**Matchning:**")
                for k,v in cm.items():
                    st.markdown(f"- {k} → `{v}`")

                if "benamning" not in cm:
                    st.warning("Kunde inte hitta en kolumn för Benämning. Välj manuellt:")
                    cm["benamning"]=st.selectbox("Benämning-kolumn",df.columns.tolist())

                läge=st.radio("Importläge",
                              ["Lägg till (behåll befintliga)","Ersätt hela prisbanken"],
                              horizontal=True)

                if st.button("✅ Importera", type="primary"):
                    ny=[]
                    for _,row in df.iterrows():
                        item={"kod":  str(row.get(cm.get("kod",""),"")),
                              "benamning": str(row.get(cm["benamning"],"")),
                              "enhet":str(row.get(cm.get("enhet",""),"st")),
                              "matpris":float(row.get(cm.get("matpris",""),0) or 0),
                              "arbpris":float(row.get(cm.get("arbpris",""),0) or 0),
                              "leverantor":""}
                        if item["benamning"] and item["benamning"]!="nan":
                            ny.append(item)
                    if läge.startswith("Lägg"):
                        st.session_state.prisbank.extend(ny)
                    else:
                        st.session_state.prisbank=ny
                    st.success(f"✅ {len(ny)} artiklar importerade.")
                    st.rerun()
            except Exception as e:
                st.error(f"Importfel: {e}")

# ──────────────────────────────────────────────────────────────
#  TAB MALLAR
# ──────────────────────────────────────────────────────────────
def tab_mallar():
    st.markdown('<div class="section-hdr">Mallar</div>',unsafe_allow_html=True)
    mallar=st.session_state.mallar; proj=st.session_state.projekt
    l,r=st.columns([1,2])
    with l:
        st.markdown("**Spara kalkyl som mall**")
        namn=st.text_input("Mallnamn",placeholder="T.ex. Standardbadrum")
        if st.button("💾 Spara",type="primary"):
            if namn and proj.get("rader"):
                mallar.append({"namn":namn,"rader":copy.deepcopy(proj["rader"]),
                               "skapad":datetime.date.today().isoformat()})
                st.session_state.mallar=mallar; st.success(f"'{namn}' sparad."); st.rerun()
            else: st.warning("Kalkylen är tom eller inget namn angivet.")
        st.divider()
        if mallar:
            st.download_button("⬇ Exportera mallar",
                data=json.dumps(mallar,ensure_ascii=False,indent=2).encode(),
                file_name="mallar.json",mime="application/json",
                use_container_width=True)
        up=st.file_uploader("📥 Importera mallar",type=["json"],key="mi")
        if up:
            try:
                imp=json.loads(up.read()); st.session_state.mallar.extend(imp)
                st.success(f"{len(imp)} mallar importerade."); st.rerun()
            except: st.error("Ogiltigt format.")
    with r:
        if not mallar: st.info("Inga mallar sparade.")
        else:
            for i,m in enumerate(mallar):
                with st.expander(f"📑 {m['namn']}  ({len(m.get('rader',[]))} rader)  – {m.get('skapad','')}"):
                    df=pd.DataFrame(m.get("rader",[]))[["Radtyp","Benämning","Byggdel","Mängd","Á-pris","Kostnad"]]
                    st.dataframe(df,hide_index=True,use_container_width=True,height=180)
                    bc1,bc2=st.columns(2)
                    if bc1.button("➕ Lägg till i kalkyl",key=f"anv{i}",type="primary"):
                        for rad in m.get("rader",[]):
                            ny=copy.deepcopy(rad); berakna(ny)
                            proj["rader"].append(ny)
                        st.session_state.projekt=proj
                        st.success(f"{len(m['rader'])} rader tillagda."); st.rerun()
                    if bc2.button("🗑 Radera",key=f"del{i}"):
                        mallar.pop(i); st.session_state.mallar=mallar; st.rerun()

# ──────────────────────────────────────────────────────────────
#  TAB BYGGDELAR
# ──────────────────────────────────────────────────────────────
def tab_byggdelar():
    proj=st.session_state.projekt; rader=proj.get("rader",[])
    bdlar=proj.get("byggdelar",BYGGDEL_DEF)
    st.markdown('<div class="section-hdr">Byggdelar – summering</div>',unsafe_allow_html=True)
    totals={}
    for r in rader:
        bd=r.get("Byggdel","") or "(Utan byggdel)"
        if bd not in totals: totals[bd]={"n":0,"k":0.0,"f":0.0}
        totals[bd]["n"]+=1; totals[bd]["k"]+=sf(r.get("Kostnad",0))
        totals[bd]["f"]+=sf(r.get("Försäljning",0))
    rows=[]
    for bd in list(bdlar)+["(Utan byggdel)"]:
        t=totals.get(bd,{"n":0,"k":0.0,"f":0.0})
        tb_=t["f"]-t["k"]; mg=(tb_/t["f"]*100) if t["f"] else 0
        rows.append({"Byggdel":bd,"Rader":t["n"],"Kostnad":round(t["k"],0),
                     "Försäljning":round(t["f"],0),"TB":round(tb_,0),"Marginal %":round(mg,1)})
    all_k=sum(r["Kostnad"] for r in rows); all_f=sum(r["Försäljning"] for r in rows)
    tb_t=all_f-all_k; mg_t=(tb_t/all_f*100) if all_f else 0
    rows.append({"Byggdel":"▶ TOTALT","Rader":sum(r["Rader"] for r in rows),
                 "Kostnad":round(all_k,0),"Försäljning":round(all_f,0),
                 "TB":round(tb_t,0),"Marginal %":round(mg_t,1)})
    st.dataframe(pd.DataFrame(rows),hide_index=True,use_container_width=True,
        column_config={
            "Kostnad":    st.column_config.NumberColumn(format="%,.0f kr"),
            "Försäljning":st.column_config.NumberColumn(format="%,.0f kr"),
            "TB":         st.column_config.NumberColumn(format="%,.0f kr"),
            "Marginal %": st.column_config.NumberColumn(format="%.1f %%"),
        })
    st.divider()
    vald=st.selectbox("Visa rader för:",["–"]+list(bdlar)+["(Utan byggdel)"])
    if vald!="–":
        bd_rader=[r for r in rader if (r.get("Byggdel","") or "(Utan byggdel)")==vald]
        if bd_rader:
            st.dataframe(pd.DataFrame(bd_rader)[
                ["Radtyp","Benämning","Mängd","Enhet","Á-pris","Kostnad","Försäljning","Leverantör"]],
                hide_index=True,use_container_width=True)
        else: st.info(f"Inga rader för {vald}.")

# ──────────────────────────────────────────────────────────────
#  TAB SLUTSIDA
# ──────────────────────────────────────────────────────────────
def tab_slutsida():
    proj=st.session_state.projekt
    st.markdown('<div class="section-hdr">Slutsida – Ekonomi</div>',unsafe_allow_html=True)
    l,r=st.columns([1,1])
    with l:
        st.markdown("**Omkostnader (kr)**")
        omk=proj.get("omkostnader",{k:0.0 for k in OMKKOSTNADER})
        ny_omk={}
        for k in OMKKOSTNADER:
            ny_omk[k]=st.number_input(k,value=float(omk.get(k,0)),
                                       min_value=0.0,step=1000.0,key=f"o_{k}")
        proj["omkostnader"]=ny_omk
        st.divider()
        st.markdown("**Påslag (%)**")
        pas=proj.get("paslag",{"Omkostnad %":10.0,"Risk %":5.0,"Vinst %":8.0,"Rabatt %":0.0})
        ny_pas={}
        for k in ["Omkostnad %","Risk %","Vinst %","Rabatt %"]:
            ny_pas[k]=st.number_input(k,value=float(pas.get(k,0)),
                                       min_value=0.0,max_value=100.0,step=0.5,key=f"p_{k}")
        proj["paslag"]=ny_pas
        st.session_state.projekt=proj

    with r:
        s=summera(proj)
        st.markdown("**Resultat**")
        items=[("Direkta kostnader",s["dir_k"],False),
               ("Omkostnader (fasta)",s["omk_s"],False),
               ("Omkostnadspåslag",s["omk_b"],False),
               ("Riskpåslag",s["ris_b"],False),
               ("Självkostnad",s["sjk"],True),
               ("Vinst",s["vin_b"],False),
               ("Försäljningspris",s["fp"],True),
               ("Täckningsbidrag",s["tb"],True)]
        for lbl,val,bold in items:
            txt=f"**{lbl}:** &nbsp; **{val:,.0f} kr**" if bold else f"{lbl}: &nbsp; {val:,.0f} kr"
            st.markdown(txt,unsafe_allow_html=True)
        st.divider()
        m1,m2=st.columns(2)
        m1.metric("Försäljningspris",kr(s["fp"]))
        m2.metric("Marginal",pct(s["mg"]))
        st.divider()
        if OPX_OK:
            buf=io.BytesIO(); _xl_slutsida(proj,buf); buf.seek(0)
            st.download_button("📊 Excel-slutsida",data=buf,
                file_name=f"{name_safe(proj)}_slutsida.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True)
        if RL_OK:
            buf=io.BytesIO(); _pdf_slutsida(proj,buf); buf.seek(0)
            st.download_button("📄 PDF-slutsida",data=buf,
                file_name=f"{name_safe(proj)}_slutsida.pdf",
                mime="application/pdf",use_container_width=True)

# ──────────────────────────────────────────────────────────────
#  EXPORT HELPERS
# ──────────────────────────────────────────────────────────────
def name_safe(proj): return (proj.get("projektnamn","projekt") or "projekt").replace(" ","_")

def _xl_kalkyl(proj,buf):
    wb=openpyxl.Workbook(); ws=wb.active; ws.title="Kalkyl"
    hdrs=["Typ","Benämning","Byggdel","Mängd","Enhet","Tim","Á-pris","Kostnad","Pål%","Försäljning","Leverantör"]
    ws.append(hdrs)
    for cell in ws[1]:
        cell.fill=PatternFill("solid",fgColor="1A3A5C"); cell.font=Font(color="FFFFFF",bold=True)
    for r in proj.get("rader",[]):
        ws.append([r.get("Radtyp",""),r.get("Benämning",""),r.get("Byggdel",""),
                   r.get("Mängd",0),r.get("Enhet",""),r.get("Timmar",0),
                   r.get("Á-pris",0),r.get("Kostnad",0),r.get("Påslag %",0),
                   r.get("Försäljning",0),r.get("Leverantör","")])
    for col,w in zip("ABCDEFGHIJK",[8,30,14,8,6,6,10,12,6,12,14]):
        ws.column_dimensions[col].width=w
    wb.save(buf)

def _xl_slutsida(proj,buf):
    wb=openpyxl.Workbook(); ws=wb.active; ws.title="Slutsida"
    s=summera(proj)
    ws["A1"]=proj.get("projektnamn",""); ws["A1"].font=Font(bold=True,size=13)
    ws.append([]); ws.append(["Post","Belopp (kr)"])
    for cell in ws[3]: cell.fill=PatternFill("solid",fgColor="1A3A5C"); cell.font=Font(color="FFFFFF",bold=True)
    for p,v in [("Direkta kostnader",s["dir_k"]),("Omkostnader",s["omk_s"]),
                ("Självkostnad",s["sjk"]),("Vinst",s["vin_b"]),
                ("Försäljningspris",s["fp"]),("TB",s["tb"]),("Marginal %",s["mg"])]:
        ws.append([p,round(v,2)])
    ws.column_dimensions["A"].width=28; ws.column_dimensions["B"].width=18
    wb.save(buf)

def _pdf_kalkyl(proj,buf):
    doc=SimpleDocTemplate(buf,pagesize=A4,leftMargin=15*mm,rightMargin=15*mm,
                          topMargin=20*mm,bottomMargin=20*mm)
    styles=getSampleStyleSheet(); els=[]
    els.append(Paragraph(f"Kalkyl – {proj.get('projektnamn','')}",styles["Title"]))
    els.append(Spacer(1,4*mm))
    rows=[["Typ","Benämning","Byggdel","Mängd","Enhet","Á-pris","Kostnad","Försälj."]]
    for r in proj.get("rader",[]):
        rows.append([r.get("Radtyp","")[:3],r.get("Benämning","")[:30],
                     r.get("Byggdel","")[:12],f"{r.get('Mängd',0):.1f}",
                     r.get("Enhet",""),f"{r.get('Á-pris',0):.0f}",
                     f"{r.get('Kostnad',0):.0f}",f"{r.get('Försäljning',0):.0f}"])
    t=Table(rows,colWidths=[25,90,50,22,20,38,40,40],repeatRows=1)
    t.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,0),rl_colors.HexColor("#1A3A5C")),
        ("TEXTCOLOR",(0,0),(-1,0),rl_colors.white),
        ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
        ("FONTSIZE",(0,0),(-1,-1),7),
        ("ROWBACKGROUNDS",(0,1),(-1,-1),[rl_colors.HexColor("#EEF3F8"),rl_colors.white]),
        ("GRID",(0,0),(-1,-1),0.3,rl_colors.grey),
    ])); els.append(t); doc.build(els)

def _pdf_slutsida(proj,buf):
    doc=SimpleDocTemplate(buf,pagesize=A4,leftMargin=20*mm,rightMargin=20*mm,
                          topMargin=25*mm,bottomMargin=20*mm)
    styles=getSampleStyleSheet(); s=summera(proj); els=[]
    els.append(Paragraph(f"Kalkyl – {proj.get('projektnamn','')}",styles["Title"]))
    els.append(Spacer(1,8*mm))
    data=[["Post","Belopp"],
          ["Direkta kostnader",f"{s['dir_k']:,.0f} kr"],
          ["Självkostnad",f"{s['sjk']:,.0f} kr"],
          ["Försäljningspris",f"{s['fp']:,.0f} kr"],
          ["Täckningsbidrag",f"{s['tb']:,.0f} kr"],
          ["Marginal",f"{s['mg']:.2f} %"]]
    t=Table(data,colWidths=[120*mm,60*mm])
    t.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,0),rl_colors.HexColor("#1A3A5C")),
        ("TEXTCOLOR",(0,0),(-1,0),rl_colors.white),
        ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
        ("FONTSIZE",(0,0),(-1,-1),11),
        ("ROWBACKGROUNDS",(0,1),(-1,-1),[rl_colors.HexColor("#EEF3F8"),rl_colors.white]),
        ("GRID",(0,0),(-1,-1),0.5,rl_colors.grey),
        ("TOPPADDING",(0,0),(-1,-1),5),("BOTTOMPADDING",(0,0),(-1,-1),5),
    ])); els.append(t); doc.build(els)

# ──────────────────────────────────────────────────────────────
#  MAIN
# ──────────────────────────────────────────────────────────────
def main():
    init(); sidebar()
    sida=st.session_state.get("sida","🏠  Start")
    if   "Start"      in sida: tab_start()
    elif "Projektinfo" in sida: tab_projekt()
    elif "Kalkyl"     in sida: tab_kalkyl()
    elif "Prisbank"   in sida: tab_prisbank()
    elif "Mallar"     in sida: tab_mallar()
    elif "Byggdelar"  in sida: tab_byggdelar()
    elif "Slutsida"   in sida: tab_slutsida()
    else: tab_start()

if __name__=="__main__": main()
