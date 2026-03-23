#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Kalkylprogram Bygg & Entreprenad  v1.0
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog, simpledialog
import json, os, sys, shutil, datetime, copy, platform, subprocess
from pathlib import Path

try:
    import pandas as pd
    PANDAS_OK = True
except ImportError:
    PANDAS_OK = False

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPX_OK = True
except ImportError:
    OPX_OK = False

try:
    from reportlab.lib import colors as rl_colors
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import mm
    RL_OK = True
except ImportError:
    RL_OK = False

# ──────────────────────────────────────────────────────────────
#  CONSTANTS
# ──────────────────────────────────────────────────────────────
APP_TITLE  = "Kalkylprogram – Bygg & Entreprenad"
APP_VER    = "1.0"

C = {
    "bg":       "#f4f6f9",
    "panel":    "#ffffff",
    "primary":  "#1a3a5c",
    "secondary":"#2e6da4",
    "accent":   "#f0a500",
    "text":     "#1a1a1a",
    "muted":    "#6b7280",
    "success":  "#16a34a",
    "danger":   "#dc2626",
    "roweven":  "#f0f4f8",
    "rowodd":   "#ffffff",
    "hdr_bg":   "#1a3a5c",
    "hdr_fg":   "#ffffff",
    "sel":      "#bfd7ef",
}

ENHETER      = ["st","m","m²","m³","kg","ton","tim","pers","ls","rund"]
RADTYPER     = ["Material","Arbete","UE"]
STATUS_LIST  = ["Kalkyl","Offert","Pågående","Vunnen","Förlorad","Avslutad","Pausad"]
MAPPAR       = ["Kalkyl","Underlag","Ritningar","Inköp","UE","Export","Övrigt"]
BYGGDEL_DEF  = ["Badrum","Kök","Vägg","Etapp 1","Rivning","Snickeri",
                "Grund","Fasad","Tak","El","VVS","Övrigt"]

BKI = {
    "Flerbostadshus": {"2018":175.2,"2019":180.1,"2020":183.5,"2021":192.4,
                       "2022":215.8,"2023":228.3,"2024":232.1,"2025":238.5,"2026":242.0},
    "Småhus":         {"2018":168.4,"2019":173.2,"2020":177.8,"2021":186.3,
                       "2022":208.9,"2023":221.7,"2024":225.4,"2025":231.0,"2026":234.5},
    "ROT/Ombyggnad":  {"2018":171.3,"2019":176.5,"2020":180.2,"2021":189.1,
                       "2022":212.1,"2023":224.9,"2024":228.7,"2025":234.8,"2026":238.2},
}
BKI_BASÅR = "2020"

# ──────────────────────────────────────────────────────────────
#  UTILITIES
# ──────────────────────────────────────────────────────────────
def fmt(v, dec=2):
    try:
        f = float(v)
        s = f"{f:,.{dec}f}"
        return s.replace(",", "\u2009").replace(".", ",")   # thin-space thousands, comma decimal
    except:
        return str(v) if v else ""

def sfloat(v, d=0.0):
    try:
        return float(str(v).replace("\u2009","").replace(" ","").replace(",","."))
    except:
        return d

def app_dir():
    return Path(sys.executable).parent if getattr(sys,"frozen",False) else Path(__file__).parent

def settings_path(): return app_dir() / "settings.json"
def mallar_path():   return app_dir() / "mallar.json"
def prisbank_path(): return app_dir() / "prisbank.json"

def load_json(p, default):
    try:
        with open(p,"r",encoding="utf-8") as f: return json.load(f)
    except: return default

def save_json(p, data):
    try:
        with open(p,"w",encoding="utf-8") as f: json.dump(data,f,ensure_ascii=False,indent=2)
    except: pass

def open_file(path):
    try:
        if platform.system()=="Windows": os.startfile(path)
        elif platform.system()=="Darwin": subprocess.run(["open",path])
        else: subprocess.run(["xdg-open",path])
    except: pass

def create_folders(base):
    try:
        b = Path(base); b.mkdir(parents=True, exist_ok=True)
        for m in MAPPAR: (b/m).mkdir(exist_ok=True)
        return True
    except: return False

# ──────────────────────────────────────────────────────────────
#  DATA MODEL
# ──────────────────────────────────────────────────────────────
def empty_rad():
    return {"id":"","radtyp":"Material","kod":"","benamning":"","beskrivning":"",
            "byggdel":"","mangd":0.0,"enhet":"st","timmar":0.0,"apris":0.0,
            "kostnad":0.0,"paslag":0.0,"forsaljning":0.0,"leverantor":"","kommentar":""}

def empty_projekt():
    return {
        "projektnamn":"","projektnummer":"","kund":"","bestallar":"","adress":"",
        "datum":datetime.date.today().isoformat(),"kalkylansvarig":"",
        "status":"Kalkyl","kommentar":"","projektmapp":"",
        "rader":[], "dokument":[], "byggdelar":list(BYGGDEL_DEF),
        "omkostnader":{"Arbetsledning":0.0,"Etablering":0.0,"Administration":0.0,
                       "Transporter":0.0,"Garanti":0.0,"Risk":0.0,"Övrigt":0.0},
        "paslag":{"Omkostnad %":10.0,"Risk %":5.0,"Vinst %":8.0,"Rabatt %":0.0}
    }

def berakna(rad):
    mt = rad.get("radtyp","Material")
    m  = sfloat(rad.get("mangd",0))
    t  = sfloat(rad.get("timmar",0))
    a  = sfloat(rad.get("apris",0))
    p  = sfloat(rad.get("paslag",0))
    k  = (t*a) if mt=="Arbete" else (m*a)
    rad["kostnad"]   = round(k, 2)
    rad["forsaljning"] = round(k*(1+p/100), 2)
    return rad

def summera_projekt(proj):
    rader = proj.get("rader",[])
    dir_k = sum(sfloat(r.get("kostnad",0)) for r in rader)
    dir_f = sum(sfloat(r.get("forsaljning",0)) for r in rader)
    omk   = proj.get("omkostnader",{})
    omk_s = sum(sfloat(v) for v in omk.values())
    p     = proj.get("paslag",{})
    o_pct = sfloat(p.get("Omkostnad %",0))
    r_pct = sfloat(p.get("Risk %",0))
    v_pct = sfloat(p.get("Vinst %",0))
    d_pct = sfloat(p.get("Rabatt %",0))
    omk_b = dir_k * o_pct/100
    ris_b = dir_k * r_pct/100
    sjk   = dir_k + omk_s + omk_b + ris_b
    vin_b = sjk * v_pct/100
    fp    = sjk + vin_b
    rab   = fp * d_pct/100
    fp_n  = fp - rab
    tb    = fp_n - dir_k
    mg    = (tb/fp_n*100) if fp_n else 0
    return {"dir_k":dir_k,"dir_f":dir_f,"omk_s":omk_s,"omk_b":omk_b,"ris_b":ris_b,
            "sjk":sjk,"vin_b":vin_b,"fp":fp_n,"tb":tb,"mg":mg}
# ──────────────────────────────────────────────────────────────
#  STYLING HELPERS
# ──────────────────────────────────────────────────────────────
def apply_style(root):
    s = ttk.Style(root)
    s.theme_use("clam")
    s.configure(".", background=C["bg"], foreground=C["text"],
                font=("Segoe UI",10), fieldbackground=C["panel"])
    s.configure("TNotebook", background=C["primary"], tabmargins=[2,4,2,0])
    s.configure("TNotebook.Tab", background=C["primary"], foreground=C["hdr_fg"],
                padding=[14,6], font=("Segoe UI",10,"bold"))
    s.map("TNotebook.Tab",
          background=[("selected",C["accent"]),("active",C["secondary"])],
          foreground=[("selected",C["text"]),("active","white")])
    s.configure("TFrame", background=C["bg"])
    s.configure("Panel.TFrame", background=C["panel"], relief="flat")
    s.configure("TLabel", background=C["bg"], foreground=C["text"])
    s.configure("Panel.TLabel", background=C["panel"])
    s.configure("Header.TLabel", background=C["primary"], foreground=C["hdr_fg"],
                font=("Segoe UI",12,"bold"), padding=[10,6])
    s.configure("TButton", background=C["secondary"], foreground="white",
                font=("Segoe UI",9,"bold"), padding=[8,4], relief="flat")
    s.map("TButton",background=[("active",C["primary"])],foreground=[("active","white")])
    s.configure("Accent.TButton", background=C["accent"], foreground=C["text"])
    s.map("Accent.TButton",background=[("active","#d4900a")])
    s.configure("Danger.TButton", background=C["danger"], foreground="white")
    s.configure("TEntry", fieldbackground=C["panel"], relief="flat", padding=[4,3])
    s.configure("TCombobox", fieldbackground=C["panel"])
    s.configure("TScrollbar", background=C["bg"], troughcolor=C["bg"], relief="flat")
    s.configure("Treeview", background=C["panel"], fieldbackground=C["panel"],
                foreground=C["text"], rowheight=24, font=("Segoe UI",9))
    s.configure("Treeview.Heading", background=C["primary"], foreground=C["hdr_fg"],
                font=("Segoe UI",9,"bold"), relief="flat")
    s.map("Treeview", background=[("selected",C["sel"])], foreground=[("selected",C["text"])])

def lbl(parent, text, bold=False, big=False, style=None, **kw):
    f = ("Segoe UI", 11 if big else 10, "bold" if bold else "normal")
    st = style or "TLabel"
    return ttk.Label(parent, text=text, font=f, style=st, **kw)

def btn(parent, text, cmd, style="TButton", **kw):
    return ttk.Button(parent, text=text, command=cmd, style=style, **kw)

def sep(parent, orient="horizontal"):
    return ttk.Separator(parent, orient=orient)

def scrolled_tree(parent, columns, headings, widths, show="headings", height=20):
    frame = ttk.Frame(parent)
    vsb = ttk.Scrollbar(frame, orient="vertical")
    hsb = ttk.Scrollbar(frame, orient="horizontal")
    tree = ttk.Treeview(frame, columns=columns, show=show,
                        yscrollcommand=vsb.set, xscrollcommand=hsb.set, height=height)
    vsb.configure(command=tree.yview)
    hsb.configure(command=tree.xview)
    for col, hd, w in zip(columns, headings, widths):
        tree.heading(col, text=hd, anchor="w")
        tree.column(col, width=w, minwidth=40, anchor="w")
    tree.grid(row=0,column=0,sticky="nsew")
    vsb.grid(row=0,column=1,sticky="ns")
    hsb.grid(row=1,column=0,sticky="ew")
    frame.rowconfigure(0,weight=1)
    frame.columnconfigure(0,weight=1)
    tree.tag_configure("even", background=C["roweven"])
    tree.tag_configure("odd",  background=C["rowodd"])
    return frame, tree

# ──────────────────────────────────────────────────────────────
#  RAD-DIALOG  (lägg till / redigera kalkylrad)
# ──────────────────────────────────────────────────────────────
class RadDialog(tk.Toplevel):
    def __init__(self, parent, rad=None, byggdelar=None, title="Lägg till rad"):
        super().__init__(parent)
        self.title(title)
        self.resizable(True, True)
        self.configure(bg=C["bg"])
        self.result = None
        self._byggdelar = byggdelar or list(BYGGDEL_DEF)
        self._rad = copy.deepcopy(rad) if rad else empty_rad()
        self._build()
        self.grab_set()
        self.geometry("620x560")
        self.update_idletasks()
        # center
        x = parent.winfo_rootx() + (parent.winfo_width()-620)//2
        y = parent.winfo_rooty() + (parent.winfo_height()-560)//2
        self.geometry(f"+{x}+{y}")
        self.wait_window()

    def _fld(self, label, row, var, options=None, width=28):
        ttk.Label(self._form, text=label, background=C["panel"],
                  font=("Segoe UI",9)).grid(row=row,column=0,sticky="w",padx=6,pady=3)
        if options is not None:
            cb = ttk.Combobox(self._form, textvariable=var, values=options,
                              width=width, state="readonly")
            cb.grid(row=row,column=1,sticky="ew",padx=6,pady=3)
            return cb
        else:
            e = ttk.Entry(self._form, textvariable=var, width=width+2)
            e.grid(row=row,column=1,sticky="ew",padx=6,pady=3)
            return e

    def _build(self):
        r = self._rad
        # header
        ttk.Label(self, text="  Kalkylrad", style="Header.TLabel").pack(fill="x")

        frm = ttk.Frame(self, style="Panel.TFrame")
        frm.pack(fill="both",expand=True,padx=10,pady=10)
        self._form = frm
        frm.columnconfigure(1,weight=1)

        self.v_typ   = tk.StringVar(value=r["radtyp"])
        self.v_kod   = tk.StringVar(value=r.get("kod",""))
        self.v_ben   = tk.StringVar(value=r.get("benamning",""))
        self.v_bsk   = tk.StringVar(value=r.get("beskrivning",""))
        self.v_bdl   = tk.StringVar(value=r.get("byggdel",""))
        self.v_mng   = tk.StringVar(value=str(r.get("mangd",0)))
        self.v_enh   = tk.StringVar(value=r.get("enhet","st"))
        self.v_tim   = tk.StringVar(value=str(r.get("timmar",0)))
        self.v_apr   = tk.StringVar(value=str(r.get("apris",0)))
        self.v_pas   = tk.StringVar(value=str(r.get("paslag",0)))
        self.v_lev   = tk.StringVar(value=r.get("leverantor",""))
        self.v_kom   = tk.StringVar(value=r.get("kommentar",""))

        self._fld("Radtyp",    0, self.v_typ, RADTYPER)
        self._fld("Kod",       1, self.v_kod)
        self._fld("Benämning", 2, self.v_ben)
        self._fld("Beskrivning",3,self.v_bsk)
        self._fld("Byggdel",   4, self.v_bdl, self._byggdelar)
        self._fld("Mängd",     5, self.v_mng)
        self._fld("Enhet",     6, self.v_enh, ENHETER)
        self._fld("Timmar",    7, self.v_tim)
        self._fld("Á-pris",    8, self.v_apr)
        self._fld("Påslag %",  9, self.v_pas)
        self._fld("Leverantör",10,self.v_lev)
        self._fld("Kommentar", 11,self.v_kom)

        # preview
        self._prev = ttk.Label(frm, text="", background=C["panel"],
                               font=("Segoe UI",9), foreground=C["muted"])
        self._prev.grid(row=12,column=0,columnspan=2,sticky="w",padx=6,pady=4)

        for v in (self.v_mng,self.v_tim,self.v_apr,self.v_pas):
            v.trace_add("write", lambda *_: self._update_preview())
        self.v_typ.trace_add("write", lambda *_: self._update_preview())
        self._update_preview()

        # buttons
        bf = ttk.Frame(self, style="TFrame")
        bf.pack(fill="x",padx=10,pady=(0,10))
        btn(bf,"✓  Spara",  self._save,  "Accent.TButton").pack(side="right",padx=4)
        btn(bf,"Avbryt",    self.destroy).pack(side="right")

    def _update_preview(self):
        try:
            typ = self.v_typ.get()
            m = sfloat(self.v_mng.get()); t = sfloat(self.v_tim.get())
            a = sfloat(self.v_apr.get()); p = sfloat(self.v_pas.get())
            k = (t*a) if typ=="Arbete" else (m*a)
            f = k*(1+p/100)
            self._prev.config(text=f"  Kostnad: {fmt(k)} kr   Försäljning: {fmt(f)} kr")
        except: pass

    def _save(self):
        if not self.v_ben.get().strip():
            messagebox.showwarning("Saknas","Benämning måste fyllas i.",parent=self); return
        r = self._rad
        r["radtyp"]    = self.v_typ.get()
        r["kod"]       = self.v_kod.get().strip()
        r["benamning"] = self.v_ben.get().strip()
        r["beskrivning"]=self.v_bsk.get().strip()
        r["byggdel"]   = self.v_bdl.get()
        r["mangd"]     = sfloat(self.v_mng.get())
        r["enhet"]     = self.v_enh.get()
        r["timmar"]    = sfloat(self.v_tim.get())
        r["apris"]     = sfloat(self.v_apr.get())
        r["paslag"]    = sfloat(self.v_pas.get())
        r["leverantor"]= self.v_lev.get().strip()
        r["kommentar"] = self.v_kom.get().strip()
        berakna(r)
        if not r.get("id"): r["id"] = str(datetime.datetime.now().timestamp())
        self.result = r
        self.destroy()
# ──────────────────────────────────────────────────────────────
#  START TAB
# ──────────────────────────────────────────────────────────────
class StartTab(ttk.Frame):
    def __init__(self, parent, app):
        super().__init__(parent)
        self.app = app
        self.configure(style="TFrame")
        self._build()

    def _build(self):
        # Left panel – actions
        left = ttk.Frame(self, style="Panel.TFrame", width=280)
        left.pack(side="left", fill="y", padx=(16,8), pady=16)
        left.pack_propagate(False)

        ttk.Label(left, text="  Kalkylprogram", style="Header.TLabel",
                  font=("Segoe UI",14,"bold")).pack(fill="x")
        ttk.Label(left, text="  Bygg & Entreprenad", background=C["primary"],
                  foreground=C["accent"], font=("Segoe UI",10)).pack(fill="x")

        ttk.Frame(left, height=20, style="Panel.TFrame").pack()

        for txt, cmd in [
            ("📄  Nytt projekt",   self.app.new_projekt),
            ("📂  Öppna projekt",  self.app.open_projekt),
            ("💾  Spara projekt",  self.app.save_projekt),
        ]:
            b = ttk.Button(left, text=txt, command=cmd, style="TButton",
                           width=26)
            b.pack(padx=12, pady=4, fill="x")

        sep(left).pack(fill="x",padx=12,pady=10)
        ttk.Label(left, text="Senaste projekt", background=C["panel"],
                  font=("Segoe UI",9,"bold")).pack(anchor="w",padx=12)
        self._recent_frame = ttk.Frame(left, style="Panel.TFrame")
        self._recent_frame.pack(fill="x",padx=12,pady=4)
        self.refresh_recent()

        # Right panel – info
        right = ttk.Frame(self, style="Panel.TFrame")
        right.pack(side="left", fill="both", expand=True, padx=(0,16), pady=16)

        ttk.Label(right, text="  Projektöversikt", style="Header.TLabel").pack(fill="x")
        self._info = ttk.Label(right, text="\n  Inget projekt öppet.",
                               background=C["panel"], font=("Segoe UI",10),
                               justify="left", wraplength=500)
        self._info.pack(anchor="nw", padx=10, pady=10)

    def refresh_recent(self):
        for w in self._recent_frame.winfo_children(): w.destroy()
        s = load_json(settings_path(), {"recent_projects":[]})
        for p in s.get("recent_projects",[])[:8]:
            name = Path(p).stem
            b = ttk.Button(self._recent_frame, text=f"  {name}",
                           command=lambda pp=p: self.app.open_projekt(pp),
                           style="TButton", width=24)
            b.pack(fill="x", pady=1)

    def update_info(self, proj):
        if not proj:
            self._info.config(text="\n  Inget projekt öppet."); return
        s = summera_projekt(proj)
        txt = (f"\n  Projekt:    {proj.get('projektnamn','')}\n"
               f"  Nr:         {proj.get('projektnummer','')}\n"
               f"  Kund:       {proj.get('kund','')}\n"
               f"  Status:     {proj.get('status','')}\n"
               f"  Datum:      {proj.get('datum','')}\n"
               f"  Ansvarig:   {proj.get('kalkylansvarig','')}\n\n"
               f"  Kalkylrader:      {len(proj.get('rader',[]))}\n"
               f"  Direkta kostn:    {fmt(s['dir_k'])} kr\n"
               f"  Försäljningspris: {fmt(s['fp'])} kr\n"
               f"  TB:               {fmt(s['tb'])} kr\n"
               f"  Marginal:         {fmt(s['mg'],1)} %")
        self._info.config(text=txt)

# ──────────────────────────────────────────────────────────────
#  PROJEKT TAB
# ──────────────────────────────────────────────────────────────
class ProjektTab(ttk.Frame):
    def __init__(self, parent, app):
        super().__init__(parent)
        self.app = app
        self._vars = {}
        self._build()

    def _build(self):
        ttk.Label(self, text="  Projektinformation", style="Header.TLabel").pack(fill="x")

        canvas = tk.Canvas(self, background=C["bg"], highlightthickness=0)
        sb = ttk.Scrollbar(self, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=sb.set)
        sb.pack(side="right", fill="y")
        canvas.pack(fill="both", expand=True)

        inner = ttk.Frame(canvas, style="TFrame")
        win_id = canvas.create_window((0,0), window=inner, anchor="nw")

        def _resize(e):
            canvas.configure(scrollregion=canvas.bbox("all"))
            canvas.itemconfig(win_id, width=e.width)
        canvas.bind("<Configure>", _resize)
        inner.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))

        # Two columns
        col1 = ttk.Frame(inner, style="Panel.TFrame")
        col1.grid(row=0,column=0,padx=10,pady=10,sticky="nsew")
        col2 = ttk.Frame(inner, style="Panel.TFrame")
        col2.grid(row=0,column=1,padx=10,pady=10,sticky="nsew")
        inner.columnconfigure(0,weight=1)
        inner.columnconfigure(1,weight=1)

        def section(parent, title):
            ttk.Label(parent, text=f"  {title}", style="Header.TLabel",
                      font=("Segoe UI",10,"bold")).pack(fill="x",pady=(0,4))

        def field(parent, label, key, options=None):
            f = ttk.Frame(parent, style="Panel.TFrame")
            f.pack(fill="x", padx=8, pady=3)
            ttk.Label(f, text=label, background=C["panel"],
                      width=16, font=("Segoe UI",9)).pack(side="left")
            v = tk.StringVar()
            self._vars[key] = v
            if options:
                w = ttk.Combobox(f, textvariable=v, values=options,
                                 state="readonly", width=28)
            else:
                w = ttk.Entry(f, textvariable=v, width=30)
            w.pack(side="left", fill="x", expand=True)
            return v

        section(col1, "Projektuppgifter")
        field(col1, "Projektnamn",   "projektnamn")
        field(col1, "Projektnr",     "projektnummer")
        field(col1, "Kund",          "kund")
        field(col1, "Beställare",    "bestallar")
        field(col1, "Adress",        "adress")
        field(col1, "Datum",         "datum")
        field(col1, "Kalkylansvarig","kalkylansvarig")
        field(col1, "Status",        "status", STATUS_LIST)

        section(col2, "Kommentar")
        v = tk.StringVar(); self._vars["kommentar"] = v
        t = tk.Text(col2, height=5, font=("Segoe UI",9),
                    background=C["panel"], relief="flat", wrap="word")
        t.pack(fill="x",padx=8,pady=4)
        self._kommentar_text = t

        section(col2, "Projektmapp")
        mf = ttk.Frame(col2, style="Panel.TFrame")
        mf.pack(fill="x",padx=8,pady=3)
        self._vars["projektmapp"] = tk.StringVar()
        ttk.Entry(mf, textvariable=self._vars["projektmapp"], width=30,
                  state="readonly").pack(side="left",fill="x",expand=True)
        btn(mf,"Välj mapp", self._velj_mapp).pack(side="left",padx=4)
        btn(mf,"Skapa mappar", self._skapa_mappar, "Accent.TButton").pack(side="left")

        btn(col2,"💾  Spara projektinfo", self.app.save_projekt,
            "Accent.TButton").pack(padx=8,pady=10,anchor="w")

    def _velj_mapp(self):
        p = filedialog.askdirectory(title="Välj projektmapp")
        if p: self._vars["projektmapp"].set(p)

    def _skapa_mappar(self):
        p = self._vars["projektmapp"].get()
        if not p:
            messagebox.showwarning("Ingen mapp","Välj eller ange en projektmapp."); return
        if create_folders(p):
            messagebox.showinfo("Klart", f"Mappstruktur skapad:\n{p}")
        else:
            messagebox.showerror("Fel","Kunde inte skapa mappar.")

    def load(self, proj):
        for key, var in self._vars.items():
            var.set(str(proj.get(key,"") or ""))
        self._kommentar_text.delete("1.0","end")
        self._kommentar_text.insert("1.0", proj.get("kommentar",""))

    def save_to(self, proj):
        for key, var in self._vars.items():
            proj[key] = var.get()
        proj["kommentar"] = self._kommentar_text.get("1.0","end-1c")
# ──────────────────────────────────────────────────────────────
#  KALKYL TAB
# ──────────────────────────────────────────────────────────────
KALKYL_COLS = ("radtyp","kod","benamning","byggdel","mangd","enhet",
               "timmar","apris","kostnad","paslag","forsaljning","leverantor")
KALKYL_HDR  = ("Typ","Kod","Benämning","Byggdel","Mängd","Enh",
               "Tim","Á-pris","Kostnad","Pål%","Försäljning","Leverantör")
KALKYL_W    = (70,70,200,100,70,50,60,90,100,55,100,110)

class KalkylTab(ttk.Frame):
    def __init__(self, parent, app):
        super().__init__(parent)
        self.app = app
        self._build()

    def _build(self):
        ttk.Label(self, text="  Kalkyl", style="Header.TLabel").pack(fill="x")

        # Toolbar
        tb = ttk.Frame(self, style="Panel.TFrame")
        tb.pack(fill="x", padx=0, pady=0)
        for txt, cmd, st in [
            ("＋  Ny rad",       self._ny_rad,       "Accent.TButton"),
            ("✎  Redigera",     self._edit_rad,      "TButton"),
            ("⧉  Kopiera rad",  self._kopiera_rad,   "TButton"),
            ("✕  Ta bort",      self._ta_bort,       "Danger.TButton"),
            ("▲  Upp",          self._flytta_upp,    "TButton"),
            ("▼  Ner",          self._flytta_ner,    "TButton"),
        ]:
            btn(tb, txt, cmd, st).pack(side="left", padx=2, pady=4)

        sep(tb,"vertical").pack(side="left",fill="y",padx=6,pady=4)
        btn(tb,"📋  Från prisbank", self._fran_prisbank,"TButton").pack(side="left",padx=2,pady=4)
        btn(tb,"📑  Från mall",     self._fran_mall,    "TButton").pack(side="left",padx=2,pady=4)
        btn(tb,"💾  Spara som mall",self._spara_mall,   "TButton").pack(side="left",padx=2,pady=4)
        sep(tb,"vertical").pack(side="left",fill="y",padx=6,pady=4)
        btn(tb,"📊  Exportera Excel",self._export_excel,"TButton").pack(side="left",padx=2,pady=4)
        btn(tb,"📄  Exportera PDF",  self._export_pdf,  "TButton").pack(side="left",padx=2,pady=4)

        # Filter bar
        fb = ttk.Frame(self, style="TFrame")
        fb.pack(fill="x", padx=4, pady=2)
        ttk.Label(fb, text="Filtrera:").pack(side="left",padx=4)
        self._filter_var = tk.StringVar()
        self._filter_var.trace_add("write", lambda *_: self.refresh())
        ttk.Entry(fb, textvariable=self._filter_var, width=24).pack(side="left",padx=2)
        ttk.Label(fb, text="Byggdel:").pack(side="left",padx=8)
        self._bdl_var = tk.StringVar(value="Alla")
        self._bdl_cb  = ttk.Combobox(fb, textvariable=self._bdl_var,
                                      state="readonly", width=18)
        self._bdl_cb.pack(side="left")
        self._bdl_var.trace_add("write", lambda *_: self.refresh())
        ttk.Label(fb, text="Typ:").pack(side="left",padx=8)
        self._typ_var = tk.StringVar(value="Alla")
        ttk.Combobox(fb, textvariable=self._typ_var,
                     values=["Alla"]+RADTYPER,
                     state="readonly",width=12).pack(side="left")
        self._typ_var.trace_add("write", lambda *_: self.refresh())

        # Tree
        tf, self._tree = scrolled_tree(self, KALKYL_COLS, KALKYL_HDR, KALKYL_W, height=22)
        tf.pack(fill="both", expand=True, padx=4, pady=4)
        self._tree.bind("<Double-1>", lambda e: self._edit_rad())
        self._tree.bind("<Delete>",   lambda e: self._ta_bort())

        # Summary bar
        sf = ttk.Frame(self, style="Panel.TFrame")
        sf.pack(fill="x", side="bottom")
        self._sum_lbl = ttk.Label(sf, text="", background=C["panel"],
                                   font=("Segoe UI",9,"bold"), foreground=C["primary"])
        self._sum_lbl.pack(side="left", padx=10, pady=4)

    def _update_bdl_filter(self):
        bdlar = ["Alla"] + (self.app.projekt.get("byggdelar", list(BYGGDEL_DEF))
                            if self.app.projekt else list(BYGGDEL_DEF))
        self._bdl_cb["values"] = bdlar
        if self._bdl_var.get() not in bdlar: self._bdl_var.set("Alla")

    def refresh(self):
        self._update_bdl_filter()
        tree = self._tree
        tree.delete(*tree.get_children())
        if not self.app.projekt: return
        filt = self._filter_var.get().lower()
        bdl  = self._bdl_var.get()
        typ  = self._typ_var.get()
        tot_k = tot_f = 0.0
        for i, r in enumerate(self.app.projekt.get("rader",[])):
            if filt and filt not in r.get("benamning","").lower() \
                    and filt not in r.get("kod","").lower(): continue
            if bdl != "Alla" and r.get("byggdel","") != bdl: continue
            if typ != "Alla" and r.get("radtyp","") != typ: continue
            tag = "even" if i%2==0 else "odd"
            vals = (
                r.get("radtyp",""), r.get("kod",""), r.get("benamning",""),
                r.get("byggdel",""),
                fmt(r.get("mangd",0),1), r.get("enhet",""),
                fmt(r.get("timmar",0),1), fmt(r.get("apris",0)),
                fmt(r.get("kostnad",0)), fmt(r.get("paslag",0),1),
                fmt(r.get("forsaljning",0)), r.get("leverantor","")
            )
            tree.insert("","end", iid=r["id"], values=vals, tags=(tag,))
            tot_k += sfloat(r.get("kostnad",0))
            tot_f += sfloat(r.get("forsaljning",0))
        self._sum_lbl.config(
            text=f"  Rader: {len(tree.get_children())}   "
                 f"Kostnad: {fmt(tot_k)} kr   "
                 f"Försäljning: {fmt(tot_f)} kr   "
                 f"TB: {fmt(tot_f-tot_k)} kr"
        )

    def _selected_rad(self):
        sel = self._tree.selection()
        if not sel: return None, None
        rid = sel[0]
        for i,r in enumerate(self.app.projekt.get("rader",[])):
            if r.get("id")==rid: return i, r
        return None, None

    def _ny_rad(self):
        if not self.app.projekt:
            messagebox.showwarning("Inget projekt","Skapa eller öppna ett projekt."); return
        dlg = RadDialog(self, byggdelar=self.app.projekt.get("byggdelar",BYGGDEL_DEF))
        if dlg.result:
            self.app.projekt["rader"].append(dlg.result)
            self.refresh()
            self.app.mark_dirty()

    def _edit_rad(self):
        i, r = self._selected_rad()
        if r is None: return
        dlg = RadDialog(self, rad=r,
                        byggdelar=self.app.projekt.get("byggdelar",BYGGDEL_DEF),
                        title="Redigera rad")
        if dlg.result:
            self.app.projekt["rader"][i] = dlg.result
            self.refresh()
            self.app.mark_dirty()

    def _kopiera_rad(self):
        i, r = self._selected_rad()
        if r is None: return
        ny = copy.deepcopy(r)
        ny["id"] = str(datetime.datetime.now().timestamp())
        ny["benamning"] += " (kopia)"
        self.app.projekt["rader"].insert(i+1, ny)
        self.refresh()
        self.app.mark_dirty()

    def _ta_bort(self):
        i, r = self._selected_rad()
        if r is None: return
        if messagebox.askyesno("Ta bort",f"Ta bort rad '{r.get('benamning','')}'?"):
            self.app.projekt["rader"].pop(i)
            self.refresh()
            self.app.mark_dirty()

    def _flytta_upp(self):
        i, r = self._selected_rad()
        if r is None or i==0: return
        lst = self.app.projekt["rader"]
        lst[i-1], lst[i] = lst[i], lst[i-1]
        self.refresh()
        self._tree.selection_set(r["id"])

    def _flytta_ner(self):
        i, r = self._selected_rad()
        if r is None: return
        lst = self.app.projekt["rader"]
        if i >= len(lst)-1: return
        lst[i+1], lst[i] = lst[i], lst[i+1]
        self.refresh()
        self._tree.selection_set(r["id"])

    def _fran_prisbank(self):
        self.app.notebook.select(self.app.tab_idx["prisbank"])

    def _fran_mall(self):
        mallar = load_json(mallar_path(), [])
        if not mallar:
            messagebox.showinfo("Inga mallar","Inga sparade mallar hittades."); return
        win = tk.Toplevel(self)
        win.title("Välj mall")
        win.configure(bg=C["bg"])
        win.geometry("400x400")
        ttk.Label(win, text="  Välj mall", style="Header.TLabel").pack(fill="x")
        lb = tk.Listbox(win, font=("Segoe UI",10), selectmode="single",
                        background=C["panel"])
        lb.pack(fill="both",expand=True,padx=10,pady=8)
        for m in mallar: lb.insert("end", m.get("namn",""))
        def _anvand():
            sel = lb.curselection()
            if not sel: return
            m = mallar[sel[0]]
            for r in m.get("rader",[]):
                ny = copy.deepcopy(r)
                ny["id"] = str(datetime.datetime.now().timestamp())
                berakna(ny)
                self.app.projekt["rader"].append(ny)
            self.refresh()
            self.app.mark_dirty()
            win.destroy()
        btn(win,"Använd mall", _anvand, "Accent.TButton").pack(pady=6)

    def _spara_mall(self):
        sel = self._tree.selection()
        if not sel:
            messagebox.showinfo("Välj rader","Markera rader att spara som mall."); return
        namn = simpledialog.askstring("Mallnamn","Ge mallen ett namn:",parent=self)
        if not namn: return
        ids = set(sel)
        rader = [copy.deepcopy(r) for r in self.app.projekt.get("rader",[])
                 if r.get("id") in ids]
        mallar = load_json(mallar_path(), [])
        mallar.append({"namn":namn, "rader":rader,
                       "skapad":datetime.date.today().isoformat()})
        save_json(mallar_path(), mallar)
        messagebox.showinfo("Sparat",f"Mall '{namn}' sparad.")
        self.app.mallar_tab.refresh()

    def _export_excel(self):
        self.app.export_excel_kalkyl()

    def _export_pdf(self):
        self.app.export_pdf_kalkyl()
# ──────────────────────────────────────────────────────────────
#  PRISBANK TAB
# ──────────────────────────────────────────────────────────────
PB_COLS = ("kod","benamning","enhet","matpris","arbpris")
PB_HDR  = ("Kod","Benämning","Enhet","Materialpris","Arbetspris")
PB_W    = (80,260,60,110,110)

class PrisbankTab(ttk.Frame):
    def __init__(self, parent, app):
        super().__init__(parent)
        self.app = app
        self._data = []
        self._build()
        self._load_prisbank()

    def _build(self):
        ttk.Label(self, text="  Prisbank", style="Header.TLabel").pack(fill="x")

        tb = ttk.Frame(self, style="Panel.TFrame")
        tb.pack(fill="x")
        btn(tb,"📥  Importera Excel", self._import_excel, "Accent.TButton").pack(side="left",padx=4,pady=4)
        btn(tb,"＋  Ny artikel",      self._ny_artikel,   "TButton").pack(side="left",padx=2,pady=4)
        btn(tb,"✕  Ta bort",          self._ta_bort,      "Danger.TButton").pack(side="left",padx=2,pady=4)
        btn(tb,"→  Lägg till kalkyl", self._till_kalkyl,  "TButton").pack(side="left",padx=2,pady=4)

        sep(tb,"vertical").pack(side="left",fill="y",padx=6,pady=4)

        # BKI
        ttk.Label(tb, text="BKI-index:", background=C["panel"],
                  font=("Segoe UI",9,"bold")).pack(side="left",padx=4)
        self._bki_typ  = tk.StringVar(value="Flerbostadshus")
        self._bki_år   = tk.StringVar(value=str(datetime.date.today().year))
        self._bki_bas  = tk.StringVar(value=BKI_BASÅR)
        ttk.Combobox(tb, textvariable=self._bki_typ,
                     values=list(BKI.keys()), state="readonly",
                     width=18).pack(side="left",padx=2)
        ttk.Label(tb, text="Basår:", background=C["panel"],
                  font=("Segoe UI",9)).pack(side="left",padx=4)
        ttk.Combobox(tb, textvariable=self._bki_bas,
                     values=list(BKI["Flerbostadshus"].keys()),
                     state="readonly", width=7).pack(side="left",padx=2)
        ttk.Label(tb, text="→ År:", background=C["panel"],
                  font=("Segoe UI",9)).pack(side="left",padx=4)
        ttk.Combobox(tb, textvariable=self._bki_år,
                     values=list(BKI["Flerbostadshus"].keys()),
                     state="readonly", width=7).pack(side="left",padx=2)
        btn(tb,"Räkna om",self._bki_raknna,"TButton").pack(side="left",padx=4)
        self._bki_lbl = ttk.Label(tb, text="", background=C["panel"],
                                   font=("Segoe UI",9), foreground=C["success"])
        self._bki_lbl.pack(side="left",padx=6)

        # Search
        sf = ttk.Frame(self, style="TFrame")
        sf.pack(fill="x", padx=4, pady=3)
        ttk.Label(sf, text="Sök:").pack(side="left",padx=4)
        self._sok = tk.StringVar()
        self._sok.trace_add("write", lambda *_: self._filter())
        ttk.Entry(sf, textvariable=self._sok, width=30).pack(side="left",padx=2)
        self._count_lbl = ttk.Label(sf, text="", font=("Segoe UI",9),
                                     foreground=C["muted"])
        self._count_lbl.pack(side="left",padx=10)

        tf, self._tree = scrolled_tree(self, PB_COLS, PB_HDR, PB_W, height=24)
        tf.pack(fill="both", expand=True, padx=4, pady=4)
        self._tree.bind("<Double-1>", lambda e: self._till_kalkyl())

    def _load_prisbank(self):
        self._data = load_json(prisbank_path(), [])
        self._filter()

    def _filter(self):
        tree = self._tree
        tree.delete(*tree.get_children())
        q = self._sok.get().lower()
        n = 0
        for i, item in enumerate(self._data):
            if q and q not in item.get("benamning","").lower() \
                    and q not in item.get("kod","").lower(): continue
            tag = "even" if n%2==0 else "odd"
            tree.insert("","end", iid=str(i), tags=(tag,), values=(
                item.get("kod",""), item.get("benamning",""), item.get("enhet",""),
                fmt(item.get("matpris",0)), fmt(item.get("arbpris",0))
            ))
            n += 1
        self._count_lbl.config(text=f"{n} artiklar")

    def _import_excel(self):
        if not PANDAS_OK:
            messagebox.showerror("Saknas","pandas behövs för import. Kör: pip install pandas openpyxl"); return
        fp = filedialog.askopenfilename(
            title="Öppna prislista", filetypes=[("Excel","*.xlsx *.xls"),("Alla","*.*")])
        if not fp: return
        try:
            xl = pd.ExcelFile(fp)
            sheet = xl.sheet_names[0]
            if len(xl.sheet_names)>1:
                sheet = simpledialog.askstring("Flik",
                    f"Välj flik ({', '.join(xl.sheet_names)}):", initialvalue=xl.sheet_names[0])
            df = xl.parse(sheet)
            df.columns = [str(c).strip().lower() for c in df.columns]
            # Try auto-map columns
            colmap = {}
            for col in df.columns:
                if any(x in col for x in ["kod","code","nr"]): colmap["kod"]=col
                elif any(x in col for x in ["benäm","beskr","namn","desc"]): colmap["benamning"]=col
                elif any(x in col for x in ["enh","unit"]): colmap["enhet"]=col
                elif any(x in col for x in ["mat","material","kost"]): colmap["matpris"]=col
                elif any(x in col for x in ["arb","arbete","lön","tim"]): colmap["arbpris"]=col
            if "benamning" not in colmap:
                # fallback: use first columns
                cols = list(df.columns)
                mapping = dict(zip(["kod","benamning","enhet","matpris","arbpris"],cols))
                colmap = {k:v for k,v in mapping.items()}

            ny = []
            for _, row in df.iterrows():
                item = {
                    "kod":      str(row.get(colmap.get("kod",""),"")),
                    "benamning":str(row.get(colmap.get("benamning",""),"")),
                    "enhet":    str(row.get(colmap.get("enhet",""),"st")),
                    "matpris":  sfloat(row.get(colmap.get("matpris",""),0)),
                    "arbpris":  sfloat(row.get(colmap.get("arbpris",""),0)),
                }
                if item["benamning"] and item["benamning"]!="nan": ny.append(item)
            self._data.extend(ny)
            save_json(prisbank_path(), self._data)
            self._filter()
            messagebox.showinfo("Importerat",f"{len(ny)} artiklar importerade.")
        except Exception as e:
            messagebox.showerror("Fel",f"Kunde inte importera:\n{e}")

    def _ny_artikel(self):
        win = tk.Toplevel(self); win.title("Ny artikel"); win.configure(bg=C["bg"])
        win.geometry("380x280"); win.grab_set()
        frm = ttk.Frame(win, style="Panel.TFrame"); frm.pack(fill="both",expand=True,padx=10,pady=10)
        vars_ = {}
        for i,(lbl_,key,_) in enumerate([
            ("Kod","kod",""),("Benämning","benamning",""),("Enhet","enhet","st"),
            ("Materialpris","matpris","0"),("Arbetspris","arbpris","0")
        ]):
            ttk.Label(frm, text=lbl_, background=C["panel"]).grid(row=i,column=0,sticky="w",padx=6,pady=3)
            v = tk.StringVar(value=_); vars_[key]=v
            ttk.Entry(frm, textvariable=v).grid(row=i,column=1,sticky="ew",padx=6)
        frm.columnconfigure(1,weight=1)
        def _spara():
            item = {k:v.get() for k,v in vars_.items()}
            item["matpris"] = sfloat(item["matpris"]); item["arbpris"]=sfloat(item["arbpris"])
            self._data.append(item); save_json(prisbank_path(),self._data)
            self._filter(); win.destroy()
        btn(win,"Spara",_spara,"Accent.TButton").pack(pady=8)

    def _ta_bort(self):
        sel = self._tree.selection()
        if not sel: return
        idx = int(sel[0])
        if messagebox.askyesno("Ta bort","Ta bort vald artikel?"):
            self._data.pop(idx); save_json(prisbank_path(),self._data); self._filter()

    def _till_kalkyl(self):
        sel = self._tree.selection()
        if not sel: return
        if not self.app.projekt:
            messagebox.showwarning("Inget projekt","Öppna eller skapa ett projekt."); return
        idx = int(sel[0])
        pb = self._data[idx]
        r = empty_rad()
        r["id"]        = str(datetime.datetime.now().timestamp())
        r["kod"]       = pb.get("kod","")
        r["benamning"] = pb.get("benamning","")
        r["enhet"]     = pb.get("enhet","st")
        r["apris"]     = sfloat(pb.get("matpris",0))
        r["radtyp"]    = "Material"
        r["mangd"]     = 1.0
        berakna(r)
        self.app.projekt["rader"].append(r)
        self.app.kalkyl_tab.refresh()
        self.app.mark_dirty()
        self.app.notebook.select(self.app.tab_idx["kalkyl"])

    def _bki_raknna(self):
        try:
            typ = self._bki_typ.get()
            bas = self._bki_bas.get()
            nytt = self._bki_år.get()
            i_bas  = BKI[typ][bas]
            i_nytt = BKI[typ][nytt]
            faktor = i_nytt / i_bas
            self._bki_lbl.config(text=f"Faktor: {faktor:.4f}  ({bas}→{nytt})")
        except Exception as e:
            self._bki_lbl.config(text=f"Fel: {e}")
# ──────────────────────────────────────────────────────────────
#  MALLAR TAB
# ──────────────────────────────────────────────────────────────
class MallarTab(ttk.Frame):
    def __init__(self, parent, app):
        super().__init__(parent)
        self.app = app
        self._build()
        self.refresh()

    def _build(self):
        ttk.Label(self, text="  Mallar", style="Header.TLabel").pack(fill="x")
        tb = ttk.Frame(self, style="Panel.TFrame")
        tb.pack(fill="x")
        btn(tb,"✕  Ta bort mall", self._ta_bort,"Danger.TButton").pack(side="left",padx=4,pady=4)
        btn(tb,"✎  Byt namn",    self._byt_namn,"TButton").pack(side="left",padx=2,pady=4)
        btn(tb,"→  Använd",      self._anvand,  "Accent.TButton").pack(side="left",padx=2,pady=4)

        pane = ttk.PanedWindow(self, orient="horizontal")
        pane.pack(fill="both", expand=True, padx=4, pady=4)

        # Mall list
        left = ttk.Frame(pane, style="Panel.TFrame")
        pane.add(left, weight=1)
        ttk.Label(left, text="Sparade mallar", background=C["panel"],
                  font=("Segoe UI",10,"bold")).pack(anchor="w",padx=6,pady=4)
        self._lb = tk.Listbox(left, font=("Segoe UI",10),
                              background=C["panel"], selectmode="single",
                              activestyle="none", relief="flat")
        self._lb.pack(fill="both", expand=True, padx=4, pady=4)
        self._lb.bind("<<ListboxSelect>>", self._on_select)

        # Preview
        right = ttk.Frame(pane, style="Panel.TFrame")
        pane.add(right, weight=3)
        ttk.Label(right, text="Innehåll", background=C["panel"],
                  font=("Segoe UI",10,"bold")).pack(anchor="w",padx=6,pady=4)
        _, self._prev_tree = scrolled_tree(right,
            KALKYL_COLS, KALKYL_HDR, KALKYL_W, height=20)
        _.pack(fill="both", expand=True, padx=4, pady=4)

    def refresh(self):
        self._lb.delete(0,"end")
        mallar = load_json(mallar_path(), [])
        for m in mallar:
            self._lb.insert("end", f"  {m.get('namn','')}  ({len(m.get('rader',[]))} rader)")

    def _on_select(self, e):
        sel = self._lb.curselection()
        if not sel: return
        mallar = load_json(mallar_path(), [])
        if sel[0] >= len(mallar): return
        m = mallar[sel[0]]
        tree = self._prev_tree
        tree.delete(*tree.get_children())
        for i,r in enumerate(m.get("rader",[])):
            tag = "even" if i%2==0 else "odd"
            tree.insert("","end",tags=(tag,),values=(
                r.get("radtyp",""),r.get("kod",""),r.get("benamning",""),
                r.get("byggdel",""),fmt(r.get("mangd",0),1),r.get("enhet",""),
                fmt(r.get("timmar",0),1),fmt(r.get("apris",0)),
                fmt(r.get("kostnad",0)),fmt(r.get("paslag",0),1),
                fmt(r.get("forsaljning",0)),r.get("leverantor","")
            ))

    def _ta_bort(self):
        sel = self._lb.curselection()
        if not sel: return
        mallar = load_json(mallar_path(), [])
        if messagebox.askyesno("Ta bort",f"Ta bort mallen?"):
            mallar.pop(sel[0]); save_json(mallar_path(),mallar); self.refresh()

    def _byt_namn(self):
        sel = self._lb.curselection()
        if not sel: return
        mallar = load_json(mallar_path(), [])
        nytt = simpledialog.askstring("Byt namn","Nytt namn:",
                                      initialvalue=mallar[sel[0]].get("namn",""),
                                      parent=self)
        if nytt:
            mallar[sel[0]]["namn"]=nytt; save_json(mallar_path(),mallar); self.refresh()

    def _anvand(self):
        sel = self._lb.curselection()
        if not sel: return
        if not self.app.projekt:
            messagebox.showwarning("Inget projekt","Öppna ett projekt."); return
        mallar = load_json(mallar_path(), [])
        m = mallar[sel[0]]
        for r in m.get("rader",[]):
            ny = copy.deepcopy(r)
            ny["id"] = str(datetime.datetime.now().timestamp())
            berakna(ny)
            self.app.projekt["rader"].append(ny)
        self.app.kalkyl_tab.refresh()
        self.app.mark_dirty()
        self.app.notebook.select(self.app.tab_idx["kalkyl"])

# ──────────────────────────────────────────────────────────────
#  BYGGDELAR TAB
# ──────────────────────────────────────────────────────────────
BD_COLS = ("byggdel","rader","kostnad","forsaljning","tb","marginal")
BD_HDR  = ("Byggdel","Rader","Kostnad","Försäljning","TB","Marginal %")
BD_W    = (160,60,130,130,130,90)

class ByggdelarTab(ttk.Frame):
    def __init__(self, parent, app):
        super().__init__(parent)
        self.app = app
        self._build()

    def _build(self):
        ttk.Label(self, text="  Byggdelar", style="Header.TLabel").pack(fill="x")

        # Manage building parts
        tb = ttk.Frame(self, style="Panel.TFrame")
        tb.pack(fill="x")
        btn(tb,"＋  Ny byggdel",  self._ny,      "Accent.TButton").pack(side="left",padx=4,pady=4)
        btn(tb,"✕  Ta bort",     self._ta_bort, "Danger.TButton").pack(side="left",padx=2,pady=4)
        btn(tb,"↻  Uppdatera",   self.refresh,  "TButton").pack(side="left",padx=2,pady=4)

        pane = ttk.PanedWindow(self, orient="vertical")
        pane.pack(fill="both", expand=True, padx=4, pady=4)

        # Summary table
        top = ttk.Frame(pane, style="Panel.TFrame")
        pane.add(top, weight=2)
        ttk.Label(top, text="Summering per byggdel", background=C["panel"],
                  font=("Segoe UI",10,"bold")).pack(anchor="w",padx=6,pady=4)
        tf, self._tree = scrolled_tree(top, BD_COLS, BD_HDR, BD_W, height=12)
        tf.pack(fill="both", expand=True, padx=4, pady=4)
        self._tree.bind("<<TreeviewSelect>>", self._on_select)

        # Detail rows for selected byggdel
        bot = ttk.Frame(pane, style="Panel.TFrame")
        pane.add(bot, weight=3)
        self._det_lbl = ttk.Label(bot, text="Rader för vald byggdel",
                                   background=C["panel"], font=("Segoe UI",10,"bold"))
        self._det_lbl.pack(anchor="w",padx=6,pady=4)
        df, self._det_tree = scrolled_tree(bot, KALKYL_COLS, KALKYL_HDR, KALKYL_W, height=12)
        df.pack(fill="both", expand=True, padx=4, pady=4)

    def refresh(self):
        tree = self._tree
        tree.delete(*tree.get_children())
        if not self.app.projekt: return
        bdlar = self.app.projekt.get("byggdelar", list(BYGGDEL_DEF))
        rader = self.app.projekt.get("rader",[])
        # Also catch unassigned
        all_bdl = list(bdlar) + ["(Utan byggdel)"]
        totals = {bd:{"rader":0,"k":0.0,"f":0.0} for bd in all_bdl}
        for r in rader:
            bd = r.get("byggdel","") or "(Utan byggdel)"
            if bd not in totals: totals[bd]={"rader":0,"k":0.0,"f":0.0}
            totals[bd]["rader"] += 1
            totals[bd]["k"] += sfloat(r.get("kostnad",0))
            totals[bd]["f"] += sfloat(r.get("forsaljning",0))
        for i,bd in enumerate(all_bdl):
            t = totals.get(bd,{"rader":0,"k":0.0,"f":0.0})
            tb_  = t["f"]-t["k"]
            mg   = (tb_/t["f"]*100) if t["f"] else 0
            tag  = "even" if i%2==0 else "odd"
            tree.insert("","end",iid=bd,tags=(tag,),values=(
                bd, t["rader"], fmt(t["k"]), fmt(t["f"]), fmt(tb_), fmt(mg,1)
            ))
        # Totals row
        all_k = sum(t["k"] for t in totals.values())
        all_f = sum(t["f"] for t in totals.values())
        tb__ = all_f-all_k; mg__=(tb__/all_f*100) if all_f else 0
        tree.insert("","end",iid="__total__",tags=(),values=(
            "TOTALT", sum(t["rader"] for t in totals.values()),
            fmt(all_k), fmt(all_f), fmt(tb__), fmt(mg__,1)
        ))

    def _on_select(self, e):
        sel = self._tree.selection()
        if not sel or sel[0]=="__total__": return
        bd = sel[0]
        self._det_lbl.config(text=f"Rader – {bd}")
        dt = self._det_tree
        dt.delete(*dt.get_children())
        if not self.app.projekt: return
        for i,r in enumerate(self.app.projekt.get("rader",[])):
            if (r.get("byggdel","") or "(Utan byggdel)") != bd: continue
            tag = "even" if i%2==0 else "odd"
            dt.insert("","end",tags=(tag,),values=(
                r.get("radtyp",""),r.get("kod",""),r.get("benamning",""),
                r.get("byggdel",""),fmt(r.get("mangd",0),1),r.get("enhet",""),
                fmt(r.get("timmar",0),1),fmt(r.get("apris",0)),
                fmt(r.get("kostnad",0)),fmt(r.get("paslag",0),1),
                fmt(r.get("forsaljning",0)),r.get("leverantor","")
            ))

    def _ny(self):
        if not self.app.projekt:
            messagebox.showwarning("Inget projekt","Öppna ett projekt."); return
        namn = simpledialog.askstring("Ny byggdel","Namn på byggdel:",parent=self)
        if namn:
            self.app.projekt["byggdelar"].append(namn)
            self.app.mark_dirty(); self.refresh()

    def _ta_bort(self):
        sel = self._tree.selection()
        if not sel or sel[0]=="__total__": return
        bd = sel[0]
        if messagebox.askyesno("Ta bort",f"Ta bort byggdel '{bd}'?\n(Rader behåller sin byggdelskoppling)"):
            bdlar = self.app.projekt.get("byggdelar",[])
            if bd in bdlar: bdlar.remove(bd)
            self.app.mark_dirty(); self.refresh()
# ──────────────────────────────────────────────────────────────
#  DOKUMENT TAB
# ──────────────────────────────────────────────────────────────
DOK_COLS = ("namn","kategori","tillagd","storlek")
DOK_HDR  = ("Filnamn","Kategori","Tillagd","Storlek")
DOK_W    = (260,120,110,90)
DOK_KAT  = ["Ritningar","Underlag","Inköp","UE","Övrigt"]

class DokumentTab(ttk.Frame):
    def __init__(self, parent, app):
        super().__init__(parent)
        self.app = app
        self._build()

    def _build(self):
        ttk.Label(self, text="  Dokument", style="Header.TLabel").pack(fill="x")
        tb = ttk.Frame(self, style="Panel.TFrame")
        tb.pack(fill="x")
        btn(tb,"＋  Lägg till fil", self._lagg_till,"Accent.TButton").pack(side="left",padx=4,pady=4)
        btn(tb,"🗁  Öppna fil",     self._oppna,    "TButton").pack(side="left",padx=2,pady=4)
        btn(tb,"🗁  Öppna mapp",    self._oppna_mapp,"TButton").pack(side="left",padx=2,pady=4)
        btn(tb,"✕  Ta bort",       self._ta_bort,  "Danger.TButton").pack(side="left",padx=2,pady=4)

        tf, self._tree = scrolled_tree(self, DOK_COLS, DOK_HDR, DOK_W, height=24)
        tf.pack(fill="both", expand=True, padx=4, pady=4)
        self._tree.bind("<Double-1>", lambda e: self._oppna())

    def refresh(self):
        tree = self._tree
        tree.delete(*tree.get_children())
        if not self.app.projekt: return
        for i,d in enumerate(self.app.projekt.get("dokument",[])):
            tag = "even" if i%2==0 else "odd"
            p = Path(d.get("sökväg",""))
            storlek = ""
            try:
                sz = p.stat().st_size
                storlek = f"{sz//1024} kB" if sz>=1024 else f"{sz} B"
            except: storlek = "?"
            tree.insert("","end",iid=str(i),tags=(tag,),values=(
                d.get("namn",""), d.get("kategori",""),
                d.get("tillagd",""), storlek
            ))

    def _lagg_till(self):
        if not self.app.projekt:
            messagebox.showwarning("Inget projekt","Öppna ett projekt."); return
        filer = filedialog.askopenfilenames(title="Välj filer")
        if not filer: return
        # Ask category
        kat = simpledialog.askstring("Kategori",
            f"Kategori ({', '.join(DOK_KAT)}):", initialvalue="Övrigt", parent=self)
        mapp = self.app.projekt.get("projektmapp","")
        for fp in filer:
            dest = fp
            if mapp:
                # copy to project folder
                sub = kat if kat in MAPPAR else "Övrigt"
                dest_dir = Path(mapp) / sub
                dest_dir.mkdir(parents=True, exist_ok=True)
                dest = str(dest_dir / Path(fp).name)
                try: shutil.copy2(fp, dest)
                except Exception as e: dest = fp
            doc = {"namn":Path(fp).name, "kategori":kat or "Övrigt",
                   "tillagd":datetime.date.today().isoformat(), "sökväg":dest}
            self.app.projekt["dokument"].append(doc)
        self.refresh()
        self.app.mark_dirty()

    def _oppna(self):
        sel = self._tree.selection()
        if not sel: return
        idx = int(sel[0])
        dok = self.app.projekt["dokument"][idx]
        open_file(dok.get("sökväg",""))

    def _oppna_mapp(self):
        if not self.app.projekt: return
        mapp = self.app.projekt.get("projektmapp","")
        if mapp and Path(mapp).exists(): open_file(mapp)
        else: messagebox.showinfo("Ingen mapp","Ingen projektmapp inställd.")

    def _ta_bort(self):
        sel = self._tree.selection()
        if not sel: return
        idx = int(sel[0])
        if messagebox.askyesno("Ta bort","Ta bort dokumentreferensen?\n(Filen raderas ej.)"):
            self.app.projekt["dokument"].pop(idx)
            self.refresh(); self.app.mark_dirty()

# ──────────────────────────────────────────────────────────────
#  SLUTSIDA TAB
# ──────────────────────────────────────────────────────────────
class SlutsidaTab(ttk.Frame):
    def __init__(self, parent, app):
        super().__init__(parent)
        self.app = app
        self._omk_vars = {}
        self._pas_vars = {}
        self._build()

    def _build(self):
        ttk.Label(self, text="  Slutsida – Ekonomi", style="Header.TLabel").pack(fill="x")

        canvas = tk.Canvas(self, background=C["bg"], highlightthickness=0)
        sb = ttk.Scrollbar(self, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=sb.set)
        sb.pack(side="right",fill="y"); canvas.pack(fill="both",expand=True)
        inner = ttk.Frame(canvas, style="TFrame")
        wid = canvas.create_window((0,0),window=inner,anchor="nw")
        def _cfg(e): canvas.configure(scrollregion=canvas.bbox("all")); canvas.itemconfig(wid,width=e.width)
        canvas.bind("<Configure>",_cfg)
        inner.bind("<Configure>",lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        inner.columnconfigure(0,weight=1); inner.columnconfigure(1,weight=1)

        # Col 1 – Omkostnader
        col1 = ttk.Frame(inner, style="Panel.TFrame"); col1.grid(row=0,column=0,padx=10,pady=10,sticky="nsew")
        ttk.Label(col1, text="  Omkostnader", style="Header.TLabel",
                  font=("Segoe UI",10,"bold")).pack(fill="x",pady=(0,4))
        omk_keys = ["Arbetsledning","Etablering","Administration","Transporter","Garanti","Risk","Övrigt"]
        for k in omk_keys:
            f = ttk.Frame(col1, style="Panel.TFrame"); f.pack(fill="x",padx=8,pady=2)
            ttk.Label(f, text=k, background=C["panel"], width=18).pack(side="left")
            v = tk.StringVar(value="0"); self._omk_vars[k]=v
            ttk.Entry(f, textvariable=v, width=16).pack(side="left")
            ttk.Label(f, text="kr", background=C["panel"]).pack(side="left",padx=4)
            v.trace_add("write", lambda *_: self._update())

        ttk.Frame(col1,height=10,style="Panel.TFrame").pack()
        ttk.Label(col1, text="  Påslag", style="Header.TLabel",
                  font=("Segoe UI",10,"bold")).pack(fill="x",pady=(0,4))
        pas_keys = ["Omkostnad %","Risk %","Vinst %","Rabatt %"]
        for k in pas_keys:
            f = ttk.Frame(col1, style="Panel.TFrame"); f.pack(fill="x",padx=8,pady=2)
            ttk.Label(f, text=k, background=C["panel"], width=18).pack(side="left")
            v = tk.StringVar(value="0"); self._pas_vars[k]=v
            ttk.Entry(f, textvariable=v, width=10).pack(side="left")
            v.trace_add("write", lambda *_: self._update())

        btn(col1,"↻  Beräkna", self._update, "Accent.TButton").pack(padx=8,pady=10,anchor="w")

        # Col 2 – Resultat
        col2 = ttk.Frame(inner, style="Panel.TFrame"); col2.grid(row=0,column=1,padx=10,pady=10,sticky="nsew")
        ttk.Label(col2, text="  Resultat", style="Header.TLabel",
                  font=("Segoe UI",10,"bold")).pack(fill="x",pady=(0,4))
        self._res_rows = {}
        res_items = [
            ("dir_k",   "Direkta kostnader",    False),
            ("omk_s",   "Omkostnader (fasta)",  False),
            ("omk_b",   "Omkostnadspåslag",     False),
            ("ris_b",   "Riskpåslag",           False),
            ("sjk",     "Självkostnad",         True),
            ("vin_b",   "Vinst",                False),
            ("fp",      "Försäljningspris",     True),
            ("tb",      "Täckningsbidrag (TB)", True),
            ("mg",      "Marginal %",           True),
        ]
        for key,lbl_,bold in res_items:
            f = ttk.Frame(col2, style="Panel.TFrame"); f.pack(fill="x",padx=8,pady=3)
            font = ("Segoe UI",10,"bold") if bold else ("Segoe UI",10)
            fg   = C["primary"] if bold else C["text"]
            ttk.Label(f, text=lbl_, background=C["panel"],
                      width=26, font=font, foreground=fg).pack(side="left")
            v = ttk.Label(f, text="–", background=C["panel"],
                          font=font, foreground=fg, width=18, anchor="e")
            v.pack(side="left")
            self._res_rows[key] = v
            if bold: sep(col2).pack(fill="x",padx=8,pady=1)

        btn(col2,"📄  Exportera slutsida PDF", self.app.export_pdf_slutsida,
            "TButton").pack(padx=8,pady=6,anchor="w")
        btn(col2,"📊  Exportera Excel",        self.app.export_excel_slutsida,
            "TButton").pack(padx=8,pady=4,anchor="w")

    def load(self, proj):
        omk = proj.get("omkostnader",{})
        for k,v in self._omk_vars.items():
            v.set(str(omk.get(k,0)))
        pas = proj.get("paslag",{})
        for k,v in self._pas_vars.items():
            v.set(str(pas.get(k,0)))
        self._update()

    def save_to(self, proj):
        proj["omkostnader"] = {k: sfloat(v.get()) for k,v in self._omk_vars.items()}
        proj["paslag"]      = {k: sfloat(v.get()) for k,v in self._pas_vars.items()}

    def _update(self):
        if not self.app.projekt: return
        self.save_to(self.app.projekt)
        s = summera_projekt(self.app.projekt)
        for key, widget in self._res_rows.items():
            v = s.get(key, 0)
            if key=="mg":
                widget.config(text=f"{fmt(v,2)} %")
            else:
                widget.config(text=f"{fmt(v)} kr")
        self.app.mark_dirty()
# ──────────────────────────────────────────────────────────────
#  EXPORT  (Excel + PDF)
# ──────────────────────────────────────────────────────────────
def _wb_style(ws, header_row=1):
    """Apply basic styling to an openpyxl worksheet."""
    hdr_fill = PatternFill("solid", fgColor="1A3A5C")
    hdr_font = Font(color="FFFFFF", bold=True, name="Calibri", size=10)
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin")
    )
    for cell in ws[header_row]:
        cell.fill = hdr_fill; cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
        cell.border = thin
    for row in ws.iter_rows(min_row=header_row+1):
        for i,cell in enumerate(row):
            cell.border = thin
            if (cell.row - header_row) % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="EEF3F8")
    ws.freeze_panes = ws.cell(row=header_row+1, column=1)

def export_excel_kalkyl(proj, filepath):
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Kalkyl"
    hdrs = ["Typ","Kod","Benämning","Beskrivning","Byggdel","Mängd","Enh",
            "Timmar","Á-pris","Kostnad","Påslag%","Försäljning","Leverantör","Kommentar"]
    ws.append(hdrs)
    for r in proj.get("rader",[]):
        ws.append([r.get("radtyp",""),r.get("kod",""),r.get("benamning",""),
                   r.get("beskrivning",""),r.get("byggdel",""),
                   r.get("mangd",0),r.get("enhet",""),r.get("timmar",0),
                   r.get("apris",0),r.get("kostnad",0),r.get("paslag",0),
                   r.get("forsaljning",0),r.get("leverantor",""),r.get("kommentar","")])
    _wb_style(ws)
    for col, w in zip("ABCDEFGHIJKLMN",[8,8,30,20,14,8,6,8,12,12,8,14,16,20]):
        ws.column_dimensions[col].width = w
    # Totals
    n = len(proj.get("rader",[]))+2
    ws.cell(n,1,"TOTALT").font=Font(bold=True)
    ws.cell(n,10,f"=SUM(J2:J{n-1})").font=Font(bold=True)
    ws.cell(n,12,f"=SUM(L2:L{n-1})").font=Font(bold=True)

    # Byggdelar sheet
    ws2 = wb.create_sheet("Byggdelar")
    ws2.append(["Byggdel","Rader","Kostnad","Försäljning","TB","Marginal %"])
    bdlar = proj.get("byggdelar",list(BYGGDEL_DEF))
    rader = proj.get("rader",[])
    totals = {}
    for r in rader:
        bd = r.get("byggdel","") or "(Utan byggdel)"
        if bd not in totals: totals[bd]={"n":0,"k":0.0,"f":0.0}
        totals[bd]["n"]+=1; totals[bd]["k"]+=sfloat(r.get("kostnad",0))
        totals[bd]["f"]+=sfloat(r.get("forsaljning",0))
    for bd in bdlar:
        t=totals.get(bd,{"n":0,"k":0.0,"f":0.0})
        tb_=t["f"]-t["k"]; mg=(tb_/t["f"]*100) if t["f"] else 0
        ws2.append([bd,t["n"],t["k"],t["f"],tb_,round(mg,2)])
    _wb_style(ws2)

    wb.save(filepath)

def export_excel_slutsida(proj, filepath):
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Slutsida"
    s = summera_projekt(proj)
    pname = proj.get("projektnamn",""); pnr = proj.get("projektnummer","")
    ws["A1"] = f"Projekt: {pname}  ({pnr})"
    ws["A1"].font = Font(bold=True,size=13)
    ws.append([]); ws.append(["Post","Belopp (kr)"])
    _wb_style(ws,3)
    rows = [
        ("Direkta kostnader", s["dir_k"]),
        ("Omkostnader (fasta)", s["omk_s"]),
        ("Omkostnadspåslag", s["omk_b"]),
        ("Riskpåslag", s["ris_b"]),
        ("Självkostnad", s["sjk"]),
        ("Vinst", s["vin_b"]),
        ("Försäljningspris", s["fp"]),
        ("Täckningsbidrag (TB)", s["tb"]),
        ("Marginal %", s["mg"]),
    ]
    for post, bel in rows:
        ws.append([post, round(bel,2) if post!="Marginal %" else round(bel,2)])
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 18
    wb.save(filepath)

def export_pdf_kalkyl(proj, filepath):
    doc = SimpleDocTemplate(filepath, pagesize=A4,
                            leftMargin=15*mm, rightMargin=15*mm,
                            topMargin=20*mm, bottomMargin=20*mm)
    styles = getSampleStyleSheet()
    title_style = styles["Title"]; title_style.fontName="Helvetica-Bold"; title_style.fontSize=14
    normal = styles["Normal"]; normal.fontName="Helvetica"; normal.fontSize=8
    elements = []
    pname = proj.get("projektnamn","")
    pnr   = proj.get("projektnummer","")
    elements.append(Paragraph(f"Kalkyl – {pname} ({pnr})", title_style))
    elements.append(Paragraph(
        f"Datum: {proj.get('datum','')}  |  Kalkylansvarig: {proj.get('kalkylansvarig','')}",
        normal))
    elements.append(Spacer(1,6*mm))
    hdrs = [["Typ","Kod","Benämning","Byggdel","Mängd","Enh","Tim","Á-pris","Kostnad","Pål%","Försälj."]]
    rows = hdrs[:]
    for r in proj.get("rader",[]):
        rows.append([
            r.get("radtyp","")[:3], r.get("kod","")[:8],
            r.get("benamning","")[:30], r.get("byggdel","")[:14],
            fmt(r.get("mangd",0),1), r.get("enhet",""),
            fmt(r.get("timmar",0),1), fmt(r.get("apris",0)),
            fmt(r.get("kostnad",0)), fmt(r.get("paslag",0),1),
            fmt(r.get("forsaljning",0))
        ])
    col_w = [28,28,90,55,28,22,22,38,42,24,42]
    col_w = [w*mm for w in col_w]
    t = Table(rows, colWidths=col_w, repeatRows=1)
    ts = TableStyle([
        ("BACKGROUND",(0,0),(-1,0),rl_colors.HexColor("#1A3A5C")),
        ("TEXTCOLOR",(0,0),(-1,0),rl_colors.white),
        ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
        ("FONTSIZE",(0,0),(-1,-1),7),
        ("ROWBACKGROUNDS",(0,1),(-1,-1),[rl_colors.HexColor("#EEF3F8"),rl_colors.white]),
        ("GRID",(0,0),(-1,-1),0.3,rl_colors.grey),
        ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
    ])
    t.setStyle(ts)
    elements.append(t)
    elements.append(Spacer(1,6*mm))
    s = summera_projekt(proj)
    elements.append(Paragraph(
        f"<b>Direkta kostnader:</b> {fmt(s['dir_k'])} kr  |  "
        f"<b>Försäljning:</b> {fmt(s['fp'])} kr  |  "
        f"<b>TB:</b> {fmt(s['tb'])} kr  |  "
        f"<b>Marginal:</b> {fmt(s['mg'],1)} %", normal))
    doc.build(elements)

def export_pdf_slutsida(proj, filepath):
    doc = SimpleDocTemplate(filepath, pagesize=A4,
                            leftMargin=20*mm, rightMargin=20*mm,
                            topMargin=25*mm, bottomMargin=20*mm)
    styles = getSampleStyleSheet()
    title_s = styles["Title"]; title_s.fontName="Helvetica-Bold"; title_s.fontSize=16
    h2 = styles["Heading2"]; h2.fontName="Helvetica-Bold"
    normal = styles["Normal"]; normal.fontName="Helvetica"; normal.fontSize=10
    elements = []
    pname = proj.get("projektnamn","Projekt")
    pnr   = proj.get("projektnummer","")
    kund  = proj.get("kund","")
    elements.append(Paragraph(f"Kalkyl – {pname}", title_s))
    elements.append(Paragraph(
        f"Projektnr: {pnr}  |  Kund: {kund}  |  Datum: {proj.get('datum','')}  "
        f"|  Ansvarig: {proj.get('kalkylansvarig','')}", normal))
    elements.append(Spacer(1,8*mm))
    s = summera_projekt(proj)
    data = [
        ["Post","Belopp"],
        ["Direkta kostnader",       f"{fmt(s['dir_k'])} kr"],
        ["Omkostnader (fasta)",     f"{fmt(s['omk_s'])} kr"],
        ["Omkostnadspåslag",        f"{fmt(s['omk_b'])} kr"],
        ["Riskpåslag",              f"{fmt(s['ris_b'])} kr"],
        ["Självkostnad",            f"{fmt(s['sjk'])} kr"],
        ["Vinst",                   f"{fmt(s['vin_b'])} kr"],
        ["Försäljningspris (netto)",f"{fmt(s['fp'])} kr"],
        ["Täckningsbidrag (TB)",    f"{fmt(s['tb'])} kr"],
        ["Marginal",                f"{fmt(s['mg'],2)} %"],
    ]
    t = Table(data, colWidths=[120*mm, 60*mm])
    ts = TableStyle([
        ("BACKGROUND",(0,0),(-1,0),rl_colors.HexColor("#1A3A5C")),
        ("TEXTCOLOR",(0,0),(-1,0),rl_colors.white),
        ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
        ("FONTSIZE",(0,0),(-1,-1),11),
        ("ROWBACKGROUNDS",(0,1),(-1,-1),[rl_colors.HexColor("#EEF3F8"),rl_colors.white]),
        ("FONTNAME",(0,-1),(-1,-1),"Helvetica-Bold"),
        ("FONTNAME",(0,5),(-1,5),"Helvetica-Bold"),
        ("FONTNAME",(0,7),(-1,7),"Helvetica-Bold"),
        ("GRID",(0,0),(-1,-1),0.5,rl_colors.grey),
        ("TOPPADDING",(0,0),(-1,-1),5),("BOTTOMPADDING",(0,0),(-1,-1),5),
    ])
    t.setStyle(ts)
    elements.append(t)
    doc.build(elements)
# ──────────────────────────────────────────────────────────────
#  MAIN APP
# ──────────────────────────────────────────────────────────────
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title(APP_TITLE)
        self.geometry("1280x800")
        self.minsize(900,600)
        self.configure(bg=C["bg"])
        apply_style(self)
        try:
            self.iconbitmap(default="")
        except: pass

        self.projekt    = None
        self._filepath  = None
        self._dirty     = False

        self._build_menu()
        self._build_ui()
        self.protocol("WM_DELETE_WINDOW", self._on_close)

    def _build_menu(self):
        m = tk.Menu(self, bg=C["primary"], fg="white",
                    activebackground=C["accent"], activeforeground=C["text"])
        self.configure(menu=m)

        fm = tk.Menu(m, tearoff=0)
        m.add_cascade(label="Arkiv", menu=fm)
        fm.add_command(label="Nytt projekt",       command=self.new_projekt,   accelerator="Ctrl+N")
        fm.add_command(label="Öppna projekt...",   command=self.open_projekt,  accelerator="Ctrl+O")
        fm.add_command(label="Spara",              command=self.save_projekt,  accelerator="Ctrl+S")
        fm.add_command(label="Spara som...",       command=self.save_projekt_as)
        fm.add_separator()
        fm.add_command(label="Avsluta",            command=self._on_close)

        em = tk.Menu(m, tearoff=0)
        m.add_cascade(label="Export", menu=em)
        em.add_command(label="Kalkyl → Excel",     command=self.export_excel_kalkyl)
        em.add_command(label="Slutsida → Excel",   command=self.export_excel_slutsida)
        em.add_command(label="Kalkyl → PDF",       command=self.export_pdf_kalkyl)
        em.add_command(label="Slutsida → PDF",     command=self.export_pdf_slutsida)

        self.bind_all("<Control-n>", lambda e: self.new_projekt())
        self.bind_all("<Control-o>", lambda e: self.open_projekt())
        self.bind_all("<Control-s>", lambda e: self.save_projekt())

    def _build_ui(self):
        # Top status bar
        self._status_bar = ttk.Label(self, text="Inget projekt öppet",
                                      background=C["primary"], foreground=C["hdr_fg"],
                                      font=("Segoe UI",9), anchor="w", padding=[8,3])
        self._status_bar.pack(side="bottom", fill="x")

        self.notebook = ttk.Notebook(self)
        self.notebook.pack(fill="both", expand=True)

        self.start_tab    = StartTab(self.notebook, self)
        self.projekt_tab  = ProjektTab(self.notebook, self)
        self.kalkyl_tab   = KalkylTab(self.notebook, self)
        self.prisbank_tab = PrisbankTab(self.notebook, self)
        self.mallar_tab   = MallarTab(self.notebook, self)
        self.byggdelar_tab= ByggdelarTab(self.notebook, self)
        self.dokument_tab = DokumentTab(self.notebook, self)
        self.slutsida_tab = SlutsidaTab(self.notebook, self)

        tabs = [
            ("🏠  Start",       self.start_tab),
            ("📋  Projekt",     self.projekt_tab),
            ("🔢  Kalkyl",      self.kalkyl_tab),
            ("💰  Prisbank",    self.prisbank_tab),
            ("📑  Mallar",      self.mallar_tab),
            ("🏗  Byggdelar",   self.byggdelar_tab),
            ("📁  Dokument",    self.dokument_tab),
            ("📊  Slutsida",    self.slutsida_tab),
        ]
        self.tab_idx = {}
        for name, frame in tabs:
            self.notebook.add(frame, text=name)
        for i, (name, _) in enumerate(tabs):
            key = name.split()[-1].lower().rstrip(":")
            self.tab_idx[key] = i
        # Friendlier key mapping
        self.tab_idx["kalkyl"]    = 2
        self.tab_idx["prisbank"]  = 3
        self.tab_idx["mallar"]    = 4
        self.tab_idx["byggdelar"] = 5
        self.tab_idx["dokument"]  = 6
        self.tab_idx["slutsida"]  = 7

        self.notebook.bind("<<NotebookTabChanged>>", self._on_tab_change)

    def _on_tab_change(self, e):
        idx = self.notebook.index("current")
        if idx==2: self.kalkyl_tab.refresh()
        elif idx==5: self.byggdelar_tab.refresh()
        elif idx==6: self.dokument_tab.refresh()

    def mark_dirty(self):
        self._dirty = True
        name = self.projekt.get("projektnamn","") if self.projekt else ""
        self.title(f"* {APP_TITLE} – {name}")

    def _set_status(self, msg):
        self._status_bar.config(text=f"  {msg}")

    def _load_projekt(self, proj, filepath):
        self.projekt   = proj
        self._filepath = filepath
        self._dirty    = False
        name = proj.get("projektnamn","")
        self.title(f"{APP_TITLE} – {name}")
        self._set_status(f"Öppnat: {filepath}")
        self.projekt_tab.load(proj)
        self.slutsida_tab.load(proj)
        self.kalkyl_tab.refresh()
        self.byggdelar_tab.refresh()
        self.dokument_tab.refresh()
        self.start_tab.update_info(proj)
        self._add_recent(filepath)

    def _add_recent(self, filepath):
        s = load_json(settings_path(), {"recent_projects":[]})
        rp = s.get("recent_projects",[])
        if filepath in rp: rp.remove(filepath)
        rp.insert(0, filepath)
        s["recent_projects"] = rp[:10]
        save_json(settings_path(), s)
        self.start_tab.refresh_recent()

    # ── File operations ──────────────────────────────────────
    def new_projekt(self):
        if self._dirty and not self._confirm_discard(): return
        proj = empty_projekt()
        self._load_projekt(proj, None)
        self.notebook.select(1)  # go to Projekt tab

    def open_projekt(self, filepath=None):
        if self._dirty and not self._confirm_discard(): return
        if not filepath:
            filepath = filedialog.askopenfilename(
                title="Öppna projekt",
                filetypes=[("Kalkylprojekt","*.json"),("Alla","*.*")])
        if not filepath: return
        try:
            with open(filepath,"r",encoding="utf-8") as f:
                proj = json.load(f)
            self._load_projekt(proj, filepath)
        except Exception as e:
            messagebox.showerror("Fel",f"Kunde inte öppna filen:\n{e}")

    def save_projekt(self):
        if not self.projekt:
            messagebox.showinfo("Inget projekt","Skapa ett projekt först."); return
        self.projekt_tab.save_to(self.projekt)
        self.slutsida_tab.save_to(self.projekt)
        if not self._filepath:
            self.save_projekt_as(); return
        try:
            with open(self._filepath,"w",encoding="utf-8") as f:
                json.dump(self.projekt,f,ensure_ascii=False,indent=2)
            self._dirty = False
            name = self.projekt.get("projektnamn","")
            self.title(f"{APP_TITLE} – {name}")
            self._set_status(f"Sparad: {self._filepath}")
            self.start_tab.update_info(self.projekt)
        except Exception as e:
            messagebox.showerror("Fel",f"Kunde inte spara:\n{e}")

    def save_projekt_as(self):
        if not self.projekt: return
        self.projekt_tab.save_to(self.projekt)
        name = self.projekt.get("projektnamn","projekt").replace(" ","_")
        fp = filedialog.asksaveasfilename(
            title="Spara projekt som",
            defaultextension=".json",
            initialfile=f"{name}.json",
            filetypes=[("Kalkylprojekt","*.json")])
        if not fp: return
        self._filepath = fp
        self.save_projekt()

    def _confirm_discard(self):
        return messagebox.askyesno("Osparade ändringar",
                                    "Du har osparade ändringar. Vill du fortsätta ändå?")

    def _on_close(self):
        if self._dirty:
            ans = messagebox.askyesnocancel("Spara?","Vill du spara innan du stänger?")
            if ans is None: return
            if ans: self.save_projekt()
        self.destroy()

    # ── Export ───────────────────────────────────────────────
    def _export_path(self, default_name, ext, label):
        mapp = ""
        if self.projekt:
            pm = self.projekt.get("projektmapp","")
            if pm: mapp = str(Path(pm)/"Export")
        fp = filedialog.asksaveasfilename(
            title=f"Spara {label}",
            initialdir=mapp or str(Path.home()),
            initialfile=default_name,
            defaultextension=ext,
            filetypes=[(label,f"*{ext}"),("Alla","*.*")])
        return fp

    def export_excel_kalkyl(self):
        if not self.projekt: messagebox.showwarning("Inget projekt","Öppna ett projekt."); return
        if not OPX_OK: messagebox.showerror("Saknas","openpyxl behövs. Kör: pip install openpyxl"); return
        name = self.projekt.get("projektnamn","kalkyl").replace(" ","_")
        fp = self._export_path(f"{name}_kalkyl.xlsx",".xlsx","Excel-kalkyl")
        if not fp: return
        try:
            export_excel_kalkyl(self.projekt, fp)
            self._set_status(f"Excel kalkyl exporterad: {fp}")
            if messagebox.askyesno("Exporterat","Öppna filen?"): open_file(fp)
        except Exception as e: messagebox.showerror("Fel",str(e))

    def export_excel_slutsida(self):
        if not self.projekt: messagebox.showwarning("Inget projekt","Öppna ett projekt."); return
        if not OPX_OK: messagebox.showerror("Saknas","openpyxl behövs."); return
        self.slutsida_tab.save_to(self.projekt)
        name = self.projekt.get("projektnamn","slutsida").replace(" ","_")
        fp = self._export_path(f"{name}_slutsida.xlsx",".xlsx","Excel-slutsida")
        if not fp: return
        try:
            export_excel_slutsida(self.projekt, fp)
            self._set_status(f"Excel slutsida exporterad: {fp}")
            if messagebox.askyesno("Exporterat","Öppna filen?"): open_file(fp)
        except Exception as e: messagebox.showerror("Fel",str(e))

    def export_pdf_kalkyl(self):
        if not self.projekt: messagebox.showwarning("Inget projekt","Öppna ett projekt."); return
        if not RL_OK: messagebox.showerror("Saknas","reportlab behövs. Kör: pip install reportlab"); return
        name = self.projekt.get("projektnamn","kalkyl").replace(" ","_")
        fp = self._export_path(f"{name}_kalkyl.pdf",".pdf","PDF-kalkyl")
        if not fp: return
        try:
            export_pdf_kalkyl(self.projekt, fp)
            self._set_status(f"PDF exporterad: {fp}")
            if messagebox.askyesno("Exporterat","Öppna filen?"): open_file(fp)
        except Exception as e: messagebox.showerror("Fel",str(e))

    def export_pdf_slutsida(self):
        if not self.projekt: messagebox.showwarning("Inget projekt","Öppna ett projekt."); return
        if not RL_OK: messagebox.showerror("Saknas","reportlab behövs."); return
        self.slutsida_tab.save_to(self.projekt)
        name = self.projekt.get("projektnamn","slutsida").replace(" ","_")
        fp = self._export_path(f"{name}_slutsida.pdf",".pdf","PDF-slutsida")
        if not fp: return
        try:
            export_pdf_slutsida(self.projekt, fp)
            self._set_status(f"PDF exporterad: {fp}")
            if messagebox.askyesno("Exporterat","Öppna filen?"): open_file(fp)
        except Exception as e: messagebox.showerror("Fel",str(e))


# ──────────────────────────────────────────────────────────────
#  ENTRY POINT
# ──────────────────────────────────────────────────────────────
def main():
    app = App()
    app.mainloop()

if __name__ == "__main__":
    main()
